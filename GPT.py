import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import io
import openpyxl
import csv
import os
import re
import shutil
from pathlib import Path
from collections import defaultdict
import time
import warnings
warnings.filterwarnings('ignore')

# ⬇️⬇️⬇️ PASTE DI SINI ⬇️⬇️⬇️
import streamlit as st
import streamlit.components.v1 as components


def notifikasi_sukses():
    # 1. Tampilkan pop-up toast berhasil
    st.toast("✅ Upload Berhasil!", icon="🎉")

    # 2. Trik HTML + JS bypass kebijakan browser menggunakan Web Audio API
    components.html(
        """
    <button id="playBtn" style="display:none;"></button>
    <script>
        function bunyikanBip() {
            try {
                var ctx = new (window.AudioContext || window.webkitAudioContext)();
                var osc = ctx.createOscillator();
                var gain = ctx.createGain();
                osc.connect(gain); 
                gain.connect(ctx.destination);
                osc.frequency.value = 880; 
                gain.gain.setValueAtTime(0.2, ctx.currentTime); 
                osc.start();
                gain.gain.exponentialRampToValueAtTime(0.001, ctx.currentTime + 0.3);
                osc.stop(ctx.currentTime + 0.3);
            } catch(e) { console.log(e); }
        }

        // Jalankan trik klik otomatis setelah elemen siap
        setTimeout(() => {
            const btn = document.getElementById('playBtn');
            btn.addEventListener('click', bunyikanBip);
            btn.click(); // Memaksa browser mengizinkan audio lewat aksi klik
        }, 100);
    </script>
    """,
        height=0,
    )


# ⬆️⬆️⬆️ SAMPAI SINI ⬆️⬆️⬆️

# =====================================
# PAGE CONFIG
# =====================================
st.set_page_config(
    page_title="BSELLER ANALYTICS PRO",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =====================================
# FILE DATABASE
# =====================================
BASE_DIR = Path(__file__).parent
DB_FOLDER = BASE_DIR / "bseller_data"
DB_FOLDER.mkdir(exist_ok=True)

REKAP_DB_FILE = str(DB_FOLDER / "rekap_database.csv")
SHOPEE_DB_FILE = str(DB_FOLDER / "shopee_database.csv")
IKLAN_DB_FILE = str(DB_FOLDER / "iklan_database.csv")
DANA_DB_FILE = str(DB_FOLDER / "dana_database.csv")
HPP_DB_FILE = str(DB_FOLDER / "hpp_database.csv")
PEMBELIAN_DB_FILE = str(DB_FOLDER / "pembelian_database.csv")
MASTER_PRODUK_DB_FILE = str(DB_FOLDER / "master_produk.csv")
PENJUALAN_FIFO_DB_FILE = str(DB_FOLDER / "penjualan_fifo.csv")
REKAP_IKLAN_MASTER_FILE = str(DB_FOLDER / "rekap_iklan_master.csv")
INCOME_DB_FILE = str(DB_FOLDER / "income_database.csv")
STOK_BATCH_DB_FILE = str(DB_FOLDER / "stok_batch.csv")
RIWAYAT_ROAS_FILE = str(DB_FOLDER / "riwayat_roas.csv")

# =====================================
# FUNGSI FORMAT ANGKA
# =====================================
def format_angka(x):
    try:
        if x is None or pd.isna(x):
            return "0"
        if isinstance(x, str):
            x = x.replace('.', '').replace(',', '.')
        num = float(x)
        return f"{num:,.0f}".replace(',', '.')
    except (ValueError, TypeError):
        return "0"

def format_rupiah(angka):
    try:
        return f"Rp {format_angka(angka)}"
    except:
        return "Rp 0"

def format_uang(x):
    """Format angka ke format Indonesia: 1.000, 10.000, 100.000, 1.000.000"""
    try:
        if pd.isna(x) or x == '' or x is None:
            return "0"
        if isinstance(x, str):
            x = x.replace('.', '').replace(',', '').replace('Rp', '').replace(' ', '')
        num = int(float(x))
        return f"{num:,}".replace(',', '.')
    except:
        return "0"

def remove_duplicate_columns(df):
    cols = pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique():
        dup_indices = [i for i, col in enumerate(df.columns) if col == dup]
        if len(dup_indices) > 1:
            df = df.drop(df.columns[dup_indices[1:]], axis=1)
    return df

def clean_column_names(df):
    df.columns = [str(col).strip().replace('\n', ' ').replace('\r', '') for col in df.columns]
    return df

def extract_nama_toko(filename):
    nama = filename.replace('.xlsx','').replace('.xls','').replace('.csv','')
    nama = re.sub(r'\(\d+\)$', '', nama)
    nama = re.sub(r'_\d+$', '', nama)
    nama = re.sub(r'\s+\d+$', '', nama)  # ⬅️ TAMBAH INI: hapus spasi + angka di akhir
    return nama.strip()

# =====================================
# FUNGSI WARNA LABA (TAMBAHKAN INI)
# =====================================
def warna_laba(val):
    """
    Format laba/rugi dengan warna emoji
    - Positif: 🟢 +Rp 1.000
    - Negatif: 🔴 -Rp 1.000  
    - Nol: ⚪ Rp 0
    """
    try:
        if pd.isna(val) or val == 0:
            return "⚪ Rp 0"
        elif val > 0:
            return f"🟢 Rp {format_angka(val)}"
        else:
            return f"🔴 Rp {format_angka(abs(val))}"
    except:
        return "⚪ Rp 0"
# =====================================
# PROCESS INCOME FILE (PEMBAYARAN SHOPEE) - FIXED
# =====================================
def process_income_file(file_bytes, filename):
    """
    Proses file Income Statement Shopee (.xlsx)
    Mengambil sheet 'Income' untuk detail transaksi
    """
    try:
        # Buka Excel
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        # Cek sheet yang tersedia
        available_sheets = wb.sheetnames
        
        # Prioritaskan sheet 'Income'
        if 'Income' in available_sheets:
            ws = wb['Income']
        elif 'Sheet1' in available_sheets:
            ws = wb['Sheet1']
        else:
            ws = wb[available_sheets[0]]
        
        # Cari baris header (yang mengandung "No. Pesanan")
        header_row = None
        header_cols = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell and 'no. pesanan' in str(cell).lower():
                    header_row = row_idx
                    # Mapping header
                    for c_idx, c_val in enumerate(row, start=1):
                        if c_val:
                            header_cols[str(c_val).strip()] = c_idx
                    break
            if header_row:
                break
        
        if not header_row:
            return pd.DataFrame(), "Header tidak ditemukan"
        
        # Baca data
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Stop jika row kosong semua
            if all(cell is None for cell in row):
                break
            
            # Konversi ke dict
            row_dict = {}
            for col_name, col_idx in header_cols.items():
                if col_idx <= len(row):
                    val = row[col_idx - 1]
                    row_dict[col_name] = val
            
            # Hanya ambil jika ada No Pesanan
            no_pesanan = None
            for key in row_dict:
                if 'no. pesanan' in key.lower():
                    no_pesanan = row_dict[key]
                    break
            
            if no_pesanan and str(no_pesanan).strip():
                data_rows.append(row_dict)
        
        if not data_rows:
            return pd.DataFrame(), "Tidak ada data transaksi"
        
        # Buat DataFrame
        df = pd.DataFrame(data_rows)
        
        # ============================================================
        # PERBAIKAN: HAPUS DUPLIKAT KOLOM SEBELUM RENAME
        # ============================================================
        # Hapus kolom duplikat (kolom dengan nama sama)
        df = df.loc[:, ~df.columns.duplicated()]
        
        # ============================================================
        # RENAME KOLOM (HANYA KOLOM YANG ADA)
        # ============================================================
        rename_map = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            
            if 'no. pesanan' in col_lower:
                rename_map[col] = 'No_Pesanan'
            elif 'total penghasilan' in col_lower or 'total yang dilepas' in col_lower:
                rename_map[col] = 'Total_Dibayar'
            elif 'harga asli produk' in col_lower:
                rename_map[col] = 'Harga_Asli'
            elif 'total diskon produk' in col_lower:
                rename_map[col] = 'Total_Diskon'
            elif 'biaya administrasi' in col_lower:
                rename_map[col] = 'Biaya_Admin'
            elif 'biaya layanan' in col_lower:
                rename_map[col] = 'Biaya_Layanan'
            elif 'biaya proses pesanan' in col_lower:
                rename_map[col] = 'Biaya_Proses'
            elif 'premi' in col_lower:
                rename_map[col] = 'Premi'
            elif 'ongkir dibayar pembeli' in col_lower:
                rename_map[col] = 'Ongkir_Pembeli'
            elif 'tanggal dana dilepaskan' in col_lower:
                rename_map[col] = 'Tanggal_Dana'
            elif 'waktu pesanan dibuat' in col_lower:
                rename_map[col] = 'Waktu_Pesanan'
            elif 'metode pembayaran' in col_lower:
                rename_map[col] = 'Metode_Bayar'
            elif 'username (pembeli)' in col_lower:
                rename_map[col] = 'Username_Pembeli'
            elif 'jasa kirim' in col_lower or 'nama kurir' in col_lower:
                # Hanya rename pertama, skip jika sudah ada Jasa_Kirim
                if 'Jasa_Kirim' not in rename_map.values():
                    rename_map[col] = 'Jasa_Kirim'
                else:
                    # Rename dengan suffix
                    rename_map[col] = f'Jasa_Kirim_{len([v for v in rename_map.values() if "Jasa_Kirim" in str(v)])}'
        
        # Rename kolom
        df.rename(columns=rename_map, inplace=True)
        
        # ============================================================
        # HAPUS LAGI DUPLIKAT SETELAH RENAME
        # ============================================================
        df = df.loc[:, ~df.columns.duplicated()]
        
        # ============================================================
        # KONVERSI KOLOM NUMERIK
        # ============================================================
        numeric_cols = ['Total_Dibayar', 'Harga_Asli', 'Total_Diskon', 
                       'Biaya_Admin', 'Biaya_Layanan', 'Biaya_Proses', 'Premi', 'Ongkir_Pembeli']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        
        # Tambah metadata
        df['Sumber_File'] = filename
        df['Tanggal_Upload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Ambil info toko dari sheet Summary
        nama_toko = extract_nama_toko(filename)
        if 'Summary' in available_sheets:
            try:
                ws_summary = wb['Summary']
                for row in ws_summary.iter_rows(min_row=1, max_row=15, values_only=True):
                    if row[0] and 'nama toko' in str(row[0]).lower():
                        if len(row) >= 2 and row[1]:
                            nama_toko = str(row[1]).strip()
                        break
            except:
                pass
        
        # Mapping nama toko biar cocok dengan iklan_data
        if 'KL' in filename.upper() or 'KARYA' in nama_toko.upper():
            nama_toko = 'KL Shopee'
        elif 'NT' in filename.upper() or 'NUSANTARA' in nama_toko.upper():
            nama_toko = 'NT Shopee'
        elif 'PIP' in filename.upper() or 'PRIMA INTI' in nama_toko.upper():
            nama_toko = 'PIP Shopee'
        elif 'IPP' in filename.upper() or 'INTI PRIMA' in nama_toko.upper():
            nama_toko = 'IPP Shopee'
        elif 'DRC' in filename.upper():
            nama_toko = 'DRC Shopee'
        
        df['Nama_Toko'] = nama_toko   # ← Perbaikan: pakai nama_toko, bukan extract_nama_toko
        
        return df, f"OK: {len(df)} transaksi"
        
    except Exception as e:
        return pd.DataFrame(), f"Error: {str(e)}"
# =====================================
# PROCESS SHOPEE FILE - SIMPLE & WORKING
# =====================================

def process_shopee_file(df_raw, filename):
    try:
        nama_toko = extract_nama_toko(filename)
        df = df_raw.copy()
        
        # Bersihkan nama kolom
        df.columns = [str(col).strip() for col in df.columns]
        
        # Buat kolom SKU dari Nomor Referensi SKU
        if 'Nomor Referensi SKU' in df.columns:
            df['SKU'] = df['Nomor Referensi SKU']
        
        # Buat kolom Terjual dari Jumlah
        if 'Jumlah' in df.columns:
            df['Terjual'] = pd.to_numeric(df['Jumlah'], errors='coerce').fillna(0).astype(int)
        
        # Filter hanya pesanan SELESAI
        if 'Status Pesanan' in df.columns:
            df = df[df['Status Pesanan'].astype(str).str.strip().str.lower() == 'selesai']
        
        # Konversi kolom harga
        for col in ['Total Pembayaran', 'Dibayar Pembeli', 'Harga Setelah Diskon', 'Harga Awal']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Tambah info
        df['Nama Toko'] = nama_toko
        df['Sumber_File'] = filename
        df['Tanggal_Upload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Pastikan No Pesanan ada
        if 'No Pesanan' not in df.columns and 'No. Pesanan' in df.columns:
            df['No Pesanan'] = df['No. Pesanan']
        
        return df.reset_index(drop=True)
    except Exception as e:
        st.error(f"Error: {e}")
        return pd.DataFrame()

# =====================================
# DATABASE FUNCTIONS
# =====================================
def load_master_produk():
    if os.path.exists(MASTER_PRODUK_DB_FILE):
        try:
            return pd.read_csv(MASTER_PRODUK_DB_FILE)
        except:
            pass
    return pd.DataFrame(columns=['SKU','Nama_Produk','Harga_Beli','Harga_Jual_Default','Packing','Stok_Awal','Stok_Saat_Ini','Supplier_Default'])

def save_master_produk(df):
    df.to_csv(MASTER_PRODUK_DB_FILE, index=False)

def load_pembelian_data():
    if os.path.exists(PEMBELIAN_DB_FILE):
        try:
            return pd.read_csv(PEMBELIAN_DB_FILE)
        except:
            pass
    return pd.DataFrame(columns=['Tanggal','No_Invoice','Supplier','SKU','Nama_Barang','Qty','Harga_Beli','Total','Metode','Jatuh_Tempo','Status','No_Ref','Catatan'])

def save_pembelian_data(df):
    df.to_csv(PEMBELIAN_DB_FILE, index=False)

def update_stok_tambah(sku, qty):
    if not st.session_state.master_produk.empty:
        mask = st.session_state.master_produk['SKU'].astype(str).str.strip() == str(sku).strip()
        if mask.any():
            idx = st.session_state.master_produk[mask].index[0]
            st.session_state.master_produk.at[idx, 'Stok_Saat_Ini'] = int(st.session_state.master_produk.at[idx, 'Stok_Saat_Ini']) + int(qty)
            save_master_produk(st.session_state.master_produk)
            return True
    return False

# =====================================
# FUNGSI FIFO
# =====================================
def load_stok_batch():
    if os.path.exists(STOK_BATCH_DB_FILE):
        try:
            return pd.read_csv(STOK_BATCH_DB_FILE)
        except:
            pass
    return pd.DataFrame(columns=['SKU','Tanggal_Masuk','Qty_Awal','Qty_Sisa','Harga_Satuan','Sumber','Keterangan'])

def save_stok_batch(df):
    df.to_csv(STOK_BATCH_DB_FILE, index=False)

def tambah_stok_batch(sku, qty, harga, sumber='Pembelian', keterangan=''):
    if 'stok_batch' not in st.session_state:
        st.session_state.stok_batch = load_stok_batch()
    
    batch_baru = pd.DataFrame([{
        'SKU': str(sku).strip(),
        'Tanggal_Masuk': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Qty_Awal': int(qty),
        'Qty_Sisa': int(qty),
        'Harga_Satuan': int(harga),
        'Sumber': sumber,
        'Keterangan': str(keterangan)
    }])
    st.session_state.stok_batch = pd.concat([st.session_state.stok_batch, batch_baru], ignore_index=True)
    save_stok_batch(st.session_state.stok_batch)

def kurangi_stok_fifo(sku, qty_keluar):
    if 'stok_batch' not in st.session_state:
        st.session_state.stok_batch = load_stok_batch()
    
    df_batch = st.session_state.stok_batch.copy()
    mask = (df_batch['SKU'].astype(str).str.strip() == str(sku).strip()) & (df_batch['Qty_Sisa'] > 0)
    batch_tersedia = df_batch[mask].sort_values('Tanggal_Masuk')
    
    if batch_tersedia.empty:
        return 0, 0
    
    sisa_keluar = int(qty_keluar)
    total_hpp = 0
    total_terkurangi = 0
    
    for idx in batch_tersedia.index:
        if sisa_keluar <= 0:
            break
        
        qty_batch = int(st.session_state.stok_batch.at[idx, 'Qty_Sisa'])
        harga_batch = int(st.session_state.stok_batch.at[idx, 'Harga_Satuan'])
        
        if qty_batch >= sisa_keluar:
            st.session_state.stok_batch.at[idx, 'Qty_Sisa'] = qty_batch - sisa_keluar
            total_hpp += sisa_keluar * harga_batch
            total_terkurangi += sisa_keluar
            sisa_keluar = 0
        else:
            st.session_state.stok_batch.at[idx, 'Qty_Sisa'] = 0
            total_hpp += qty_batch * harga_batch
            total_terkurangi += qty_batch
            sisa_keluar -= qty_batch
    
    save_stok_batch(st.session_state.stok_batch)
    return total_hpp, total_terkurangi

def update_stok_saat_ini_dari_batch():
    if 'stok_batch' not in st.session_state or st.session_state.stok_batch.empty:
        return
    
    if not st.session_state.master_produk.empty:
        total_per_sku = st.session_state.stok_batch.groupby('SKU')['Qty_Sisa'].sum().reset_index()
        for _, row in total_per_sku.iterrows():
            sku_key = str(row['SKU']).strip()
            mask = st.session_state.master_produk['SKU'].astype(str).str.strip() == sku_key
            if mask.any():
                idx = st.session_state.master_produk[mask].index[0]
                st.session_state.master_produk.at[idx, 'Stok_Saat_Ini'] = int(row['Qty_Sisa'])
        save_master_produk(st.session_state.master_produk)

def get_hpp_fifo(sku):
    if 'stok_batch' not in st.session_state:
        st.session_state.stok_batch = load_stok_batch()
    
    mask = (st.session_state.stok_batch['SKU'].astype(str).str.strip() == str(sku).strip()) & (st.session_state.stok_batch['Qty_Sisa'] > 0)
    batch_ada = st.session_state.stok_batch[mask]
    
    if batch_ada.empty:
        if not st.session_state.master_produk.empty:
            mask_mp = st.session_state.master_produk['SKU'].astype(str).str.strip() == str(sku).strip()
            if mask_mp.any():
                return int(st.session_state.master_produk[mask_mp]['Harga_Beli'].iloc[0])
        return 0
    
    total_nilai = (batch_ada['Qty_Sisa'] * batch_ada['Harga_Satuan']).sum()
    total_qty = batch_ada['Qty_Sisa'].sum()
    return int(total_nilai / total_qty) if total_qty > 0 else 0

# =====================================
# LOAD ALL DATA
# =====================================
def load_all_data():
    if 'shopee_data' not in st.session_state:
        st.session_state.shopee_data = pd.read_csv(SHOPEE_DB_FILE) if os.path.exists(SHOPEE_DB_FILE) else pd.DataFrame()
    if 'iklan_data' not in st.session_state:
        st.session_state.iklan_data = pd.read_csv(IKLAN_DB_FILE) if os.path.exists(IKLAN_DB_FILE) else pd.DataFrame()
    if 'dana_masuk' not in st.session_state:
        st.session_state.dana_masuk = pd.read_csv(DANA_DB_FILE) if os.path.exists(DANA_DB_FILE) else pd.DataFrame(columns=['No Pesanan','Nama Toko','SKU','Qty','Status Pesanan','HPP','Dibayar Shopee','Tanggal'])
    if 'hpp_db' not in st.session_state:
        st.session_state.hpp_db = pd.read_csv(HPP_DB_FILE) if os.path.exists(HPP_DB_FILE) else pd.DataFrame(columns=['SKU','Qty','HPP','Packing'])
    if 'pembelian_data' not in st.session_state:
        st.session_state.pembelian_data = load_pembelian_data()
    if 'master_produk' not in st.session_state:
        st.session_state.master_produk = load_master_produk()
    if 'stok_batch' not in st.session_state:
        st.session_state.stok_batch = load_stok_batch()
    if 'rekap_iklan_gabungan' not in st.session_state:
        st.session_state.rekap_iklan_gabungan = pd.read_csv(REKAP_IKLAN_MASTER_FILE) if os.path.exists(REKAP_IKLAN_MASTER_FILE) else pd.DataFrame()

# =====================================
# SESSION STATE INIT
# =====================================
load_all_data()
if 'processed_shopee_files' not in st.session_state:
    st.session_state.processed_shopee_files = set()
if 'processed_iklan_files' not in st.session_state:
    st.session_state.processed_iklan_files = set()
if 'shopee_uploader_key' not in st.session_state:
    st.session_state.shopee_uploader_key = 0
if 'iklan_uploader_key' not in st.session_state:
    st.session_state.iklan_uploader_key = 0
if 'target_roas' not in st.session_state:
    st.session_state.target_roas = 3
if 'total_aset' not in st.session_state:
    st.session_state.total_aset = 0
if 'active_menu' not in st.session_state:
    st.session_state.active_menu = "🏠 Dashboard"

# ===== TAMBAHAN UNTUK UPLOAD PEMBAYARAN =====
if 'income_data' not in st.session_state:
    INCOME_DB_FILE = str(DB_FOLDER / "income_database.csv")
    if os.path.exists(INCOME_DB_FILE):
        st.session_state.income_data = pd.read_csv(INCOME_DB_FILE)
    else:
        st.session_state.income_data = pd.DataFrame()
if 'processed_income_files' not in st.session_state:
    st.session_state.processed_income_files = set()
if 'income_uploader_key' not in st.session_state:
    st.session_state.income_uploader_key = 0
# =====================================
# CSS
# =====================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
* { font-family: 'Inter', sans-serif !important; }
.stApp { background: #f8f9fa; }
.card {
    background: white; border-radius: 14px; padding: 20px; margin-bottom: 18px;
    border: 1px solid #E2E8F0; box-shadow: 0 1px 6px rgba(0,0,0,0.04);
}
.stButton > button {
    background: linear-gradient(135deg, #0D9488 0%, #0D9488 100%) !important;
    color: white !important; font-weight: 600 !important; border-radius: 10px !important;
    border: none !important; padding: 8px 16px !important; font-size: 14px !important;
}
.metric-card {
    background: white; border-radius: 14px; padding: 16px; text-align: center;
    border: 1px solid #E2E8F0; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
}
.metric-value { font-size: 50px; font-weight: 800; color: #0D9488; }
.metric-label { font-size: 12px; color: #64748B; margin-top: 5px; }

/* Perbesar font di semua tabel */
div[data-testid="stTable"] table,
div[data-testid="stDataFrame"] table,
table {
    font-size: 80px !important;
}
div[data-testid="stTable"] td,
div[data-testid="stDataFrame"] td,
table td,
table th {
    font-size: 16px !important;
}
</style>
""", unsafe_allow_html=True)

# =====================================
# SIDEBAR
# =====================================
with st.sidebar:
    st.markdown("""<div style="text-align:center;padding:20px 0;">
        <div style="font-size:48px;">🎯</div>
        <div style="font-size:22px;font-weight:800;color:#333;">BSELLER</div>
        <div style="font-size:11px;color:#555;">Premium Analytics v5.0</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    
    menu_structure = {
        "📊 MAIN": ["🏠 Dashboard"],
        "📥 DATA": ["📤 Upload Pesanan", "📢 Upload Iklan", "💳 Upload Pembayaran", "💰 Input Dana"],
        "📦 INVENTORY": ["📋 Master Produk", "🛒 Pembelian Stok", "📦 Kartu Stok"],
        "📈 ANALISIS": ["📢 Rekap Iklan", "🔍 Filter Data", "💎 Laporan Keuangan", "🧮 Simulasi ROAS"],
        "⚙️ LAIN": ["🤖 Analisis"]
    }
    
    for category, menus in menu_structure.items():
        st.markdown(f'<div style="font-size:10px;font-weight:700;color:#94A3B8;text-transform:uppercase;padding:8px 8px 4px 8px;">{category}</div>', unsafe_allow_html=True)
        for menu in menus:
            if st.button(menu, key=f"btn_{menu}", use_container_width=True):
                st.session_state.active_menu = menu
                st.rerun()

active_menu = st.session_state.active_menu

# =====================================
# HEADER
# =====================================
st.markdown(f"""
<div style="background: linear-gradient(135deg, #0D9488 20%, #0D9488 100%); color: white; border-radius: 14px; padding: 15px 24px; margin-bottom: 24px; display: flex; align-items: center; justify-content: space-between;">
    <div style="font-size:18px;font-weight:700;">🎯 BSELLER ANALYTICS</div>
    <div style="font-size:13px;background:rgba(255,255,255,0.2);padding:6px 16px;border-radius:30px;">{active_menu}</div>
</div>
""", unsafe_allow_html=True)

# =====================================
# 🏠 DASHBOARD
# =====================================
if active_menu == "🏠 Dashboard":
    tp = len(st.session_state.pembelian_data) if not st.session_state.pembelian_data.empty else 0
    tps = len(st.session_state.dana_masuk) if not st.session_state.dana_masuk.empty else 0
    tpen = pd.to_numeric(st.session_state.dana_masuk['Dibayar Shopee'], errors='coerce').sum() if not st.session_state.dana_masuk.empty and 'Dibayar Shopee' in st.session_state.dana_masuk.columns else 0
    tik = pd.to_numeric(st.session_state.iklan_data['Biaya'], errors='coerce').sum() if not st.session_state.iklan_data.empty and 'Biaya' in st.session_state.iklan_data.columns else 0
    roas = round(tpen / tik, 2) if tik > 0 else 0
    
    # Total Aset = Stok Barang (Harga Beli × Qty)
    total_aset_stok = 0
    if not st.session_state.master_produk.empty and 'Stok_Saat_Ini' in st.session_state.master_produk.columns and 'Harga_Beli' in st.session_state.master_produk.columns:
        for _, r in st.session_state.master_produk.iterrows():
            qty = int(r['Stok_Saat_Ini']) if pd.notna(r['Stok_Saat_Ini']) else 0
            hrg_beli = int(r['Harga_Beli']) if pd.notna(r['Harga_Beli']) else 0
            total_aset_stok += qty * hrg_beli
    
    st.session_state.total_aset = total_aset_stok
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.markdown(f"""<div class="metric-card"><div style="font-size:24px;">📦</div><div class="metric-value">{format_angka(tp)}</div><div class="metric-label">Total Pembelian</div></div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""<div class="metric-card"><div style="font-size:24px;">🛒</div><div class="metric-value">{format_angka(tps)}</div><div class="metric-label">Total Pesanan</div></div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""<div class="metric-card"><div style="font-size:24px;">💰</div><div class="metric-value">{format_rupiah(tpen)}</div><div class="metric-label">Total Penjualan</div></div>""", unsafe_allow_html=True)
    with col4:
        # Total Aset dengan expander untuk detail
        with st.expander(f"💎 Total Aset: {format_rupiah(st.session_state.total_aset)}", expanded=False):
            if not st.session_state.master_produk.empty:
                df_aset = st.session_state.master_produk[['SKU', 'Nama_Produk', 'Stok_Saat_Ini', 'Harga_Beli']].copy()
                df_aset = df_aset[df_aset['Stok_Saat_Ini'] > 0]
                df_aset['Nilai'] = df_aset['Stok_Saat_Ini'] * df_aset['Harga_Beli']
                df_aset = df_aset.sort_values('Nilai', ascending=False)
                for col in ['Harga_Beli', 'Nilai']:
                    df_aset[col] = df_aset[col].apply(format_rupiah)
                st.dataframe(df_aset, use_container_width=True, height=250, hide_index=True)
    with col5:
        color = "#10b981" if roas >= st.session_state.target_roas else "#ef4444"
        st.markdown(f"""<div class="metric-card"><div style="font-size:24px;">📈</div><div class="metric-value" style="color:{color}">{roas}</div><div class="metric-label">ROAS (Target: {st.session_state.target_roas})</div></div>""", unsafe_allow_html=True)
    
    # ============================================================
    # SPEEDOMETER ROAS (KOMPLIT + CHECKBOX)
    # ============================================================
    st.markdown("---")
    st.subheader("🎯 Speedometer ROAS")
    
    if not st.session_state.iklan_data.empty:
        df_sp = st.session_state.iklan_data.copy()
        col_nama = 'Nama Iklan' if 'Nama Iklan' in df_sp.columns else None
        col_kode = 'Kode Produk' if 'Kode Produk' in df_sp.columns else None
        col_terjual = 'Produk Terjual' if 'Produk Terjual' in df_sp.columns else None
        col_omzet = 'Omzet Penjualan' if 'Omzet Penjualan' in df_sp.columns else None
        col_biaya = 'Biaya' if 'Biaya' in df_sp.columns else None
        col_toko = 'Nama Toko' if 'Nama Toko' in df_sp.columns else None
        col_tgl = 'Tanggal' if 'Tanggal' in df_sp.columns else None
        
        for c in [col_terjual, col_omzet, col_biaya]:
            if c: df_sp[c] = pd.to_numeric(df_sp[c].astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(0).astype(int)
        if col_tgl: df_sp[col_tgl] = pd.to_datetime(df_sp[col_tgl], errors='coerce')
        
        toko_list = ['Semua'] + (sorted(df_sp[col_toko].dropna().unique().tolist()) if col_toko else [])
        toko_pilih = st.selectbox("🏪 Toko", toko_list, key="dash_toko")
        if toko_pilih != 'Semua': df_sp = df_sp[df_sp[col_toko] == toko_pilih]
        
        df_grouped = df_sp.groupby([col_kode, col_nama]).agg({col_terjual:'sum', col_omzet:'sum', col_biaya:'sum'}).reset_index()
        df_grouped[col_kode] = df_grouped[col_kode].astype(str).str.strip()
        df_grouped[col_nama] = df_grouped[col_nama].astype(str).str.strip()
        
        # Checkbox
        col_cek1, col_cek2 = st.columns(2)
        with col_cek1:
            lihat_zombie = st.checkbox("💀 Zombie (0 penjualan)", value=False, key="cek_zombie")
        with col_cek2:
            lihat_rugi = st.checkbox("🔴 Rugi (ROAS < Min)", value=False, key="cek_rugi")
        
        # Default: hanya yang ada penjualan
        if not lihat_zombie:
            df_show = df_grouped[df_grouped[col_terjual] > 0].copy()
        else:
            df_show = df_grouped.copy()
        
        if df_show.empty:
            st.info("📭 Tidak ada iklan.")
        else:
            daftar_iklan = [f"{row[col_kode]} - {row[col_nama][:50]}" for _, row in df_show.iterrows()]
            pilih = st.selectbox("🎯 Pilih Iklan", daftar_iklan, key="dash_pilih")
            kode_pilih = pilih.split(' - ')[0].strip()
            match = df_show[df_show[col_kode] == kode_pilih]
            
            if not match.empty:
                row = match.iloc[0]
                nama = str(row[col_nama]); terjual = int(row[col_terjual])
                omzet = int(row[col_omzet]); biaya = int(row[col_biaya])
                ra = round(omzet/biaya, 1) if biaya > 0 else 0
                
                kode_to_sku = {
                    '23702055121':'Sling5x3M','28171008905':'Sling4x3M','55200717478':'Sling5x5M',
                    '9084726015':'07-470','12684449528':'1108-150','44857049532':'TAFF-30ML',
                    '26643764706':'PIPE-1PC','11008281950':'59-135s','47102879955':'Velcro-8inch',
                    '22584521485':'45-600','8035948643':'AU-AC0984','8574235086':'HK-LS1200',
                    '16834420649':'1108-150','9412670154':'07-470',
                }
                sku = kode_to_sku.get(kode_pilih, '-')
                hpp = 0; hj_default = 0
                
                if sku != '-' and not st.session_state.master_produk.empty:
                    mp = st.session_state.master_produk
                    m = mp[mp['SKU'].astype(str).str.upper() == sku.upper()]
                    if not m.empty:
                        hpp = int(m['Harga_Beli'].iloc[0]) + int(m['Packing'].iloc[0])
                        if 'Harga_Jual_Default' in mp.columns:
                            val = pd.to_numeric(m['Harga_Jual_Default'].iloc[0], errors='coerce')
                            hj_default = int(val) if pd.notna(val) else 0
                
                if hj_default == 0: hj_default = omzet // terjual if terjual > 0 else 0
                roas_min = round(hj_default / (hj_default - hpp), 1) if (hj_default - hpp) > 0 else 0
                laba_pc = hj_default - hpp
                
                # Filter Rugi
                if lihat_rugi and ra >= roas_min:
                    st.info("ℹ️ Iklan ini UNTUNG, tidak termasuk kategori rugi.")
                    st.stop()
                
                col_sp1, col_sp2 = st.columns([1, 1])
                with col_sp1:
                    fig = go.Figure(go.Indicator(
                        mode="gauge+number+delta",
                        value=ra, delta={'reference': roas_min, 'increasing': {'color': "green"}},
                        gauge={
                            'axis': {'range': [0, max(ra*1.5, roas_min*2, 10)]},
                            'bar': {'color': "green" if ra >= roas_min else "red"},
                            'steps': [
                                {'range': [0, roas_min], 'color': "rgba(255,0,0,0.2)"},
                                {'range': [roas_min, max(ra, roas_min*2)], 'color': "rgba(0,255,0,0.2)"}],
                            'threshold': {'line': {'color': "red", 'width': 3}, 'value': roas_min}
                        },
                        title={'text': f"ROAS<br><sup>Min: {roas_min}x</sup>"}
                    ))
                    fig.update_layout(height=300, margin=dict(l=20, r=20, t=50, b=20))
                    st.plotly_chart(fig, use_container_width=True)
                
                with col_sp2:
                    st.markdown("### 📊 Detail")
                    st.metric("📦 SKU", sku)
                    st.metric("💰 Harga Jual", format_rupiah(hj_default))
                    st.metric("💎 HPP", format_rupiah(hpp))
                    st.metric("💵 Laba/PC", format_rupiah(laba_pc))
                    st.metric("🎯 ROAS Minimal", f"{roas_min}x")
                    st.metric("📈 ROAS Aktual", f"{ra}x")
                    
                    if ra >= roas_min and ra > 0:
                        st.success(f"✅ UNTUNG! ROAS {ra}x > Minimal {roas_min}x")
                    elif ra > 0:
                        st.error(f"❌ RUGI! ROAS {ra}x < Minimal {roas_min}x")
                    else:
                        st.warning("💀 ZOMBIE! 0 penjualan")
                
                st.markdown("---")
                c1,c2,c3 = st.columns(3)
                c1.metric("💰 Omzet", format_rupiah(omzet))
                c2.metric("📢 Biaya Iklan", format_rupiah(biaya))
                c3.metric("📦 Terjual", f"{terjual} pcs")
                
                if col_tgl:
                    st.markdown("---")
                    st.subheader("📈 Tren ROAS Harian")
                    df_tren = df_sp[df_sp[col_kode].astype(str).str.strip() == kode_pilih].copy()
                    df_tren['Tanggal'] = df_tren[col_tgl].dt.date
                    tren = df_tren.groupby('Tanggal').agg({col_omzet:'sum', col_biaya:'sum'}).reset_index()
                    tren['ROAS'] = round(tren[col_omzet] / tren[col_biaya].replace(0, 1), 1)
                    
                    fig2 = go.Figure()
                    fig2.add_trace(go.Bar(x=tren['Tanggal'], y=tren[col_omzet], name='Omzet', marker_color='green'))
                    fig2.add_trace(go.Bar(x=tren['Tanggal'], y=tren[col_biaya], name='Biaya', marker_color='red'))
                    fig2.add_trace(go.Scatter(x=tren['Tanggal'], y=tren['ROAS']*10000, name='ROAS (x10K)', yaxis='y2', line=dict(color='blue', width=3)))
                    fig2.update_layout(height=250, margin=dict(l=10, r=10, t=30, b=10), legend=dict(orientation='h', y=1.1))
                    st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("📭 Upload data iklan dulu.")

# =====================================
# 🤖 ANALISIS - DOKTER SPESIALIS + BUDGET
# =====================================
if active_menu == "🤖 Analisis":
    st.subheader("🩺 Dokter Spesialis Toko")
    
    if st.session_state.shopee_data.empty:
        st.error("❌ Upload data penjualan dulu!")
        st.stop()
    if st.session_state.master_produk.empty:
        st.error("❌ Upload Master Produk dulu!")
        st.stop()
    
    df = st.session_state.shopee_data.copy()
    if 'Status Pesanan' in df.columns:
        batal = ['batal', 'dibatalkan', 'canceled']
        df = df[~df['Status Pesanan'].astype(str).str.lower().isin(batal)]
    
    hpp_map = {}
    for _, r in st.session_state.master_produk.iterrows():
        hpp_map[str(r['SKU']).strip().upper()] = int(r.get('Harga_Beli', 0) or 0) + int(r.get('Packing', 0) or 0)
    
    df['Qty'] = pd.to_numeric(df.get('Jumlah Produk di Pesan', 1), errors='coerce').fillna(1).astype(int)
    df['HPP_Unit'] = df['Nomor Referensi SKU'].apply(lambda x: hpp_map.get(str(x).strip().upper(), 0))
    df['HPP_Total'] = df['HPP_Unit'] * df['Qty']
    df['Harga'] = pd.to_numeric(df.get('Harga Awal', df.get('Total Pembayaran', 0)), errors='coerce').fillna(0) * df['Qty'] if 'Harga Awal' in df.columns or 'Total Pembayaran' in df.columns else 0
    df['Estimasi'] = (df['Harga'] * 0.78).astype(int)
    
    sku = df.groupby(['Nomor Referensi SKU', 'Nama Produk']).agg(Qty=('Qty','sum'), Omzet=('Harga','sum'), HPP=('HPP_Total','sum'), Estimasi=('Estimasi','sum')).reset_index()
    sku['Laba'] = sku['Estimasi'] - sku['HPP']
    sku['Margin'] = ((sku['Laba'] / sku['Estimasi']) * 100).fillna(0)
    
    if not st.session_state.iklan_data.empty and 'Kode Produk' in st.session_state.iklan_data.columns:
        df_ik = st.session_state.iklan_data.copy()
        df_ik['Biaya'] = pd.to_numeric(df_ik['Biaya'].astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(0)
        mapping = {'23702055121':'Sling5x3M','28171008905':'Sling4x3M','9084726015':'07-470','12684449528':'1108-150','11008281950':'59-135s','22584521485':'45-600'}
        df_ik['SKU'] = df_ik['Kode Produk'].astype(str).map(mapping)
        iklan = df_ik.groupby('SKU')['Biaya'].sum().reset_index()
        iklan.columns = ['Nomor Referensi SKU', 'Biaya_Iklan']
        sku = pd.merge(sku, iklan, on='Nomor Referensi SKU', how='left')
        sku['Biaya_Iklan'] = sku['Biaya_Iklan'].fillna(0)
    else:
        sku['Biaya_Iklan'] = 0
    
    sku['Laba_Bersih'] = sku['Laba'] - sku['Biaya_Iklan']
    sku['ROAS'] = round(sku['Estimasi'] / sku['Biaya_Iklan'].replace(0, 1), 1)
    
    hasil = []
    for _, r in sku.iterrows():
        m = r['Margin']
        laba = int(r['Laba_Bersih'])
        hpp_unit = hpp_map.get(str(r['Nomor Referensi SKU']).strip().upper(), 0)
        harga_ideal = int(hpp_unit / 0.78 * 1.2) if hpp_unit > 0 else 0
        
        if m < 0:
            status = "🔴 KRITIS"
            resep = f"NIK HARGA → {format_rupiah(harga_ideal)} | STOP IKLAN"
        elif m < 10:
            status = "🟡 WASPADA"
            resep = f"NIK HARGA → {format_rupiah(harga_ideal)}"
        elif m < 20:
            status = "🟢 SEHAT"
            resep = "PERTAHANKAN"
        else:
            status = "🟣 BINTANG"
            resep = "GENJOT IKLAN 3X"
        
        hasil.append({'SKU': r['Nomor Referensi SKU'], 'Nama': r['Nama Produk'], 'Qty': int(r['Qty']),
            'Margin': f"{m:.0f}%", 'ROAS': f"{r['ROAS']}x", 'Laba': laba, 'Iklan': int(r['Biaya_Iklan']),
            'Status': status, 'Resep': resep, 'Margin_Num': m, 'Harga_Ideal': harga_ideal})
    
    df_hasil = pd.DataFrame(hasil).sort_values('Margin_Num')
    
    kritis = df_hasil[df_hasil['Margin_Num'] < 0]
    waspada = df_hasil[(df_hasil['Margin_Num'] >= 0) & (df_hasil['Margin_Num'] < 10)]
    bintang = df_hasil[df_hasil['Margin_Num'] >= 20]
    
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("🔴 KRITIS", len(kritis))
    c2.metric("🟡 WASPADA", len(waspada))
    c3.metric("🟣 BINTANG", len(bintang))
    c4.metric("💰 LABA", format_rupiah(int(df_hasil['Laba'].sum())))
    
    if len(kritis) > 0:
        st.error(f"🚨 OPERASI {len(kritis)} SKU KRITIS!")
        for _, r in kritis.iterrows():
            with st.expander(f"🔴 {r['SKU']} - {r['Nama']} (Rugi: {format_rupiah(abs(r['Laba']))})", expanded=True):
                st.write(f"💊 {r['Resep']}")
    
    display = df_hasil[['Status','SKU','Nama','Qty','Margin','ROAS','Laba','Iklan','Resep']].copy()
    display['Laba'] = display['Laba'].apply(format_rupiah)
    display['Iklan'] = display['Iklan'].apply(format_rupiah)
    st.dataframe(display, use_container_width=True, height=400, hide_index=True)
    
    # ===================================
    # BUDGET IKLAN AI
    # ===================================
    st.markdown("---")
    st.subheader("📈 Budget Iklan AI")
    
    if not st.session_state.iklan_data.empty:
        if 'budget_default' not in st.session_state:
            st.session_state.budget_default = 80000
        budget_default = st.session_state.budget_default
        
        col_b1, col_b2, col_b3 = st.columns(3)
        with col_b1:
            toko_list = ['Semua'] + st.session_state.iklan_data['Nama Toko'].dropna().unique().tolist() if 'Nama Toko' in st.session_state.iklan_data.columns else ['Semua']
            toko_bud = st.selectbox("🏪 Toko", toko_list, key="bud_toko")
        with col_b2:
            if 'Tanggal' in st.session_state.iklan_data.columns:
                df_t = st.session_state.iklan_data.copy()
                df_t['Tanggal'] = pd.to_datetime(df_t['Tanggal'], errors='coerce')
                tgl_bud = st.date_input("📅 Tanggal", [df_t['Tanggal'].min().date(), df_t['Tanggal'].max().date()], key="bud_tgl")
            else:
                tgl_bud = []
        with col_b3:
            target_roas = st.number_input("🎯 Target ROAS", value=7.0, step=0.5, key="bud_roas")
        
        # Opsi Budget
        st.markdown("**💰 Budget Default:**")
        col_o1, col_o2, col_o3, col_o4 = st.columns(4)
        with col_o1:
            if st.button("💎 100.000", use_container_width=True): st.session_state.budget_default = 100000; st.rerun()
        with col_o2:
            if st.button("💰 50.000", use_container_width=True): st.session_state.budget_default = 50000; st.rerun()
        with col_o3:
            if st.button("💵 25.000", use_container_width=True): st.session_state.budget_default = 25000; st.rerun()
        with col_o4:
            st.metric("Saat Ini", f"{budget_default:,.0f}".replace(',', '.'))
        
        df_ikbud = st.session_state.iklan_data.copy()
        df_ikbud['Biaya'] = pd.to_numeric(df_ikbud['Biaya'].astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(budget_default)
        df_ikbud['Omzet'] = pd.to_numeric(df_ikbud.get('Omzet Penjualan', 0).astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(0) if 'Omzet Penjualan' in df_ikbud.columns else 0
        df_ikbud['Terjual'] = pd.to_numeric(df_ikbud.get('Produk Terjual', 0).astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(0) if 'Produk Terjual' in df_ikbud.columns else 0
        
        if toko_bud != 'Semua' and 'Nama Toko' in df_ikbud.columns:
            df_ikbud = df_ikbud[df_ikbud['Nama Toko'] == toko_bud]
        if 'Tanggal' in df_ikbud.columns and len(tgl_bud) == 2:
            df_ikbud['Tgl'] = pd.to_datetime(df_ikbud['Tanggal'], errors='coerce')
            df_ikbud = df_ikbud[(df_ikbud['Tgl'].dt.date >= tgl_bud[0]) & (df_ikbud['Tgl'].dt.date <= tgl_bud[1])]
        
        if not df_ikbud.empty:
            # Database SKU dari Penjualan
            nama_to_sku = {}
            if not st.session_state.shopee_data.empty:
                for _, r in st.session_state.shopee_data[['Nomor Referensi SKU', 'Nama Produk']].drop_duplicates().iterrows():
                    nama_to_sku[str(r['Nama Produk']).upper()] = str(r['Nomor Referensi SKU']).strip().upper()
            
            def cari_sku(nama):
                if pd.isna(nama): return '-'
                n = str(nama).upper()
                if n in nama_to_sku: return nama_to_sku[n]
                for db_n, db_s in nama_to_sku.items():
                    if ' '.join(n.split()[:2]) in db_n: return db_s
                return '-'
            
            def cari_nama(sku):
                if sku == '-': return '-'
                m = st.session_state.shopee_data[st.session_state.shopee_data['Nomor Referensi SKU'].astype(str).str.upper() == sku.upper()]
                return str(m['Nama Produk'].iloc[0]) if not m.empty else '-'
            
            cols_group = ['Tgl', 'Kode Produk', 'Nama Toko']
            if 'Nama Iklan' in df_ikbud.columns: cols_group.append('Nama Iklan')
            
            ik_group = df_ikbud.groupby(cols_group).agg(Terjual=('Terjual','sum'), Biaya=('Biaya','sum'), Omzet=('Omzet','sum')).reset_index()
            ik_group['ROAS_Aktual'] = round(ik_group['Omzet'] / ik_group['Biaya'].replace(0, 1), 1)
            ik_group['Budget_Harian'] = ik_group.apply(lambda x: int(x['Biaya']*(x['ROAS_Aktual']/target_roas)) if x['ROAS_Aktual']>target_roas else (int(x['Biaya']*0.5) if x['ROAS_Aktual']>target_roas*0.5 else budget_default), axis=1)
            
            def sts(r):
                if r['ROAS_Aktual']>target_roas*1.5: return "🚀 GAS POLL"
                elif r['ROAS_Aktual']>target_roas: return "✅ NAIK"
                elif r['ROAS_Aktual']>target_roas*0.5: return "⚠️ TAHAN"
                elif r['Biaya']>0: return "⛔ STOP"
                else: return "💀 NO DATA"
            ik_group['Status'] = ik_group.apply(sts, axis=1)
            ik_group['SKU'] = ik_group['Nama Iklan'].apply(cari_sku) if 'Nama Iklan' in ik_group.columns else '-'
            ik_group['Nama_Produk'] = ik_group['SKU'].apply(cari_nama)
            
            total_biaya = int(ik_group['Biaya'].sum())
            total_budget = int(ik_group['Budget_Harian'].sum())
            
            c1,c2,c3 = st.columns(3)
            c1.metric("📢 Total Biaya", f"{total_biaya:,.0f}".replace(',', '.'))
            c2.metric("💰 Total Budget", f"{total_budget:,.0f}".replace(',', '.'))
            c3.metric("📈 ROAS", f"{round(ik_group['Omzet'].sum()/total_biaya,1)}x" if total_biaya>0 else "0x")
            
            ik_group['Tgl_Str'] = ik_group['Tgl'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else '-')
            
            display_bud = pd.DataFrame({
                'Tgl': ik_group['Tgl_Str'],
                'Nama Toko': ik_group['Nama Toko'],
                'Kode Produk': ik_group.get('Kode Produk', '-'),
                'Nama Produk': ik_group['Nama_Produk'],
                'SKU': ik_group['SKU'],
                'Terjual': ik_group['Terjual'],
                'Budget Harian': ik_group['Budget_Harian'].apply(lambda x: f"{x:,.0f}".replace(',', '.')),
                'Biaya': ik_group['Biaya'].apply(lambda x: f"{x:,.0f}".replace(',', '.')),
                'ROAS Aktual': ik_group['ROAS_Aktual'].apply(lambda x: f"{x}x"),
                'Status': ik_group['Status']
            })
            
            edited = st.data_editor(display_bud, column_config={"Budget Harian": st.column_config.TextColumn("✏️ Budget Harian")}, use_container_width=True, height=400, hide_index=True)
            
            col_s1, col_s2 = st.columns(2)
            with col_s1:
                if st.button("💾 Simpan", use_container_width=True):
                    st.session_state.riwayat_budget = st.session_state.get('riwayat_budget', [])
                    total_edited = total_budget
                    if edited is not None:
                        try: total_edited = sum(int(str(x).replace('.','')) for x in edited['Budget Harian'])
                        except: pass
                    st.session_state.riwayat_budget.append({'Tgl': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M'), 'Toko': toko_bud, 'Total': total_edited, 'ROAS': target_roas})
                    st.success("✅ Tersimpan!")
            with col_s2:
                if st.button("🔄 Reset", use_container_width=True): st.rerun()
            
            if st.session_state.get('riwayat_budget'):
                with st.expander("📋 Riwayat"):
                    df_r = pd.DataFrame(st.session_state.riwayat_budget)
                    df_r['Total'] = df_r['Total'].apply(lambda x: f"{x:,.0f}".replace(',', '.'))
                    st.dataframe(df_r, use_container_width=True, hide_index=True)
        else:
            st.info("📭 Tidak ada data")
    else:
        st.info("📭 Upload data iklan dulu!")
    
    st.download_button("⬇️ Download CSV", df_hasil.to_csv(index=False).encode('utf-8'), "diagnosa.csv", use_container_width=True)
# =====================================
# 📤 UPLOAD PENJUALAN - TAMPILAN LENGKAP
# =====================================
if active_menu == "📤 Upload Pesanan":
    st.markdown("### 📁 Upload File Data Penjualan Shopee (XLSX)")
    
    TARGET = [
        'No Pesanan', 'Status Pesanan', 'Alasan Pembatalan', 'Status Pembatalan/ Pengembalian',
        'No. Resi', 'Opsi Pengiriman', 'Antar ke counter/ pick-up', 'Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)',
        'Waktu Pengiriman Diatur', 'Waktu Pesanan Dibuat', 'Waktu Pembayaran Dilakukan', 'Metode Pembayaran',
        'SKU Induk', 'Nama Produk', 'Nomor Referensi SKU', 'Nama Variasi', 'Harga Awal', 'Harga Setelah Diskon',
        'Terjual', 'Returned quantity', 'Dibayar Pembeli', 'Total Diskon', 'Diskon Dari Penjual', 'Diskon Dari Shopee',
        'Berat Produk', 'Jumlah Produk di Pesan', 'Total Berat', 'Voucher Ditanggung Penjual', 'Cashback Koin',
        'Voucher Ditanggung Shopee', 'Paket Diskon', 'Paket Diskon (Diskon dari Shopee)', 'Paket Diskon (Diskon dari Penjual)',
        'Potongan Koin Shopee', 'Diskon Kartu Kredit', 'Ongkos Kirim Dibayar oleh Pembeli', 'Estimasi Potongan Biaya Pengiriman',
        'Ongkos Kirim Pengembalian Barang', 'Total Pembayaran', 'Perkiraan Ongkos Kirim', 'Catatan dari Pembeli', 'Catatan',
        'Username (Pembeli)', 'Nama Penerima', 'No. Telepon', 'Alamat Pengiriman', 'Kota/Kabupaten', 'Provinsi',
        'Waktu Pesanan Selesai'
    ]
    
    KOLOM_UANG = ['Harga Awal', 'Harga Setelah Diskon', 'Dibayar Pembeli', 'Total Diskon', 'Diskon Dari Penjual',
                  'Diskon Dari Shopee', 'Potongan Koin Shopee', 'Diskon Kartu Kredit', 'Ongkos Kirim Dibayar oleh Pembeli',
                  'Estimasi Potongan Biaya Pengiriman', 'Ongkos Kirim Pengembalian Barang', 'Total Pembayaran', 'Perkiraan Ongkos Kirim']
    
    def format_indonesia(x):
        try:
            if pd.isna(x): return "0"
            return f"{int(float(x)):,}".replace(",", ".")
        except: return "0"
    
    def to_int(val):
        try:
            if pd.isna(val): return 0
            if isinstance(val, str):
                val = val.replace('Rp','').replace(' ','').replace('.','').replace(',','')
            return int(float(val))
        except: return 0
    
    uploaded_files = st.file_uploader("Pilih file Excel", type=["xlsx", "xls"], 
                                       key=f"shopee_upload_{st.session_state.shopee_uploader_key}", 
                                       accept_multiple_files=True, label_visibility="collapsed")
    
    if uploaded_files:
        for f in uploaded_files:
            if f.name not in st.session_state.processed_shopee_files:
                try:
                    df = pd.read_excel(f, engine="openpyxl")
                    df.columns = [str(c).strip() for c in df.columns]
                    
                    hasil = pd.DataFrame()
                    for col in TARGET:
                        match = None
                        for orig in df.columns:
                            if col.replace('.', '').lower() == orig.replace('.', '').lower():
                                match = orig; break
                        if not match:
                            for orig in df.columns:
                                if col.replace('.', '').lower() in orig.replace('.', '').lower():
                                    match = orig; break
                        hasil[col] = df[match] if match else None
                    
                    if 'Status Pesanan' in hasil.columns:
                        status_batal = ['batal', 'dibatalkan', 'canceled', 'cancelled', 'Batal', 'Dibatalkan', 'CANCEL', 'Cancel']
                        before = len(hasil)
                        hasil = hasil[~hasil['Status Pesanan'].astype(str).str.lower().isin(status_batal)]
                        if before > len(hasil): st.info(f"🗑️ Menghapus {before - len(hasil)} data Batal")
                    
                    for col in KOLOM_UANG:
                        if col in hasil.columns: hasil[col] = hasil[col].apply(to_int)
                    
                    for col in ['Harga Awal', 'Harga Setelah Diskon', 'Dibayar Pembeli', 'Total Pembayaran']:
                        if col in hasil.columns:
                            mask = (hasil[col] > 0) & (hasil[col] < 1000)
                            if mask.any():
                                hasil.loc[mask, col] = hasil.loc[mask, col] * 1000
                    
                    for col in ['Terjual', 'Returned quantity', 'Berat Produk', 'Jumlah Produk di Pesan', 'Total Berat', 'Cashback Koin']:
                        if col in hasil.columns: hasil[col] = hasil[col].apply(to_int)
                    
                    hasil['Nama Toko'] = extract_nama_toko(f.name)
                    hasil['Sumber_File'] = f.name
                    hasil['Tanggal_Upload'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    if hasil.empty:
                        st.warning(f"⚠️ {f.name}: Tidak ada data valid")
                        continue
                    
                    if not st.session_state.shopee_data.empty:
                        existing_ids = st.session_state.shopee_data['No Pesanan'].astype(str).tolist() if 'No Pesanan' in st.session_state.shopee_data else []
                        new_data = hasil[~hasil['No Pesanan'].astype(str).isin(existing_ids)] if existing_ids else hasil
                        if not new_data.empty:
                            st.session_state.shopee_data = pd.concat([st.session_state.shopee_data, new_data], ignore_index=True)
                    else:
                        st.session_state.shopee_data = hasil
                    
                    st.session_state.shopee_data.to_csv(SHOPEE_DB_FILE, index=False)
                    st.session_state.processed_shopee_files.add(f.name)
                    
                    if 'Status Pesanan' in hasil.columns and 'Nomor Referensi SKU' in hasil.columns and 'Jumlah Produk di Pesan' in hasil.columns:
                        df_selesai = hasil[hasil['Status Pesanan'].astype(str).str.strip().str.lower() == 'selesai']
                        total_hpp_fifo = 0
                        for _, pesanan in df_selesai.iterrows():
                            sku_fifo = str(pesanan['Nomor Referensi SKU']).strip()
                            qty_fifo = int(pd.to_numeric(pesanan['Jumlah Produk di Pesan'], errors='coerce').fillna(0))
                            if sku_fifo and qty_fifo > 0:
                                hpp_keluar, _ = kurangi_stok_fifo(sku_fifo, qty_fifo)
                                total_hpp_fifo += hpp_keluar
                        update_stok_saat_ini_dari_batch()
                        if total_hpp_fifo > 0:
                            st.caption(f"📦 HPP FIFO terjual: {format_rupiah(total_hpp_fifo)}")
                    
                    st.success(f"✅ {f.name}: {len(hasil)} baris")
                    st.markdown('<audio autoplay src="https://www.soundjay.com/buttons/sounds/button-09.mp3"></audio>', unsafe_allow_html=True)
                    
                except Exception as e:
                    st.error(f"❌ {f.name}: {str(e)}")
        
        st.session_state.shopee_uploader_key += 1
        st.rerun()
    
    if not st.session_state.shopee_data.empty:
        df = st.session_state.shopee_data.copy()
        
        # Filter Toko + Tanggal + Status + Cari
        col_f1, col_f2, col_f3, col_f4 = st.columns(4)
        with col_f1:
            if 'Nama Toko' in df.columns:
                t_list = ['Semua'] + sorted(df['Nama Toko'].dropna().unique().tolist())
                t_pilih = st.selectbox("🏪 Toko", t_list, key="filter_toko_penjualan")
        with col_f2:
            if 'Waktu Pesanan Dibuat' in df.columns:
                df['Tgl_Pesan'] = pd.to_datetime(df['Waktu Pesanan Dibuat'], errors='coerce')
                today = datetime.now().date()
                week_ago = today - timedelta(days=7)
                range_tgl = st.date_input("📅 Tanggal", value=(week_ago, today), key="filter_tgl_penjualan", on_change=lambda: st.rerun())
        with col_f3:
            if 'Status Pesanan' in df.columns:
                s_list = ['Semua'] + sorted(df['Status Pesanan'].dropna().unique().tolist())
                s_pilih = st.selectbox("📋 Status", s_list, key="filter_status_penjualan")
        with col_f4:
            cari = st.text_input("🔎 Cari No. Pesanan / Produk", key="cari_penjualan")
        
        # Terapkan filter setelah semua input
        if t_pilih != 'Semua':
            df = df[df['Nama Toko'] == t_pilih]
        if len(range_tgl) == 2:
            df = df[(df['Tgl_Pesan'].dt.date >= range_tgl[0]) & (df['Tgl_Pesan'].dt.date <= range_tgl[1])]
        if s_pilih != 'Semua':
            df = df[df['Status Pesanan'] == s_pilih]
        if cari:
            mask = (df['No Pesanan'].astype(str).str.contains(cari, case=False, na=False)) | \
                   (df['Nama Produk'].astype(str).str.contains(cari, case=False, na=False))
            df = df[mask]
        
        col1, col2, col3, col4 = st.columns(4)
        with col1: st.metric("📊 Total Data", format_indonesia(len(df)))
        with col2:
            if 'Harga Setelah Diskon' in df.columns and 'Jumlah Produk di Pesan' in df.columns:
                total_penjualan = (df['Harga Setelah Diskon'] * df['Jumlah Produk di Pesan']).sum()
                st.metric("💰 Total Penjualan", format_indonesia(total_penjualan))
            elif 'Harga Setelah Diskon' in df.columns:
                st.metric("💰 Total Penjualan", format_indonesia(df['Harga Setelah Diskon'].sum()))
        with col3:
            if st.button("🗑️ Hapus Semua", key="hapus_semua_penjualan"):
                st.session_state.shopee_data = pd.DataFrame()
                if os.path.exists(SHOPEE_DB_FILE): os.remove(SHOPEE_DB_FILE)
                st.session_state.processed_shopee_files = set()
                st.success("✅ Dihapus!"); st.rerun()
        with col4:
            if st.button("🔄 Reset", key="reset_uploader_penjualan"):
                st.session_state.shopee_uploader_key += 1
                st.session_state.processed_shopee_files = set()
                st.success("✅ Reset!"); st.rerun()
        
        display = df.copy()
        for col in KOLOM_UANG:
            if col in display.columns: display[col] = display[col].apply(format_indonesia)
        for col in ['Terjual', 'Returned quantity', 'Berat Produk', 'Jumlah Produk di Pesan', 'Total Berat', 'Cashback Koin']:
            if col in display.columns: display[col] = display[col].apply(format_indonesia)
        
        st.caption(f"📊 Menampilkan {len(display)} dari {len(st.session_state.shopee_data)} baris")
        st.dataframe(display[[c for c in TARGET if c in display.columns]], use_container_width=True, height=500, hide_index=True)
        st.download_button("📥 Download CSV", df.to_csv(index=False).encode('utf-8'), f"shopee_{datetime.now():%Y%m%d}.csv", "text/csv")
        st.caption("📌 Format: 10.000 | Harga <1000 auto x1000 | Status Batal dihapus")
    else:
        st.info("📭 Upload file Excel Shopee")
# =====================================
# 📢 UPLOAD IKLAN (FIX - PERIODE OTOMATIS)
# =====================================
elif active_menu == "📢 Upload Iklan":
    st.markdown("### 📁 Upload File Data Iklan Shopee (CSV)")
    st.caption("Upload file CSV hasil export dari Shopee Ads | Tanggal otomatis dari Periode")
    
    st.markdown("---")
    
    def fmt_ribuan(x):
        try:
            if pd.isna(x): return "0"
            return f"{int(float(x)):,}".replace(",", ".")
        except: return "0"
    
    NUMERIC_COLS = ['Dilihat', 'Jumlah Klik', 'Konversi', 'Konversi Langsung', 
                    'Produk Terjual', 'Terjual Langsung', 'Omzet Penjualan', 
                    'Penjualan Langsung (GMV Langsung)', 'Biaya', 'Voucher Amount', 'Vouchered Sales']
    
    uploaded_files_csv = st.file_uploader(
        "Pilih file CSV", type=["csv"],
        key=f"iklan_upload_{st.session_state.iklan_uploader_key}",
        accept_multiple_files=True, label_visibility="collapsed"
    )
    
    if uploaded_files_csv:
        new_data, new_files = [], []
        for uploaded_file in uploaded_files_csv:
            if uploaded_file.name not in st.session_state.processed_iklan_files:
                try:
                    raw_data = uploaded_file.getvalue()
                    text = None
                    for enc in ["utf-8-sig", "utf-8", "utf-16", "latin1"]:
                        try: text = raw_data.decode(enc); break
                        except: continue
                    if text is None: continue
                    
                    lines = text.splitlines()
                    
                    # Ambil Periode dari baris ke-5
                    tgl_otomatis = datetime.now().strftime('%Y-%m-%d')
                    for line in lines[:10]:
                        if 'Periode' in line:
                            parts = line.split(';')
                            if len(parts) >= 2:
                                periode = parts[1].strip()
                                tgl_mulai = periode.split(' - ')[0].strip()
                                try:
                                    tgl_otomatis = datetime.strptime(tgl_mulai, '%d/%m/%Y').strftime('%Y-%m-%d')
                                except: pass
                            break
                    
                    header_idx = -1
                    for i, line in enumerate(lines):
                        clean_line = line.replace('\ufeff', '').replace('﻿', '').strip()
                        if "Urutan" in clean_line and "Nama Iklan" in clean_line and "Status" in clean_line:
                            header_idx = i; break
                    if header_idx == -1: continue
                    
                    headers = [h.strip() for h in lines[header_idx].replace('\ufeff','').replace('﻿','').split(";")]
                    data_rows = []
                    for line in lines[header_idx + 1:]:
                        line = line.replace('\ufeff', '').replace('﻿', '')
                        if line.strip() and not line.strip().startswith(";"):
                            row = line.split(";")
                            while len(row) < len(headers): row.append("")
                            data_rows.append(row[:len(headers)])
                    
                    df = pd.DataFrame(data_rows, columns=headers)
                    df = df.dropna(axis=1, how='all').dropna(axis=0, how='all').reset_index(drop=True)
                    
                    nama_toko = extract_nama_toko(uploaded_file.name)
                    

                    # Auto-detect toko dari nama file
                    nama_file_upper = uploaded_file.name.upper()
                    if nama_file_upper.startswith('KL') or 'KARYA' in nama_file_upper:
                        nama_toko = 'KL Shopee'
                    elif nama_file_upper.startswith('NT') or 'NUSANTARA' in nama_file_upper:
                        nama_toko = 'NT Shopee'
                    elif nama_file_upper.startswith('PIP') or 'PRIMA INTI' in nama_file_upper:
                        nama_toko = 'PIP Shopee'
                    elif nama_file_upper.startswith('IPP') or 'INTI PRIMA' in nama_file_upper:
                        nama_toko = 'IPP Shopee'
                    elif nama_file_upper.startswith('DRC'):
                        nama_toko = 'DRC Shopee'
                    df["Nama Toko"] = nama_toko
                    df["Sumber File"] = uploaded_file.name
                    df["Tanggal Upload"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    df["Tanggal"] = tgl_otomatis  # ← DARI PERIODE CSV
                    
                    for col in NUMERIC_COLS:
                        if col in df.columns:
                            df[col] = df[col].astype(str).str.replace(r'[^\d\.\-]', '', regex=True)
                            df[col] = df[col].str.replace('.', '', regex=False)
                            df[col] = df[col].str.replace(',', '.', regex=False)
                            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
                    
                    if 'Nama Iklan' in df.columns:
                        df = df[df['Nama Iklan'].astype(str).str.strip() != '']
                    
                    new_data.append(df)
                    new_files.append(uploaded_file.name)
                    st.success(f"✅ {uploaded_file.name} ({len(df)} baris) - Tgl: {tgl_otomatis}")
                except Exception as e:
                    st.error(f"❌ Error: {uploaded_file.name} - {str(e)}")
        
        if new_data:
            # Concat semua file baru
            merged = pd.concat(new_data, ignore_index=True)
            
            # Baca ulang data yang sudah ada di CSV
            if os.path.exists(IKLAN_DB_FILE):
                data_lama = pd.read_csv(IKLAN_DB_FILE)
            else:
                data_lama = pd.DataFrame()
            
            # Gabung lama + baru
            if not data_lama.empty:
                data_gabung = pd.concat([data_lama, merged], ignore_index=True)
            else:
                data_gabung = merged
            
            # Hapus duplikat
            if 'Nama Iklan' in data_gabung.columns and 'Nama Toko' in data_gabung.columns and 'Tanggal' in data_gabung.columns:
                before = len(data_gabung)
                data_gabung = data_gabung.drop_duplicates(
                    subset=['Nama Iklan', 'Nama Toko', 'Tanggal'], 
                    keep='last'
                )
            
            # Simpan ke CSV
            data_gabung.to_csv(IKLAN_DB_FILE, index=False, encoding="utf-8-sig")
            
            # Update session
            st.session_state.iklan_data = data_gabung
            
            st.success(f"🎉 {len(merged)} data baru ditambahkan! Total: {len(data_gabung)} baris")
            st.session_state.iklan_uploader_key += 1
            st.rerun()
    
    st.markdown("---")
    st.markdown("### 📊 Data Iklan Tersimpan")
    
    if not st.session_state.iklan_data.empty:
        df_iklan = st.session_state.iklan_data.copy()
        
        # Filter Toko + Tanggal + Cari
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            if 'Nama Toko' in df_iklan.columns:
                tl = ['Semua'] + sorted(df_iklan['Nama Toko'].dropna().unique().tolist())
                st_pilih = st.selectbox("🏪 Toko", tl, key="filter_toko_upload")
                if st_pilih != 'Semua':
                    df_iklan = df_iklan[df_iklan['Nama Toko'] == st_pilih]
        with col_f2:
            if 'Tanggal' in df_iklan.columns:
                df_iklan['Tgl'] = pd.to_datetime(df_iklan['Tanggal'], errors='coerce')
                tgl_min = df_iklan['Tgl'].min().date()
                tgl_max = df_iklan['Tgl'].max().date()
                range_tgl_iklan = st.date_input("📅 Tanggal", value=(tgl_min, tgl_max), key="filter_tgl_upload")
                if len(range_tgl_iklan) == 2:
                    df_iklan = df_iklan[(df_iklan['Tgl'].dt.date >= range_tgl_iklan[0]) & (df_iklan['Tgl'].dt.date <= range_tgl_iklan[1])]
        with col_f3:
            cari_iklan = st.text_input("🔎 Cari Iklan", key="cari_iklan_upload")
            if cari_iklan and 'Nama Iklan' in df_iklan.columns:
                df_iklan = df_iklan[df_iklan['Nama Iklan'].astype(str).str.contains(cari_iklan, case=False, na=False)]
        
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("📊 Jumlah Baris", fmt_ribuan(len(df_iklan)))
        c2.metric("📋 Jumlah Kolom", len(df_iklan.columns))
        if 'Biaya' in df_iklan.columns: c3.metric("💰 Total Biaya Iklan", fmt_ribuan(df_iklan['Biaya'].sum()))
        with c4:
            if st.button("🗑️ Hapus Semua", type="secondary", key="hapus_semua_iklan"):
                st.session_state.iklan_data = pd.DataFrame()
                st.session_state.processed_iklan_files = set()
                if os.path.exists(IKLAN_DB_FILE): os.remove(IKLAN_DB_FILE)
                st.success("✅ Dihapus!"); st.rerun()
        
        display_iklan = df_iklan.copy()
        for col in NUMERIC_COLS:
            if col in display_iklan.columns: display_iklan[col] = display_iklan[col].apply(fmt_ribuan)
        if 'Tanggal' in display_iklan.columns:
            display_iklan['Tanggal'] = pd.to_datetime(display_iklan['Tanggal']).dt.strftime('%d/%m/%Y')
        
        st.caption(f"📊 Menampilkan {len(display_iklan)} dari {len(st.session_state.iklan_data)} baris")
        st.dataframe(display_iklan, use_container_width=True, height=500, hide_index=True)
        csv_data = st.session_state.iklan_data.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button("📥 Download CSV", csv_data, f"iklan_{datetime.now():%Y%m%d}.csv", "text/csv")
    else:
        st.info("📭 Belum ada data iklan.")

# =====================================
# 💳 UPLOAD PEMBAYARAN (INCOME + FILTER LENGKAP)
# =====================================
elif active_menu == "💳 Upload Pembayaran":
    st.header("💳 Upload Pembayaran (Income Shopee)")
    st.caption("Upload file Excel Income Statement dari Shopee")
    
    if 'income_data' not in st.session_state:
        if os.path.exists(INCOME_DB_FILE):
            st.session_state.income_data = pd.read_csv(INCOME_DB_FILE)
        else:
            st.session_state.income_data = pd.DataFrame()
    
    if 'processed_income_files' not in st.session_state:
        st.session_state.processed_income_files = set()
    
    if 'income_uploader_key' not in st.session_state:
        st.session_state.income_uploader_key = 0
    
    uploaded_files = st.file_uploader(
        "Pilih file Excel Income Shopee",
        type=["xlsx", "xls"],
        key=f"income_upload_{st.session_state.income_uploader_key}",
        accept_multiple_files=True
    )
    
    if uploaded_files:
        for f in uploaded_files:
            if f.name not in st.session_state.processed_income_files:
                with st.spinner(f"⏳ Memproses {f.name}..."):
                    df_income, message = process_income_file(f.read(), f.name)
                    
                    if df_income.empty:
                        st.error(f"❌ {f.name}: {message}")
                        continue
                    
                    if not st.session_state.income_data.empty and 'No_Pesanan' in st.session_state.income_data.columns:
                        existing_ids = st.session_state.income_data['No_Pesanan'].astype(str).tolist()
                        new_data = df_income[~df_income['No_Pesanan'].astype(str).isin(existing_ids)]
                    else:
                        new_data = df_income
                    
                    if not new_data.empty:
                        st.session_state.income_data = pd.concat([st.session_state.income_data, new_data], ignore_index=True)
                        st.session_state.income_data.to_csv(INCOME_DB_FILE, index=False)
                        st.success(f"✅ {f.name}: {len(new_data)} transaksi baru")
                    else:
                        st.info(f"📭 {f.name}: Semua data sudah ada")
                    
                    st.session_state.processed_income_files.add(f.name)
        
        st.session_state.income_uploader_key += 1
        st.rerun()
    
    st.markdown("---")
    st.subheader("📊 Data Pembayaran Tersimpan")
    
    if not st.session_state.income_data.empty:
        df_show = st.session_state.income_data.copy()
        
        if 'Tanggal_Dana' in df_show.columns:
            df_show['Tanggal_Dana'] = pd.to_datetime(df_show['Tanggal_Dana'], errors='coerce')
        
        # Filter Toko + Tanggal + Cari
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            if 'Nama_Toko' in df_show.columns:
                t_list = ['Semua'] + sorted(df_show['Nama_Toko'].dropna().unique().tolist())
                t_pilih = st.selectbox("🏪 Toko", t_list, key="filter_toko_income")
                if t_pilih != 'Semua':
                    df_show = df_show[df_show['Nama_Toko'].astype(str).str.strip() == t_pilih]
        with col_f2:
            if 'Tanggal_Dana' in df_show.columns:
                tgl_min = df_show['Tanggal_Dana'].min().date()
                tgl_max = df_show['Tanggal_Dana'].max().date()
                today = datetime.now().date()
                range_tgl = st.date_input("📅 Tanggal", value=(today, today), key="filter_tgl_income")
                if len(range_tgl) == 2:
                    df_show = df_show[(df_show['Tanggal_Dana'].dt.date >= range_tgl[0]) & (df_show['Tanggal_Dana'].dt.date <= range_tgl[1])]
        with col_f3:
            cari = st.text_input("🔎 Cari No. Pesanan", key="cari_income")
            if cari and 'No_Pesanan' in df_show.columns:
                df_show = df_show[df_show['No_Pesanan'].astype(str).str.contains(cari, case=False, na=False)]
        
        if 'Tanggal_Dana' in df_show.columns:
            df_show['Tanggal_Dana'] = df_show['Tanggal_Dana'].dt.strftime('%d/%m/%Y')
        
        st.caption(f"📊 Menampilkan {len(df_show)} dari {len(st.session_state.income_data)} transaksi")
        
        kolom_prioritas = ['Nama_Toko', 'No_Pesanan', 'Tanggal_Dana', 'Total_Dibayar', 'Harga_Asli', 'Biaya_Admin', 'Biaya_Layanan']
        kolom_tampil = [c for c in kolom_prioritas if c in df_show.columns]
        
        st.dataframe(df_show[kolom_tampil], use_container_width=True, height=400, hide_index=True)
        
        if 'Total_Dibayar' in st.session_state.income_data.columns:
            total = pd.to_numeric(st.session_state.income_data['Total_Dibayar'], errors='coerce').sum()
            st.metric("💰 Total Pembayaran", format_rupiah(int(total)))
        
        csv_data = st.session_state.income_data.to_csv(index=False).encode('utf-8-sig')
        st.download_button("📥 Download CSV", csv_data, f"income_{datetime.now():%Y%m%d}.csv", "text/csv", use_container_width=True)
        
        if st.button("🗑️ Hapus Semua Data", type="secondary"):
            st.session_state.income_data = pd.DataFrame()
            st.session_state.processed_income_files = set()
            if os.path.exists(INCOME_DB_FILE): os.remove(INCOME_DB_FILE)
            st.success("✅ Data dihapus!"); st.rerun()
    else:
        st.info("📭 Belum ada data.")
# =====================================
# 💰 INPUT DANA (RAPI - DUA TAB)
# =====================================
elif active_menu == "💰 Input Dana":
    st.header("📊 Rekap Per Hari")
    
    subtab1, subtab2 = st.tabs(["💎 Cair (Sudah Dibayar)", "📦 Pending (Belum Dibayar)"])
    
    hpp_map = {}
    nama_map = {}
    if not st.session_state.master_produk.empty:
        for _, row in st.session_state.master_produk.iterrows():
            sku_key = str(row['SKU']).strip().upper()
            try:
                hrg_beli = int(float(row['Harga_Beli'])) if pd.notna(row['Harga_Beli']) else 0
            except:
                hrg_beli = 0
            try:
                packing = int(float(row['Packing'])) if pd.notna(row['Packing']) else 0
            except:
                packing = 0
            hpp_map[sku_key] = hrg_beli + packing
            nama_map[sku_key] = str(row['Nama_Produk'])

        # ============================================================
        # SUBTAB 1: PENDING
        # ============================================================
        with subtab1:
            st.caption("Data dari Upload Penjualan | Estimasi berdasarkan rata-rata Cair per SKU")
            
            if st.session_state.shopee_data.empty:
                st.info("📭 Belum ada data penjualan.")
            else:
                df_pending = st.session_state.shopee_data.copy()
                
                if 'Nomor Referensi SKU' in df_pending.columns:
                    df_pending['SKU'] = df_pending['Nomor Referensi SKU']
                if 'Jumlah Produk di Pesan' in df_pending.columns:
                    df_pending['Qty'] = pd.to_numeric(df_pending['Jumlah Produk di Pesan'], errors='coerce').fillna(0).astype(int)
                elif 'Jumlah' in df_pending.columns:
                    df_pending['Qty'] = pd.to_numeric(df_pending['Jumlah'], errors='coerce').fillna(0).astype(int)
                # Harga Jual dari Harga Awal
                if 'Harga Awal' in df_pending.columns:
                    df_pending['Harga_Jual'] = pd.to_numeric(df_pending['Harga Awal'], errors='coerce').fillna(0)
                elif 'Total Pembayaran' in df_pending.columns:
                    df_pending['Harga_Jual'] = pd.to_numeric(df_pending['Total Pembayaran'], errors='coerce').fillna(0)
                elif 'Dibayar Pembeli' in df_pending.columns:
                    df_pending['Harga_Jual'] = pd.to_numeric(df_pending['Dibayar Pembeli'], errors='coerce').fillna(0)
                if 'Waktu Pesanan Dibuat' in df_pending.columns:
                    df_pending['Tanggal'] = pd.to_datetime(df_pending['Waktu Pesanan Dibuat'])
                                
                col1, col2 = st.columns(2)
                with col1:
                    today = datetime.now().date()
                    range_tgl = st.date_input("📅 Tanggal", value=(today, today), key="pending_tgl")
                with col2:
                    daftar_toko = ['Semua Toko'] + sorted(df_pending['Nama Toko'].dropna().unique().tolist()) if 'Nama Toko' in df_pending.columns else ['Semua Toko']
                    toko_pilih = st.selectbox("🏪 Toko", daftar_toko, key="pending_toko")
                
                if len(range_tgl) == 2 and 'Tanggal' in df_pending.columns:
                    df_pending = df_pending[(df_pending['Tanggal'].dt.date >= range_tgl[0]) & (df_pending['Tanggal'].dt.date <= range_tgl[1])]
                if toko_pilih != 'Semua Toko' and 'Nama Toko' in df_pending.columns:
                    df_pending = df_pending[df_pending['Nama Toko'] == toko_pilih]
                
                # Filter Income (pakai Tanggal Dana)
                if 'income_data' in st.session_state and not st.session_state.income_data.empty:
                    if 'No_Pesanan' in st.session_state.income_data.columns and 'Tanggal_Dana' in st.session_state.income_data.columns:
                        df_inc = st.session_state.income_data.copy()
                        df_inc['Tanggal_Dana'] = pd.to_datetime(df_inc['Tanggal_Dana'], errors='coerce')
                        sudah_cair = df_inc[df_inc['Tanggal_Dana'].notna()]['No_Pesanan'].dropna().astype(str).str.strip().unique()
                        if len(sudah_cair) > 0:
                            col_pes = 'No Pesanan' if 'No Pesanan' in df_pending.columns else ('No. Pesanan' if 'No. Pesanan' in df_pending.columns else None)
                            if col_pes:
                                sebelum = len(df_pending)
                                df_pending[col_pes] = df_pending[col_pes].fillna('').astype(str).str.strip()
                                df_pending = df_pending[~df_pending[col_pes].isin(sudah_cair)]
                                st.caption(f"🔍 {sebelum - len(df_pending)} pesanan sudah cair, disembunyikan. Sisa: {len(df_pending)}")
                
                # Rasio Map
                rasio_map = {}
                if 'income_data' in st.session_state and not st.session_state.income_data.empty:
                    df_inc = st.session_state.income_data.copy()
                    if 'No_Pesanan' in df_inc.columns and 'Total_Dibayar' in df_inc.columns:
                        df_sp = st.session_state.shopee_data.copy()
                        if 'Nomor Referensi SKU' in df_sp.columns:
                            df_sp['SKU'] = df_sp['Nomor Referensi SKU']
                        if 'Dibayar Pembeli' in df_sp.columns:
                            df_sp['Harga_Jual'] = pd.to_numeric(df_sp['Dibayar Pembeli'], errors='coerce').fillna(0)
                        merged = df_inc.merge(df_sp[['No Pesanan', 'SKU', 'Harga_Jual']], left_on='No_Pesanan', right_on='No Pesanan', how='inner')
                        if not merged.empty:
                            merged['Rasio'] = merged['Total_Dibayar'] / merged['Harga_Jual'].replace(0, 1)
                            rasio_map = merged.groupby(merged['SKU'].str.upper())['Rasio'].mean().to_dict()
                
                rekap_pending = []
                if not df_pending.empty and 'SKU' in df_pending.columns:
                    for tgl in sorted(df_pending['Tanggal'].dt.date.unique()):
                        df_tgl = df_pending[df_pending['Tanggal'].dt.date == tgl]
                        for sku in df_tgl['SKU'].dropna().unique():
                            sku_key = str(sku).strip().upper()
                            if not sku_key: continue
                            df_sku = df_tgl[df_tgl['SKU'].astype(str).str.strip().str.upper() == sku_key]
                            hpp_satuan = hpp_map.get(sku_key, 0)
                            rasio = rasio_map.get(sku_key, 0.85)
                            toko = str(df_sku['Nama Toko'].iloc[0]) if 'Nama Toko' in df_sku.columns else ''
                            nama_produk = nama_map.get(sku_key, sku_key)
                            
                            # Loop per No Pesanan (1 invoice = 1 baris)
                            for no_pes in df_sku['No Pesanan'].dropna().astype(str).unique():
                                qty_inv = int(df_sku[df_sku['No Pesanan'].astype(str) == no_pes]['Qty'].sum()) if 'Qty' in df_sku.columns else 1
                                hpp_inv = hpp_satuan * qty_inv
                                hrg_inv = int(df_sku[df_sku['No Pesanan'].astype(str) == no_pes]['Harga_Jual'].sum()) if 'Harga_Jual' in df_sku.columns else 0
                                est_inv = int(hrg_inv * rasio)
                                
                                rekap_pending.append({
                                    'No Pesanan': no_pes,
                                    'Tanggal': tgl.strftime('%d/%m/%Y'),
                                    'Nama Toko': toko,
                                    'SKU': sku_key,
                                    'Nama Produk': nama_produk,
                                    'Qty': qty_inv,
                                    'HPP': hpp_inv,
                                    'Harga Jual': hrg_inv,
                                    'Estimasi Diterima': est_inv,
                                    'Rasio': f"{rasio*100:.0f}%"
                                })
                
                if rekap_pending:
                    df_rp = pd.DataFrame(rekap_pending)
                    st.session_state.df_rekap_pending = df_rp
                    
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("📦 Total Qty", format_angka(int(df_rp['Qty'].sum())))
                    c2.metric("💰 Total Harga Jual", format_rupiah(int(df_rp['Harga Jual'].sum())))
                    c3.metric("📊 Total HPP", format_rupiah(int(df_rp['HPP'].sum())))
                    c4.metric("💎 Estimasi Diterima", format_rupiah(int(df_rp['Estimasi Diterima'].sum())))
                    
                    st.markdown(f"### 🧾 Total Invoice: {len(df_rp)}")
                    
                    st.markdown("---")
                    df_display = df_rp.copy()
                    for col in ['HPP', 'Harga Jual', 'Estimasi Diterima']:
                        df_display[col] = df_display[col].apply(lambda x: f"Rp {int(x):,}".replace(',', '.'))
                    st.dataframe(df_display, use_container_width=True, height=400, hide_index=True)
                    st.download_button("📥 Download CSV", df_rp.to_csv(index=False).encode('utf-8-sig'), f"pending_{datetime.now():%Y%m%d}.csv", use_container_width=True)
                else:
                    st.info("📭 Tidak ada data")

        # ============================================================
        # SUBTAB 2: CAIR (dari Upload Pembayaran) - REVISI FINAL
        # ============================================================
        with subtab2:
            st.caption("Data dari Upload Pembayaran (Income Shopee)")
            
            if 'income_data' not in st.session_state or st.session_state.income_data.empty:
                st.info("📭 Belum ada data pembayaran. Upload di menu 💳 Upload Pembayaran.")
            else:
                df_cair = st.session_state.income_data.copy()
                
                # Gabung dengan shopee_data
                df_shopee = st.session_state.shopee_data.copy()
                if not df_shopee.empty:
                    if 'Nomor Referensi SKU' in df_shopee.columns:
                        df_shopee['SKU'] = df_shopee['Nomor Referensi SKU']
                    if 'Jumlah Produk di Pesan' in df_shopee.columns:
                        df_shopee['Qty'] = pd.to_numeric(df_shopee['Jumlah Produk di Pesan'], errors='coerce').fillna(0).astype(int)
                    elif 'Jumlah' in df_shopee.columns:
                        df_shopee['Qty'] = pd.to_numeric(df_shopee['Jumlah'], errors='coerce').fillna(0).astype(int)
                    if 'Waktu Pesanan Dibuat' in df_shopee.columns:
                        df_shopee['Tanggal'] = pd.to_datetime(df_shopee['Waktu Pesanan Dibuat'])
                    if 'No Pesanan' in df_shopee.columns and 'No_Pesanan' in df_cair.columns:
                        df_cair = df_cair.merge(df_shopee[['No Pesanan', 'SKU', 'Qty', 'Tanggal']],
                                               left_on='No_Pesanan', right_on='No Pesanan', how='left')
                
                if 'Tanggal' not in df_cair.columns:
                    df_cair['Tanggal'] = pd.to_datetime(df_cair.get('Tanggal_Dana', datetime.now()))
                df_cair['Tanggal'] = pd.to_datetime(df_cair['Tanggal'], errors='coerce')
                
                # Range tanggal
                tgl_min_p = df_cair['Tanggal'].min().date() if not df_cair.empty else datetime.now().date()
                tgl_max_p = df_cair['Tanggal'].max().date() if not df_cair.empty else datetime.now().date()
                
                tgl_min_i, tgl_max_i = None, None
                if 'iklan_data' in st.session_state and not st.session_state.iklan_data.empty:
                    df_i = st.session_state.iklan_data.copy()
                    if 'Tanggal' in df_i.columns:
                        df_i['Tanggal'] = pd.to_datetime(df_i['Tanggal'], errors='coerce').dropna()
                        if not df_i.empty:
                            tgl_min_i = df_i['Tanggal'].min().date()
                            tgl_max_i = df_i['Tanggal'].max().date()
                
                tgl_min = min(tgl_min_p, tgl_min_i) if tgl_min_i else tgl_min_p
                tgl_max = max(tgl_max_p, tgl_max_i) if tgl_max_i else tgl_max_p
                
                col1, col2 = st.columns(2)
                with col1:
                    today = datetime.now().date()
                    range_tgl2 = st.date_input("📅 Tanggal", value=(today, today), key="cair_tgl")
                with col2:
                    daftar_toko2 = ['Semua Toko']
                    if 'iklan_data' in st.session_state and not st.session_state.iklan_data.empty:
                        if 'Nama Toko' in st.session_state.iklan_data.columns:
                            daftar_toko2 += sorted(st.session_state.iklan_data['Nama Toko'].dropna().unique().tolist())
                    if len(daftar_toko2) == 1 and 'Nama_Toko' in df_cair.columns:
                        daftar_toko2 += sorted(df_cair['Nama_Toko'].dropna().unique().tolist())
                    toko_pilih2 = st.selectbox("🏪 Toko", daftar_toko2, key="cair_toko")
                
                if len(range_tgl2) == 2:
                    tgl_awal2, tgl_akhir2 = range_tgl2
                else:
                    tgl_awal2, tgl_akhir2 = tgl_min, tgl_max
                
                # Filter penjualan (TANGGAL + TOKO)
                df_cair = df_cair[(df_cair['Tanggal'].dt.date >= tgl_awal2) & (df_cair['Tanggal'].dt.date <= tgl_akhir2)]
                
                # Filter toko untuk Income juga
                if toko_pilih2 != 'Semua Toko' and 'Nama_Toko' in df_cair.columns:
                    # Mapping nama toko dari Income ke format yang sama dengan dropdown
                    toko_map = {
                        'primaintiperkakas': 'PIP Shopee',
                        'pip': 'PIP Shopee',
                        'intiprimaperkakas': 'IPP Shopee',
                        'ipp': 'IPP Shopee',
                        'karyalogam': 'KL Shopee',
                        'kl': 'KL Shopee',
                        'nusantaratool': 'NT Shopee',
                        'nt': 'NT Shopee',
                        'drc': 'DRC Shopee',
                    }
                    # Buat kolom sementara untuk matching
                    df_cair['Toko_Match'] = df_cair['Nama_Toko'].astype(str).str.lower().str.replace(' ', '').str.replace('_', '')
                    df_cair['Toko_Match'] = df_cair['Toko_Match'].map(toko_map).fillna(df_cair['Nama_Toko'])
                    df_cair = df_cair[df_cair['Toko_Match'].astype(str).str.strip() == toko_pilih2.strip()]
                
                # Iklan per tanggal
                iklan_per_tgl = {}
                if not st.session_state.iklan_data.empty and 'Tanggal' in st.session_state.iklan_data.columns and 'Biaya' in st.session_state.iklan_data.columns:
                    df_iklan = st.session_state.iklan_data.copy()
                    df_iklan['Tanggal'] = pd.to_datetime(df_iklan['Tanggal'], errors='coerce')
                    df_iklan['Biaya'] = pd.to_numeric(df_iklan['Biaya'], errors='coerce').fillna(0)
                    df_iklan = df_iklan[(df_iklan['Tanggal'].dt.date >= tgl_awal2) & (df_iklan['Tanggal'].dt.date <= tgl_akhir2)]
                    if toko_pilih2 != 'Semua Toko' and 'Nama Toko' in df_iklan.columns:
                        df_iklan = df_iklan[df_iklan['Nama Toko'].astype(str).str.strip() == toko_pilih2.strip()]
                    if not df_iklan.empty:
                        iklan_grouped = df_iklan.groupby(df_iklan['Tanggal'].dt.date)['Biaya'].sum()
                        iklan_per_tgl = iklan_grouped.to_dict()
                
                # Proses rekap cair - DIHITUNG PER NO PESANAN DULU, BARU DIBAGI PER SKU
                rekap_cair = []
                if not df_cair.empty and 'SKU' in df_cair.columns and 'No_Pesanan' in df_cair.columns:
                    
                    # Step 1: Hitung Total_Dibayar per No_Pesanan
                    pesanan_total = df_cair.groupby('No_Pesanan').agg({
                        'Total_Dibayar': 'first',
                        'Tanggal': 'first',
                        'Nama_Toko': 'first',
                        'Qty': 'sum'
                    }).reset_index()
                    
                    for _, pesanan in pesanan_total.iterrows():
                        no_pes = pesanan['No_Pesanan']
                        total_bayar = int(pesanan['Total_Dibayar']) if pd.notna(pesanan['Total_Dibayar']) else 0
                        tgl = pesanan['Tanggal'].date() if pd.notna(pesanan['Tanggal']) else datetime.now().date()
                        toko = str(pesanan['Nama_Toko']) if pd.notna(pesanan['Nama_Toko']) else ''
                        total_qty_pesanan = int(pesanan['Qty']) if 'Qty' in pesanan_total.columns and pd.notna(pesanan['Qty']) else 1
                        
                        # Ambil semua SKU dari No_Pesanan ini
                        df_pes = df_cair[df_cair['No_Pesanan'] == no_pes]
                        
                        for _, row_sku in df_pes.iterrows():
                            sku = str(row_sku['SKU']).strip().upper() if pd.notna(row_sku['SKU']) else ''
                            if not sku:
                                continue
                            qty = int(row_sku['Qty']) if 'Qty' in df_pes.columns and pd.notna(row_sku['Qty']) else 1
                            
                            # Bagi rata ke semua SKU dalam 1 pesanan
                            jml_sku = len(df_pes)
                            if jml_sku > 0:
                                diterima_proporsional = int(total_bayar / jml_sku)
                                # Sisa pembagian diberikan ke SKU pertama
                                if sku == df_pes.iloc[0]['SKU'].strip().upper():
                                    diterima_proporsional += total_bayar % jml_sku
                            else:
                                diterima_proporsional = total_bayar
                            
                            hpp_satuan = hpp_map.get(sku, 0)
                            total_hpp = hpp_satuan * qty
                            nama_produk = nama_map.get(sku, sku)
                            
                            rekap_cair.append({
                                'No_Pesanan': no_pes,
                                'Tanggal': tgl.strftime('%d/%m/%Y'),
                                'Nama Toko': toko,
                                'SKU': sku,
                                'Nama Produk': nama_produk,
                                'Qty': qty,
                                'HPP': total_hpp,
                                'Diterima': diterima_proporsional,
                                'Estimasi Laba': diterima_proporsional - total_hpp
                            })
                
                # Daftar toko
                daftar_toko_ada = []
                if toko_pilih2 != 'Semua Toko':
                    daftar_toko_ada = [toko_pilih2]
                else:
                    # Gabung daftar toko dari iklan DAN income
                    daftar_toko_ada = []
                    if not st.session_state.iklan_data.empty and 'Nama Toko' in st.session_state.iklan_data.columns:
                        daftar_toko_ada += st.session_state.iklan_data['Nama Toko'].dropna().unique().tolist()
                    if not df_cair.empty and 'Nama_Toko' in df_cair.columns:
                        daftar_toko_ada += df_cair['Nama_Toko'].dropna().unique().tolist()
                    daftar_toko_ada = sorted(set(daftar_toko_ada))
                    if not daftar_toko_ada:
                        daftar_toko_ada = ['-']
                
                # Summary semua tanggal
                from datetime import timedelta
                semua_tgl = []
                cur = tgl_awal2
                while cur <= tgl_akhir2:
                    semua_tgl.append(cur)
                    cur += timedelta(days=1)
                
                df_rekap = pd.DataFrame(rekap_cair) if rekap_cair else pd.DataFrame()
                summary_list = []
                
                for tgl in semua_tgl:
                    tgl_str = tgl.strftime('%d/%m/%Y')
                    biaya_tgl = int(iklan_per_tgl.get(tgl, 0))
                    mask = df_rekap['Tanggal'] == tgl_str if not df_rekap.empty else pd.Series([False])
                    
                    # Kumpulkan toko yang sudah dapat jatah dari penjualan
                    toko_dari_penjualan = set()
                    
                    if mask.any():
                        df_t = df_rekap[mask]
                        for tok in df_t['Nama Toko'].unique():
                            df_tok = df_t[df_t['Nama Toko'] == tok]
                            summary_list.append({
                                'Tanggal': tgl_str, 'Nama Toko': tok,
                                'Qty': int(df_tok['Qty'].sum()), 'Diterima': int(df_tok['Diterima'].sum()),
                                'HPP': int(df_tok['HPP'].sum()), 'Biaya Iklan': 0, 'Status': '✅ Penjualan'
                            })
                            toko_dari_penjualan.add(tok)
                    
                    # Kalau ada biaya iklan, pastikan SEMUA toko dapat jatah
                    if biaya_tgl > 0:
                        if not daftar_toko_ada or daftar_toko_ada == ['-']:
                            # Fallback: tidak ada daftar toko
                            if not toko_dari_penjualan:
                                summary_list.append({
                                    'Tanggal': tgl_str, 'Nama Toko': 'Semua Toko',
                                    'Qty': 0, 'Diterima': 0, 'HPP': 0,
                                    'Biaya Iklan': 0, 'Status': '🔴 RUGI (Iklan)'
                                })
                        else:
                            # Tambahkan toko yang BELUM ada di summary_list
                            for tok in daftar_toko_ada:
                                if tok not in toko_dari_penjualan:
                                    summary_list.append({
                                        'Tanggal': tgl_str, 'Nama Toko': tok,
                                        'Qty': 0, 'Diterima': 0, 'HPP': 0,
                                        'Biaya Iklan': 0, 'Status': '🔴 RUGI (Iklan)'
                                    })
                
                if not summary_list:
                    # Buat DataFrame kosong dengan struktur lengkap
                    df_sum = pd.DataFrame(columns=['Tanggal', 'Nama Toko', 'Qty', 'Diterima', 'HPP', 'Biaya Iklan', 'Laba Bersih', 'Status'])
                else:
                    df_sum = pd.DataFrame(summary_list)
                    
                    # ... (kode distribusi iklan tetap sama)
                    
                    # Distribusi iklan REAL dari data iklan per toko per tanggal
                    iklan_real_map = {}
                    if not st.session_state.iklan_data.empty and 'Tanggal' in st.session_state.iklan_data.columns and 'Biaya' in st.session_state.iklan_data.columns and 'Nama Toko' in st.session_state.iklan_data.columns:
                        df_iklan_real = st.session_state.iklan_data.copy()
                        df_iklan_real['Tanggal'] = pd.to_datetime(df_iklan_real['Tanggal'], errors='coerce')
                        df_iklan_real['Biaya'] = pd.to_numeric(df_iklan_real['Biaya'], errors='coerce').fillna(0)
                        df_iklan_real = df_iklan_real[(df_iklan_real['Tanggal'].dt.date >= tgl_awal2) & (df_iklan_real['Tanggal'].dt.date <= tgl_akhir2)]
                        for (tgl, toko), grp in df_iklan_real.groupby([df_iklan_real['Tanggal'].dt.date, 'Nama Toko']):
                            iklan_real_map[(tgl, str(toko).strip())] = int(grp['Biaya'].sum())
                    
                    # Terapkan biaya iklan real ke df_sum
                    for idx in df_sum.index:
                        tgl_str = df_sum.at[idx, 'Tanggal']
                        toko = str(df_sum.at[idx, 'Nama Toko']).strip()
                        try:
                            tgl_dt = datetime.strptime(tgl_str, '%d/%m/%Y').date()
                            df_sum.at[idx, 'Biaya Iklan'] = iklan_real_map.get((tgl_dt, toko), 0)
                        except:
                            pass
                    
                    df_sum['Laba Bersih'] = df_sum['Diterima'] - df_sum['HPP'] - df_sum['Biaya Iklan']
                    
                    # Tambah info invoice pending
                    df_sum['Pending'] = ''
                    if 'df_rekap_pending' in st.session_state and not st.session_state.df_rekap_pending.empty:
                        df_p = st.session_state.df_rekap_pending.copy()
                        if 'Tanggal' in df_p.columns and 'Nama Toko' in df_p.columns:
                            p_count = df_p.groupby(['Tanggal', 'Nama Toko']).size().reset_index(name='count')
                            for _, rp in p_count.iterrows():
                                mask = (df_sum['Tanggal'] == rp['Tanggal']) & (df_sum['Nama Toko'] == rp['Nama Toko'])
                                if mask.any():
                                    df_sum.loc[mask, 'Pending'] = f"📋 {rp['count']} inv"
                    
                    # Metrics
                    c1, c2, c3, c4, c5, c6 = st.columns(6)
                    c1.metric("📦 Total Qty", format_angka(int(df_sum['Qty'].sum())))
                    c2.metric("🛒 Transaksi", format_angka(len(df_sum[df_sum['Qty'] > 0])))
                    c3.metric("💎 Total Diterima", format_rupiah(int(df_sum['Diterima'].sum())))
                    c4.metric("📊 Total HPP", format_rupiah(int(df_sum['HPP'].sum())))
                    c5.metric("📢 Total Iklan", format_rupiah(int(df_sum['Biaya Iklan'].sum())))
                    c6.metric("💰 Laba Bersih", format_rupiah(int(df_sum['Laba Bersih'].sum())))
                                           
                # Ringkasan Per Tanggal
                st.markdown("---")
                st.subheader("📅 Ringkasan Per Tanggal")
                if not df_sum.empty:
                    disp = df_sum.copy()
                    disp['Pending'] = disp['Pending'].fillna('')
                    disp['Diterima'] = disp['Diterima'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                    disp['HPP'] = disp['HPP'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                    disp['Biaya Iklan'] = disp['Biaya Iklan'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                    disp['Laba Bersih'] = disp['Laba Bersih'].fillna(0).astype(int).apply(warna_laba)
                    st.dataframe(disp, use_container_width=True, height=300, hide_index=True)
                else:
                    st.dataframe(df_sum, use_container_width=True, height=150, hide_index=True)
                    st.caption("ℹ️ Tabel akan terisi saat ada data yang sesuai filter")               
                with st.expander("📊 Perkiraan Keuntungan", expanded=False):
                    # Estimasi & HPP dari shopee_data
                    df_est = st.session_state.shopee_data.copy()
                    
                    # Buang yang batal
                    if 'Status Pesanan' in df_est.columns:
                        status_batal = ['batal', 'dibatalkan', 'canceled', 'cancelled', 'Batal', 'Dibatalkan', 
                                       'CANCEL', 'Cancel', 'BATAL', 'DIBATALKAN', 'Pesanan Dibatalkan']
                        df_est = df_est[~df_est['Status Pesanan'].astype(str).str.strip().str.lower().isin([s.lower() for s in status_batal])]
                    
                    if 'Waktu Pesanan Dibuat' in df_est.columns:
                        df_est['Tanggal'] = pd.to_datetime(df_est['Waktu Pesanan Dibuat'], errors='coerce')
                        if len(range_tgl2) == 2:
                            df_est = df_est[(df_est['Tanggal'].dt.date >= range_tgl2[0]) & (df_est['Tanggal'].dt.date <= range_tgl2[1])]
                    if toko_pilih2 != 'Semua Toko' and 'Nama Toko' in df_est.columns:
                        df_est = df_est[df_est['Nama Toko'] == toko_pilih2]
                    
                    # Ambil SEMUA biaya iklan (termasuk produk tidak laku)
                    total_iklan_dict = {}
                    total_iklan_all = 0
                    if not st.session_state.iklan_data.empty and 'Tanggal' in st.session_state.iklan_data.columns and 'Biaya' in st.session_state.iklan_data.columns:
                        df_ik_est = st.session_state.iklan_data.copy()
                        df_ik_est['Tanggal'] = pd.to_datetime(df_ik_est['Tanggal'], errors='coerce')
                        df_ik_est = df_ik_est[(df_ik_est['Tanggal'].dt.date >= range_tgl2[0]) & (df_ik_est['Tanggal'].dt.date <= range_tgl2[1])]
                        if toko_pilih2 != 'Semua Toko' and 'Nama Toko' in df_ik_est.columns:
                            df_ik_est = df_ik_est[df_ik_est['Nama Toko'] == toko_pilih2]
                        total_iklan_all = int(df_ik_est['Biaya'].sum())
                        iklan_group = df_ik_est.groupby([df_ik_est['Tanggal'].dt.date, 'Nama Toko'])['Biaya'].sum()
                        for (tgl, toko), val in iklan_group.items():
                            total_iklan_dict[(tgl.strftime('%d/%m/%Y'), str(toko).strip())] = int(val)
                    
                    if not df_est.empty and 'Nomor Referensi SKU' in df_est.columns:
                        df_est['Tanggal'] = df_est['Tanggal'].dt.strftime('%d/%m/%Y')
                        if 'Jumlah Produk di Pesan' in df_est.columns:
                            df_est['Qty'] = pd.to_numeric(df_est['Jumlah Produk di Pesan'], errors='coerce').fillna(0).astype(int)
                        
                        df_est['HPP_Unit'] = df_est['Nomor Referensi SKU'].apply(lambda x: hpp_map.get(str(x).strip().upper(), 0))
                        df_est['HPP_Total'] = df_est['HPP_Unit'] * df_est['Qty'].fillna(1)
                        
                        if 'Harga Awal' in df_est.columns:
                            df_est['Harga_Awal_Num'] = pd.to_numeric(df_est['Harga Awal'], errors='coerce').fillna(0)
                            df_est['Harga_Awal_Num'] = df_est['Harga_Awal_Num'] * df_est['Qty'].fillna(1)
                            df_est['Harga'] = df_est['Harga_Awal_Num']
                        elif 'Total Pembayaran' in df_est.columns:
                            df_est['Harga'] = pd.to_numeric(df_est['Total Pembayaran'], errors='coerce').fillna(0)
                        else:
                            df_est['Harga'] = 0
                        
                        df_est['Estimasi'] = (df_est['Harga'] * 0.78).astype(int)
                        
                        est_group = df_est.groupby(['Tanggal', 'Nama Toko']).agg(
                            Estimasi=('Estimasi', 'sum'),
                            HPP=('HPP_Total', 'sum')
                        ).reset_index()
                        
                        # Gabungkan dengan data iklan (outer join)
                        if total_iklan_dict:
                            iklan_rows = []
                            for (tgl, toko), biaya in total_iklan_dict.items():
                                iklan_rows.append({'Tanggal': tgl, 'Nama Toko': toko, 'Iklan': biaya})
                            df_iklan_full = pd.DataFrame(iklan_rows)
                            est_group = pd.merge(est_group, df_iklan_full, on=['Tanggal', 'Nama Toko'], how='outer')
                            est_group[['Estimasi', 'HPP', 'Iklan']] = est_group[['Estimasi', 'HPP', 'Iklan']].fillna(0).astype(int)
                        else:
                            est_group['Iklan'] = 0
                        
                        est_group['Laba'] = est_group['Estimasi'] - est_group['HPP'] - est_group['Iklan']
                        est_group = est_group.sort_values('Estimasi', ascending=False).head(10)
                        
                        total_est_all = int(df_est['Estimasi'].sum())
                        total_hpp_all = int(df_est['HPP_Total'].sum())
                        total_laba_all = total_est_all - total_hpp_all - total_iklan_all
                        
                        st.caption(f"💰 Estimasi: {format_rupiah(total_est_all)} | 📊 HPP: {format_rupiah(total_hpp_all)} | 📢 Iklan: {format_rupiah(total_iklan_all)} | 💎 {'Laba' if total_laba_all >= 0 else 'Rugi'}: {format_rupiah(abs(total_laba_all))}")
                        
                        est_group['Estimasi'] = est_group['Estimasi'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                        est_group['HPP'] = est_group['HPP'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                        est_group['Iklan'] = est_group['Iklan'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                        est_group['Laba'] = est_group['Laba'].apply(warna_laba)
                        
                        st.dataframe(est_group[['Tanggal', 'Nama Toko', 'Estimasi', 'HPP', 'Iklan', 'Laba']], 
                                    use_container_width=True, height=300, hide_index=True)
                    else:
                        # Tidak ada penjualan, tapi tetap tampilkan biaya iklan
                        if total_iklan_all > 0:
                            st.caption(f"💰 Estimasi: Rp 0 | 📊 HPP: Rp 0 | 📢 Iklan: {format_rupiah(total_iklan_all)} | 💎 Rugi: {format_rupiah(total_iklan_all)}")
                        else:
                            st.caption("📭 Tidak ada data")

                    
                
                # Detail Per Produk
                st.markdown("---")
                st.subheader("📋 Detail Per Produk")
                
                df_det_all = st.session_state.shopee_data.copy()
                if 'Waktu Pesanan Dibuat' in df_det_all.columns:
                    df_det_all['Tanggal'] = pd.to_datetime(df_det_all['Waktu Pesanan Dibuat'], errors='coerce')
                    if len(range_tgl2) == 2:
                        df_det_all = df_det_all[(df_det_all['Tanggal'].dt.date >= range_tgl2[0]) & (df_det_all['Tanggal'].dt.date <= range_tgl2[1])]
                if toko_pilih2 != 'Semua Toko' and 'Nama Toko' in df_det_all.columns:
                    df_det_all = df_det_all[df_det_all['Nama Toko'] == toko_pilih2]
                
                if not df_det_all.empty and 'Nomor Referensi SKU' in df_det_all.columns:
                    df_det_all['Tanggal'] = df_det_all['Tanggal'].dt.strftime('%d/%m/%Y')
                    if 'Jumlah Produk di Pesan' in df_det_all.columns:
                        df_det_all['Qty'] = pd.to_numeric(df_det_all['Jumlah Produk di Pesan'], errors='coerce').fillna(0).astype(int)
                    if 'Harga Awal' in df_det_all.columns:
                        df_det_all['Harga_Jual'] = pd.to_numeric(df_det_all['Harga Awal'], errors='coerce').fillna(0)
                    elif 'Total Pembayaran' in df_det_all.columns:
                        df_det_all['Harga_Jual'] = pd.to_numeric(df_det_all['Total Pembayaran'], errors='coerce').fillna(0)
                    
                    det_display = df_det_all.groupby(['Tanggal', 'Nama Toko', 'Nomor Referensi SKU', 'Nama Produk', 'Status Pesanan']).agg(
                        Qty=('Qty', 'sum'),
                        Harga=('Harga_Jual', 'sum')
                    ).reset_index()
                    
                    if not det_display.empty:
                        total_inv_det = df_det_all['No Pesanan'].nunique() if 'No Pesanan' in df_det_all.columns else len(det_display)
                        total_qty_det = int(det_display['Qty'].sum())
                        total_hrg_det = int(det_display['Harga'].sum())
                        st.caption(f"💰 Total: {format_rupiah(total_hrg_det)} | 📦 Qty: {total_qty_det} | 🧾 Invoice: {total_inv_det}")
                        
                        det_display['Harga'] = det_display['Harga'].apply(lambda x: f"Rp {int(x):,}".replace(',','.'))
                        st.dataframe(det_display, use_container_width=True, height=400, hide_index=True)
                    else:
                        st.caption("📭 Tidak ada data")
                else:
                    st.caption("📭 Tidak ada data di tanggal ini")

# =====================================
# 📢 REKAP IKLAN 
# =====================================
elif active_menu == "📢 Rekap Iklan":
    st.header("📢 Rekap Iklan")
    st.caption("Pantau efektivitas iklan | Target ROAS & Diskon editable")
    
    if st.session_state.iklan_data.empty:
        st.info("📭 Belum ada data iklan.")
    else:
        df = st.session_state.iklan_data.copy()
        col_nama = 'Nama Iklan' if 'Nama Iklan' in df.columns else None
        col_kode = 'Kode Produk' if 'Kode Produk' in df.columns else None
        col_terjual = 'Produk Terjual' if 'Produk Terjual' in df.columns else None
        col_omzet = 'Omzet Penjualan' if 'Omzet Penjualan' in df.columns else None
        col_biaya = 'Biaya' if 'Biaya' in df.columns else None
        col_toko = 'Nama Toko' if 'Nama Toko' in df.columns else None
        col_tgl = 'Tanggal' if 'Tanggal' in df.columns else None
        
        for c in [col_terjual, col_omzet, col_biaya]:
            if c: df[c] = pd.to_numeric(df[c].astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce').fillna(0).astype(int)
        if col_tgl: df[col_tgl] = pd.to_datetime(df[col_tgl], errors='coerce')
        
        hpp_map = {}
        harga_jual_map = {}
        if not st.session_state.master_produk.empty:
            for _, row in st.session_state.master_produk.iterrows():
                sk = str(row['SKU']).strip().upper()
                try:
                    hrg_beli = int(float(row['Harga_Beli'])) if pd.notna(row['Harga_Beli']) else 0
                except:
                    hrg_beli = 0
                try:
                    packing = int(float(row['Packing'])) if pd.notna(row['Packing']) else 0
                except:
                    packing = 0
                hpp_map[sk] = hrg_beli + packing
                if 'Harga_Jual_Default' in row and pd.notna(row['Harga_Jual_Default']):
                    try:
                        harga_jual_map[sk] = int(float(row['Harga_Jual_Default']))
                    except:
                        pass
        
        st.caption(f"DEBUG: 07-470 hpp = {hpp_map.get('07-470', 'TIDAK ADA')}")
        st.caption(f"DEBUG: TAFF-30ML hpp = {hpp_map.get('TAFF-30ML', 'TIDAK ADA')}")
        
        harga_map = {}
        if not st.session_state.shopee_data.empty:
            df_sp = st.session_state.shopee_data.copy()
            col_sku_sp = 'Nomor Referensi SKU' if 'Nomor Referensi SKU' in df_sp.columns else ('SKU' if 'SKU' in df_sp.columns else None)
            col_harga = 'Harga Awal' if 'Harga Awal' in df_sp.columns else None
            if col_sku_sp and col_harga:
                df_sp[col_harga] = pd.to_numeric(df_sp[col_harga], errors='coerce').fillna(0)
                if 'Waktu Pesanan Dibuat' in df_sp.columns:
                    df_sp = df_sp.sort_values('Waktu Pesanan Dibuat', ascending=False)
                if 'Nama Toko' in df_sp.columns:
                    for (toko, sku), grp in df_sp.groupby(['Nama Toko', col_sku_sp]):
                        sk_key = str(sku).strip().upper()
                        toko_key = str(toko).strip()
                        if sk_key and sk_key != 'NAN':
                            vals = grp[col_harga].dropna()
                            vals = vals[vals > 1000]
                            if len(vals) > 0:
                                harga_map[(toko_key, sk_key)] = int(vals.iloc[0])
                for sk, grp in df_sp.groupby(col_sku_sp):
                    sk_key = str(sk).strip().upper()
                    if sk_key and sk_key != 'NAN':
                        vals = grp[col_harga].dropna()
                        vals = vals[vals > 1000]
                        if len(vals) > 0:
                            if sk_key not in harga_map:
                                harga_map[sk_key] = int(vals.iloc[0])
        
        # Load harga per toko dari file edit cepat
        harga_per_toko = {}
        harga_toko_file = str(DB_FOLDER / "harga_per_toko.csv")
        if os.path.exists(harga_toko_file):
            df_ht = pd.read_csv(harga_toko_file)
            for _, row in df_ht.iterrows():
                harga_per_toko[(str(row['Nama_Toko']).strip(), str(row['SKU']).strip().upper())] = int(row['Harga_Jual'])
                
                # Juga simpan harga umum (fallback)
                for sk, grp in df_sp.groupby(col_sku_sp):
                    sk_key = str(sk).strip().upper()
                    if sk_key and sk_key != 'NAN':
                        vals = grp[col_harga].dropna()
                        vals = vals[vals > 1000]
                        if len(vals) > 0:
                            if sk_key not in harga_map:
                                harga_map[sk_key] = int(vals.iloc[0])
                else:
                    for sk, grp in df_sp.groupby(col_sku_sp):
                        sk_key = str(sk).strip().upper()
                        if sk_key and sk_key != 'NAN':
                            vals = grp[col_harga].dropna()
                            vals = vals[vals > 1000]
                            if len(vals) > 0:
                                harga_map[sk_key] = int(vals.iloc[0])
                     
        kode_to_sku = {
            # ==========================================
            # PIP SHOPEE (PRIMA INTI PERKAKAS)
            # ==========================================
            # SLING 1 TON
            '45900299218': 'Sling1x2M','52850279265': 'Sling1x4M','28943418785': 'Sling1x5M',
            '51150284056': 'Sling1x6M','24832156936': 'Sling1x3M',

            # SLING 2 TON 
            '43573843187': 'Sling2x2M','40323838649': 'Sling2x3M','42023851912': 'Sling2x4M',
            '42523843023': 'Sling2x5M','26009177149': 'Sling2x6M',

            # SLING 3 TON
            '54100480246': 'Sling3x2M','54350475570': 'Sling3x3M','26359737303': 'Sling3x4M',                    '46500490852': 'Sling3x5M',

            # SLING 4 TON
            '40673762846': 'Sling4x3M','28171008905': 'Sling4x4M','24947412361': 'Sling4x5M',
            '28142823646': 'Sling4x6M','29342814404': 'Sling4x8M',

            # SLING 5 TON
            '55200717478': 'Sling5x5M',
 
            # Selesai Cek
            '26643764706':'PIPE-1PC','11008281950': '59-135s','12684449528': '1108-150',
            '19494596228':'45-600','18522270317': 'MEIJER-10BAR','45102277297': 'EDL-2701',                     '9084726015':'07-470','47102879955': 'Velcro-8inch', '44857049532': 'TAFF-30ML',  
            '8433584840':'VELOZ-GG15INCH',

            # Belum selesai tunggu terjual                       
            # SLING 5 TON
            '23702055121': 'Sling5x3M','51350722513': 'Sling5x4M',
            '49500742481': 'Sling5x6M','46950732871': 'Sling5x8M',
            # TOOLS & LAINNYA                  
            '8035948643': 'AU-AC0984',

            # ==========================================
            # IPP SHOPEE (INTI PRIMA PERKAKAS)
            # ==========================================
            # SLING 2 TON
            '54450235614':'Sling2x2M','40325856989':'Sling2x3M','48150255860':'Sling2x4M',
            '27159003581':'Sling2x5M','55500245644':'Sling2x6M',

            # SLING 2 TON

            # Selesai Cek
            '54757033046':'TAFF-30ML','50552867054':'Velcro-8inch',


            '5290892246': 'Sling5x3M','27159003581': 'Sling2x1M','11815955337': 'Tekiro-KunciL',
            '16734400654': '1108-150','24685496037': 'Sling3x2M','51307036393': 'Ratchet2x10M',
            '9684731731': '07-470','41176083453': 'Sling5x6M','55500245644': 'Sling2x6M',
            '16000218285': 'AU-AC0984','28321013169':'Sling4x3M',
            '49750756242': 'Sling5x5M','27802700183': 'Sling1x2-6M',
            '48150255860': 'Sling2x4M','52352253086': 'EDL-2701','46050305301': 'Sling1x2M',
            '43725980576': 'Sling4x8M','45650524273': 'Sling4x6M','22084522116': 'Kapak-Kayu',
            '53507041417': 'Ratchet2x12M','52550285288': 'Sling1x6M','52250285268': 'Sling1x3M',
            '54300240730': 'Sling2x5M','43326093300': 'Sling5x3M','42276097656': 'Sling5x4M',
            '55800746031': 'Sling5x8M','54852840806': 'Extractor-Pipa','46450524354': 'Sling4x4M',
            '52450285198': 'Sling1x5M','49300295532': 'Sling1x4M',
            '29043466257': 'Sling4x5M','52150331673': 'Sling3x3M','52700336564': 'Sling3x5M',
            '56700499173': 'Sling4x3M_2','52600336532':'Sling3x4M',
            
            # ==========================================
            # KL SHOPEE (KARYA LOGAM)
            # ==========================================
            # SLING 1 TON
            '44850300466':'Sling1x2M','25432156514':'Sling1x3M','54510644376':'Sling1x4M',
            '54360655404':'Sling1x5M','49960680607':'Sling1x6M',

            # SLING 2 TON
            '41624901642':'Sling2x2M','27393140537':'Sling2x3M','55460661119':'Sling2x4M',
            '41374898484':'Sling2x5M','29009007996':'Sling2x6M',

            # SLING 3 TON
            '42480953565':'Sling3x2M','54710658807':'Sling3x3M','50660663898':'Sling3x4M',
            '29409746651':'Sling3x5M','57650365477':'Sling3x6M',

            # SLING 4 TON
            '45710679559':'Sling4x3M','56410659533':'Sling4x4M','48500590542':'Sling4x5M',
            '54650575545':'Sling4x6M','25989567618':'Sling4x8M',

            # SLING 5 TON
            '44560733186':'Sling5x3M','45610731941':'Sling5x4M','56910711960':'Sling5x5M',
            '52360712045':'Sling5x6M','41668166981':'Sling5x8M',

            # Ratchet 2 TON
            '49460797770':'Ratchet2x12M',

            # Selesai Cek 
            '27801159274':'45-600','51306573597':'TAFF-30ML','50602248478':'EDL-2701',
            '55601267942':'MEIJER-10BAR','22853747855':'07-470','19210625728':'59-135s',
            '43630983209':'59-135',
                      
            # Belum Cek
            '51102866921': 'Velcro-8inch',
            '47402871044': 'PIPE-1PC',
            
            # ==========================================
            # NT SHOPEE (NUSANTARA TOOL)
            # ==========================================
            # SLING 4 TON
            '47900589280':'Sling4x3M','52300554443':'Sling4x4M','41426007917':'Sling4x5M',
            '55150569291':'Sling4x6M','45700579435':'Sling4x8M',
            
            # Selesai Cek
            '14134122235': '07-470','27601159301':'45-600',
            # Belum Cek
            '17078874343': 'MEIJER-10BAR','27821013219': 'Sling4x3M',
            '54002257357': 'EDL-2701','54752846337': 'PIPE-1PC','54650354087': 'Sling3x3M',
            '47050374118': 'Sling3x4M','40377136956': 'Velcro-8inch','55550363711': 'Sling3x2M',
            '16134383758': '1108-150','55600727679': 'Sling5x4M','25885482122': 'Sling3x2M',
            '47150374263': 'Sling3x5M',
            
            # ==========================================
            # DRC SHOPEE
            # ==========================================
            # SLING 5 TON
            '26414173045': 'Sling5x3M','25379872796': 'Sling5x4M','26164260621': 'Sling5x5M',
            '52301094773': 'Sling5x6M','41621355363': 'Sling5x8M',
            # Belum Cek
            '52806767424': 'Ratchet3x8M','47306827500': 'Ratchet2x8M','47756788963': 'Ratchet5x12M',
            '46156800723': 'Ratchet5x15M','54056767324': 'Ratchet3x15M','45006827239': 'Ratchet2x10M',
            '48056818871': 'Ratchet2x12M','48256367225': 'Sling4x3M','9412670154': '07-470',
            '50802248482': 'EDL-2701',
            '11949481895': 'AU-AC0984',
            '52606176296': 'Webbing Sling 1 Ton 2–6 Meter','11745936543': 'MEIJER-10BAR','16834420649': '1108-150',
            '46502891636': 'Velcro-8inch','46702861402': 'PIPE-1PC','41678842328': 'Sling4x8M',
            '56906176131': 'Sling1x6M','57356180589': 'Sling1x2M','47606200919': 'Sling1x5M',
            '57956338634': 'Sling4x4M','52406343351': 'Sling4x6M','56956338672': 'Sling4x3M',
            '18106066820': 'Sling5x3M','22584521485': '45-600',
            '43129781011': 'Ratchet5x12M','46506196128': 'Sling1x4M','46808247877': 'Ratchet2x12M',
            '48058247682': 'Ratchet2x10M','48358285290': 'Ratchet3x12M','48658276071': 'Ratchet3x8M',
            '49906792855': 'Ratchet5x8M','50758256182': 'Ratchet3x15M','54358260787': 'Ratchet3x10M',
            '56008213846': 'Ratchet2x8M','56408265055': 'Ratchet3x8M','57856343398': 'Sling4x5M',
            '8574235086': 'HK-LS1200','9261805343': '59-135s','10727519333':'VELOZ-GG15INCH',
}
        
        c1, c2 = st.columns(2)
        with c1:
            today = datetime.now().date()
            r = st.date_input("📅 Tanggal", value=(today, today), key="iklan_tgl")
        with c2:
            toko_list = ['Semua'] + (sorted(df[col_toko].dropna().unique().tolist()) if col_toko else [])
            toko_pilih = st.selectbox("🏪 Toko", toko_list, key="iklan_toko")
    
        if r and len(r)==2 and col_tgl: df = df[(df[col_tgl].dt.date>=r[0])&(df[col_tgl].dt.date<=r[1])]
        if toko_pilih!='Semua' and col_toko: df = df[df[col_toko]==toko_pilih]
        
        # Input Rasio Manual
        rasio_manual = st.slider("🔧 Rasio Estimasi (Manual)", min_value=0.70, max_value=0.90, value=0.78, step=0.01, key="rasio_manual")
        
        if df.empty:
            st.info("📭 Tidak ada data.")
        else:
            df[col_nama] = df[col_nama].astype(str).str.strip().str.replace('\s+', ' ', regex=True)
            df = df[df[col_kode].notna() & (df[col_kode].astype(str).str.strip() != '')]
            
            # === AMBIL DATA SEMUA (KECUALI BATAL) ===
            df_shopee_semua = pd.DataFrame()
            if not st.session_state.shopee_data.empty:
                df_shopee_semua = st.session_state.shopee_data.copy()
                if 'Status Pesanan' in df_shopee_semua.columns:
                    status_batal = ['batal', 'dibatalkan', 'canceled', 'cancelled', 'Batal', 'Dibatalkan', 
                                   'CANCEL', 'Cancel', 'BATAL', 'DIBATALKAN', 'Pesanan Dibatalkan']
                    df_shopee_semua = df_shopee_semua[~df_shopee_semua['Status Pesanan'].astype(str).str.strip().str.lower().isin([s.lower() for s in status_batal])]
                if 'Waktu Pesanan Dibuat' in df_shopee_semua.columns:
                    df_shopee_semua['Tanggal'] = pd.to_datetime(df_shopee_semua['Waktu Pesanan Dibuat'], errors='coerce')
                if 'Nomor Referensi SKU' in df_shopee_semua.columns:
                    df_shopee_semua['SKU'] = df_shopee_semua['Nomor Referensi SKU']
                if 'Jumlah Produk di Pesan' in df_shopee_semua.columns:
                    df_shopee_semua['Qty'] = pd.to_numeric(df_shopee_semua['Jumlah Produk di Pesan'], errors='coerce').fillna(0).astype(int)
                if 'Total Pembayaran' in df_shopee_semua.columns:
                    df_shopee_semua['Omzet_Real'] = pd.to_numeric(df_shopee_semua['Total Pembayaran'], errors='coerce').fillna(0)
                
                if col_toko and 'Nama Toko' in df_shopee_semua.columns:
                    if toko_pilih != 'Semua':
                        df_shopee_semua = df_shopee_semua[df_shopee_semua['Nama Toko'].astype(str).str.strip() == toko_pilih]
                if col_tgl and 'Tanggal' in df_shopee_semua.columns:
                    if r and len(r) == 2:
                        df_shopee_semua = df_shopee_semua[(df_shopee_semua['Tanggal'].dt.date >= r[0]) & (df_shopee_semua['Tanggal'].dt.date <= r[1])]
            
            semua_map = {}
            if not df_shopee_semua.empty and 'SKU' in df_shopee_semua.columns:
                if 'Nama Toko' in df_shopee_semua.columns:
                    for (toko, sku), grp in df_shopee_semua.groupby(['Nama Toko', 'SKU']):
                        sku_key = str(sku).strip().upper()
                        toko_key = str(toko).strip()
                        if sku_key and sku_key != 'NAN':
                            qty_real = int(grp['Qty'].sum()) if 'Qty' in grp.columns else len(grp)
                            omzet_real = int(grp['Omzet_Real'].sum()) if 'Omzet_Real' in grp.columns else 0
                            semua_map[(toko_key, sku_key)] = (qty_real, omzet_real)
                else:
                    for sku, grp in df_shopee_semua.groupby('SKU'):
                        sku_key = str(sku).strip().upper()
                        if sku_key and sku_key != 'NAN':
                            qty_real = int(grp['Qty'].sum()) if 'Qty' in grp.columns else len(grp)
                            omzet_real = int(grp['Omzet_Real'].sum()) if 'Omzet_Real' in grp.columns else 0
                            semua_map[sku_key] = (qty_real, omzet_real)
            
            group_cols = [col_kode, col_nama]
            if col_toko:
                group_cols.append(col_toko)
            
            df_grouped = df.groupby(group_cols).agg({
                col_terjual: 'sum', 
                col_omzet: 'sum', 
                col_biaya: 'sum'
            }).reset_index()
            
            if 'target_roas_data' not in st.session_state:
                if os.path.exists(str(DB_FOLDER / "target_roas.csv")):
                    df_tr = pd.read_csv(str(DB_FOLDER / "target_roas.csv"))
                    st.session_state.target_roas_data = dict(zip(df_tr['key'], df_tr['target']))
                else:
                    st.session_state.target_roas_data = {}
            if 'diskon_global' not in st.session_state: st.session_state.diskon_global = 0
            
            rasio_map_iklan = {}
            if 'income_data' in st.session_state and not st.session_state.income_data.empty:
                df_inc_rasio = st.session_state.income_data.copy()
                if 'No_Pesanan' in df_inc_rasio.columns and 'Total_Dibayar' in df_inc_rasio.columns:
                    df_sp_rasio = st.session_state.shopee_data.copy()
                    if 'Nomor Referensi SKU' in df_sp_rasio.columns:
                        df_sp_rasio['SKU'] = df_sp_rasio['Nomor Referensi SKU']
                    if 'Total Pembayaran' in df_sp_rasio.columns:
                        df_sp_rasio['Harga_Jual'] = pd.to_numeric(df_sp_rasio['Total Pembayaran'], errors='coerce').fillna(0)
                    if 'No Pesanan' in df_sp_rasio.columns and 'SKU' in df_sp_rasio.columns:
                        merged_rasio = df_inc_rasio.merge(df_sp_rasio[['No Pesanan', 'SKU', 'Harga_Jual']], 
                                                          left_on='No_Pesanan', right_on='No Pesanan', how='inner')
                        if not merged_rasio.empty:
                            merged_rasio['Rasio'] = merged_rasio['Total_Dibayar'] / merged_rasio['Harga_Jual'].replace(0, 1)
                            rasio_grouped = merged_rasio.groupby(merged_rasio['SKU'].str.upper())['Rasio'].mean()
                            rasio_map_iklan = rasio_grouped.to_dict()
            
                       
            hasil = []
            for _, row in df_grouped.iterrows():
                nama = str(row[col_nama])
                kode = str(row[col_kode])
                toko_iklan = str(row[col_toko]) if col_toko else ''
                
                sku = kode_to_sku.get(kode, '-')
                
                # === DATA IKLAN MENTAH ===
                terjual_iklan = int(row[col_terjual]) if col_terjual else 0
                omzet_iklan = int(row[col_omzet]) if col_omzet else 0
                biaya = int(row[col_biaya]) if col_biaya else 0
                
                # === AMBIL TERJUAL & OMZET PER TOKO (FALLBACK KE DATA IKLAN) ===
                if sku != '-':
                    real_data = semua_map.get((toko_iklan, sku.upper()), (0, 0))
                    if real_data[0] == 0:
                        real_data = semua_map.get(sku.upper(), (0, 0))
                    terjual_dari_iklan = real_data[0] if real_data[0] > 0 else terjual_iklan
                    omzet_dari_iklan = real_data[1] if real_data[1] > 0 else omzet_iklan
                else:
                    terjual_dari_iklan = 0
                    omzet_dari_iklan = 0
                
                # HPP pakai FIFO (otomatis update kalau harga beli berubah)
                if sku != '-':
                    hpp_fifo_val = get_hpp_fifo(sku.upper())
                    hpp_satuan = hpp_fifo_val if hpp_fifo_val > 0 else hpp_map.get(sku.upper(), 0)
                else:
                    hpp_satuan = 0
                
                # Harga jual satuan (REAL) - per toko
                hj = 0
                if sku != '-':
                    if toko_iklan:
                        hj = harga_per_toko.get((toko_iklan.strip(), sku.upper()), 0)
                    if hj == 0:
                        hj = harga_jual_map.get(sku.upper(), 0)
                    if hj == 0 and toko_iklan:
                        hj = harga_map.get((toko_iklan, sku.upper()), 0)
                    if hj == 0:
                        hj = harga_map.get(sku.upper(), 0)
                if hj == 0 and terjual_dari_iklan > 0 and omzet_dari_iklan > 0:
                    hj = omzet_dari_iklan // terjual_dari_iklan
                
                # === ESTIMASI DITERIMA PAKAI RASIO MANUAL (LEBIH AMAN) ===
                estimasi_pc = hj
                if sku != '-' and hj > 0:
                    # Pakai rasio manual dari slider (default 0.78)
                    estimasi_pc = int(hj * rasio_manual)
                
                # === Omzet pakai Estimasi ===
                if estimasi_pc > 0 and terjual_dari_iklan > 0:
                    omzet_hitung = estimasi_pc * terjual_dari_iklan
                elif hj > 0 and terjual_dari_iklan > 0:
                    omzet_hitung = hj * terjual_dari_iklan
                else:
                    omzet_hitung = omzet_dari_iklan
                
                up = estimasi_pc - hpp_satuan if estimasi_pc > 0 else 0
                total_hpp = hpp_satuan * terjual_dari_iklan
                laba_bersih = omzet_hitung - total_hpp - biaya
                ra = round(omzet_hitung / biaya, 1) if biaya > 0 else 0
                
                key = f"{toko_iklan}_{kode}_{nama}" if toko_iklan else f"{kode}_{nama}"
                if key not in st.session_state.target_roas_data: 
                    st.session_state.target_roas_data[key] = 0
                
                hasil.append({
                    'No': len(hasil) + 1,
                    'Nama Toko': toko_iklan,
                    'Kode Iklan': kode,
                    'SKU': sku,
                    'Nama Iklan': nama,
                    'H.Jual/PC': hj,
                    'Estimasi/PC': estimasi_pc,
                    'HPP/PC': hpp_satuan,
                    'Terjual': terjual_dari_iklan,
                    'Untung/PC': up,
                    'Target_ROAS': st.session_state.target_roas_data[key],
                    'Omzet': omzet_hitung,
                    'Biaya': biaya,
                    'ROAS_Aktual': ra,
                    'ROAS_Min': round(estimasi_pc / (estimasi_pc - hpp_satuan), 1) if estimasi_pc > 0 and (estimasi_pc - hpp_satuan) > 0 else 0,
                    'Laba_Bersih': laba_bersih,
                    '_key': key
                })
            
            dfh = pd.DataFrame(hasil)
            
            # === TAMPILKAN JUMLAH TOKO ===
            jumlah_toko = dfh['Nama Toko'].nunique() if not dfh.empty else 0
            st.caption(f"🔍 Total: {len(dfh)} baris | Toko: {jumlah_toko} | Total Biaya: {format_rupiah(dfh['Biaya'].sum())}")
            
            # TAMPILKAN SEMUA TERMASUK YANG SKU '-'
            # if not dfh.empty and 'SKU' in dfh.columns:
            #     dfh = dfh[dfh['SKU'] != '-']
            
            if dfh.empty:
                st.info("📭 Tidak ada data.")
            else:
                
                st.markdown("---"); st.subheader("📊 Ringkasan")
                c1,c2,c3,c4,c5 = st.columns(5)
                c1.metric("📦 Terjual", format_angka(dfh['Terjual'].sum()))
                c2.metric("💰 Omzet", format_rupiah(dfh['Omzet'].sum()))
                c3.metric("📢 Biaya", format_rupiah(dfh['Biaya'].sum()))
                c4.metric("📈 ROAS", f"{round(dfh['Omzet'].sum()/dfh['Biaya'].sum(),1) if dfh['Biaya'].sum()>0 else 0}x")
                c5.metric("💎 Laba", format_rupiah(dfh['Laba_Bersih'].sum()))
                
                # Rincian Per Tanggal (dari dfh - sudah include HPP & fallback)
                st.markdown("---"); st.subheader("📅 Rincian Per Tanggal Per Toko")
                
                df_rincian = df.copy()
                df_rincian['Tgl'] = df_rincian[col_tgl].dt.strftime('%d/%m/%Y')
                
                rincian = df_rincian.groupby([col_toko, 'Tgl']).agg({
                    col_terjual: 'sum',
                    col_omzet: 'sum',
                    col_biaya: 'sum'
                }).reset_index()
                
                rincian.columns = ['Nama Toko', 'Tanggal', 'Terjual', 'Omzet', 'Biaya']
                rincian['Laba'] = rincian['Omzet'] - rincian['Biaya']
                rincian['ROAS'] = rincian.apply(lambda x: round(x['Omzet']/x['Biaya'],1) if x['Biaya']>0 else 0, axis=1)
                
                rincian_display = rincian.sort_values(['Tanggal', 'Nama Toko'])
                rincian_display['Omzet'] = rincian_display['Omzet'].apply(format_rupiah)
                rincian_display['Biaya'] = rincian_display['Biaya'].apply(format_rupiah)
                rincian_display['Laba'] = rincian_display['Laba'].apply(
                    lambda x: f"🟢 {int(x):,}".replace(',','.') if x > 0 else f"🔴 {int(x):,}".replace(',','.')
                )
                rincian_display['ROAS'] = rincian_display['ROAS'].apply(lambda x: f"{x}x")
                
                st.dataframe(rincian_display, use_container_width=True, height=400, hide_index=True)

                st.markdown("---")
                st.subheader("📋 Rincian Iklan")
                
                # Filter SKU + Edit ROAS
                col_edit1, col_edit2, col_edit3 = st.columns([1, 2, 1])
                with col_edit1:
                    sku_filter_bawah = st.text_input("🔎 Filter SKU", key="filter_sku_bwh", placeholder="Ketik SKU...")
                with col_edit2:
                    if sku_filter_bawah:
                        sku_list2 = [s.strip().upper() for s in sku_filter_bawah.split(',')]
                        mask_filter = pd.Series(False, index=dfh.index)
                        for s in sku_list2:
                            mask_filter = mask_filter | dfh['SKU'].astype(str).str.upper().str.contains(s, na=False)
                        dfh = dfh[mask_filter]
                    
                    if not dfh.empty:
                        sku_terfilter = dfh['SKU'].unique().tolist()
                        pilihan = [f"{h['No']}. [{h['Nama Toko']}] {h['Kode Iklan']} - {h['Nama Iklan'][:35]}" for h in hasil if h['SKU'] in sku_terfilter]
                    else:
                        pilihan = []
                    
                    if pilihan:
                        pil = st.selectbox("✏️ Edit ROAS", pilihan, key="pil_roas3")
                with col_edit3:
                    tgt = st.number_input("Target ROAS", value=3.0, step=0.5, key="tgt_roas3")
                    if st.button("💾 Simpan ROAS", key="simpan_roas_btn", use_container_width=True):
                        if pilihan:
                            idx = int(pil.split('.')[0])-1
                            key_roas = hasil[idx]['_key']
                            old_target = st.session_state.target_roas_data.get(key_roas, 0)
                            st.session_state.target_roas_data[key_roas] = tgt
                            pd.DataFrame(list(st.session_state.target_roas_data.items()), 
                                        columns=['key','target']).to_csv(str(DB_FOLDER / "target_roas.csv"), index=False)
                            riwayat_baru = pd.DataFrame([{
                                'Tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                'Key': key_roas, 'Target_Lama': old_target, 'Target_Baru': tgt,
                                'Kode_Iklan': hasil[idx]['Kode Iklan'], 'SKU': hasil[idx]['SKU'],
                                'Nama_Iklan': hasil[idx]['Nama Iklan'][:50]
                            }])
                            if os.path.exists(RIWAYAT_ROAS_FILE):
                                riwayat_all = pd.concat([pd.read_csv(RIWAYAT_ROAS_FILE), riwayat_baru], ignore_index=True)
                            else:
                                riwayat_all = riwayat_baru
                            riwayat_all.to_csv(RIWAYAT_ROAS_FILE, index=False)
                            st.success(f"✅ Target ROAS: {old_target}x → {tgt}x"); st.rerun()
 
                if os.path.exists(RIWAYAT_ROAS_FILE):
                    with st.expander("📜 Riwayat Target ROAS", expanded=False):
                        df_riwayat = pd.read_csv(RIWAYAT_ROAS_FILE).sort_values('Tanggal', ascending=False).head(20)
                        st.dataframe(df_riwayat, use_container_width=True, height=200, hide_index=True)
             
                if not dfh.empty:
                    with st.expander("🏆 TOP 10 Iklan & Ringkasan", expanded=False):
                        c1, c2, c3, c4, c5, c6 = st.columns(6)
                        with c1:
                            st.success("**10 Terlaris**"); top_omzet = dfh.nlargest(10, 'Omzet')[['Nama Iklan', 'Omzet']].copy()
                            top_omzet['Omzet'] = top_omzet['Omzet'].apply(format_rupiah); top_omzet.index = range(1, len(top_omzet)+1)
                            st.dataframe(top_omzet, use_container_width=True, height=300)
                        with c2:
                            st.error("**10 Rugi**"); top_rugi = dfh.nsmallest(10, 'Laba_Bersih')[['Nama Iklan', 'Laba_Bersih']].copy()
                            top_rugi = top_rugi[top_rugi['Laba_Bersih'] < 0]
                            if not top_rugi.empty: top_rugi['Laba_Bersih'] = top_rugi['Laba_Bersih'].apply(lambda x: format_rupiah(abs(x))); top_rugi.index = range(1, len(top_rugi)+1); st.dataframe(top_rugi, use_container_width=True, height=300)
                            else: st.caption("🎉 Tidak ada!")
                        with c3:
                            st.info("📊 **Total/Toko**"); total_all = dfh.groupby('Nama Toko').agg({'Laba_Bersih':'sum','Omzet':'sum','Biaya':'sum'}).reset_index()
                            total_all['Laba_Bersih'] = total_all['Laba_Bersih'].apply(lambda x: f"🟢 {format_rupiah(x)}" if x>0 else f"🔴 {format_rupiah(abs(x))}")
                            st.dataframe(total_all[['Nama Toko','Laba_Bersih']], use_container_width=True, height=300, hide_index=True)
                        with c4:
                            st.warning("📦 **SKU Terjual**"); top_sku = dfh.groupby('SKU').agg({'Terjual':'sum'}).reset_index().nlargest(10,'Terjual'); top_sku.index = range(1, len(top_sku)+1)
                            st.dataframe(top_sku, use_container_width=True, height=300)
                        with c5:
                            st.success("💎 **SKU Profit**"); top_profit = dfh.groupby('SKU').agg({'Laba_Bersih':'sum'}).reset_index(); top_profit = top_profit[top_profit['Laba_Bersih']>0].nlargest(10,'Laba_Bersih')
                            if not top_profit.empty: top_profit['Laba_Bersih'] = top_profit['Laba_Bersih'].apply(format_rupiah); top_profit.index = range(1, len(top_profit)+1); st.dataframe(top_profit, use_container_width=True, height=300)
                            else: st.caption("Belum ada")
                        with c6:
                            st.error("🔻 **SKU Rugi**"); top_rugi_sku = dfh.groupby('SKU').agg({'Laba_Bersih':'sum'}).reset_index(); top_rugi_sku = top_rugi_sku[top_rugi_sku['Laba_Bersih']<0].nsmallest(10,'Laba_Bersih')
                            if not top_rugi_sku.empty: top_rugi_sku['Laba_Bersih'] = top_rugi_sku['Laba_Bersih'].apply(lambda x: format_rupiah(abs(x))); top_rugi_sku.index = range(1, len(top_rugi_sku)+1); st.dataframe(top_rugi_sku, use_container_width=True, height=300)
                            else: st.caption("🎉 Tidak ada!")

                kol = ['No','Nama Toko','Kode Iklan','SKU','Nama Iklan','H.Jual/PC','Estimasi/PC','HPP/PC','Terjual','Untung/PC','Target_ROAS','ROAS_Min','Omzet','Biaya','ROAS_Aktual','Laba_Bersih']
                dfs = dfh[kol].copy()
                for c in ['H.Jual/PC','Estimasi/PC','HPP/PC','Untung/PC','Omzet','Biaya','Laba_Bersih']:
                    dfs[c] = dfs[c].fillna(0).astype(int).apply(lambda x: f"{int(x):,}".replace(',','.'))
                dfs['ROAS_Aktual'] = dfs['ROAS_Aktual'].fillna(0).apply(lambda x: f"{x}x")
                dfs['ROAS_Min'] = dfs['ROAS_Min'].fillna(0).apply(lambda x: f"{x}x")
                dfs['Target_ROAS'] = dfs['Target_ROAS'].fillna(0).apply(lambda x: f"{x}x")
                
                def warna_laba(val):
                    try:
                        v = int(float(str(val).replace('.','').replace(',','.')))
                        if v > 0: return f"🟢 {v:,}".replace(',','.')
                        elif v < 0: return f"🔴 {v:,}".replace(',','.')
                        return f"⚪ 0"
                    except: return f"⚪ 0"
                
                dfs['Laba_Bersih'] = dfs['Laba_Bersih'].astype(str).apply(warna_laba)                
                
                with st.expander("✏️ Edit Harga Jual (Cepat)", expanded=False):
                    col_ed1, col_ed2 = st.columns(2)
                    with col_ed1:
                        sku_edit = st.text_input("SKU (bisa paste)", key="sku_edit_cepat", placeholder="Ketik SKU...").strip().upper()
                    with col_ed2:
                        hj_edit = st.number_input("Harga Jual Baru", min_value=0, step=1000, key="hj_edit_cepat")
                    
                    if st.button("💾 Simpan Harga", key="simpan_harga_cepat", use_container_width=True):
                        if sku_edit and hj_edit > 0 and toko_pilih != 'Semua':
                            harga_toko_file = str(DB_FOLDER / "harga_per_toko.csv")
                            if os.path.exists(harga_toko_file):
                                df_ht = pd.read_csv(harga_toko_file)
                                mask_ht = (df_ht['SKU'].str.upper() == sku_edit) & (df_ht['Nama_Toko'].str.strip() == toko_pilih.strip())
                                if mask_ht.any():
                                    df_ht.loc[mask_ht, 'Harga_Jual'] = hj_edit
                                else:
                                    df_ht = pd.concat([df_ht, pd.DataFrame([{'SKU':sku_edit, 'Nama_Toko':toko_pilih, 'Harga_Jual':hj_edit}])])
                            else:
                                df_ht = pd.DataFrame([{'SKU':sku_edit, 'Nama_Toko':toko_pilih, 'Harga_Jual':hj_edit}])
                            df_ht.to_csv(harga_toko_file, index=False)
                            st.success(f"✅ {sku_edit} @ {toko_pilih}: Harga={hj_edit}")
                        else:
                            st.error("❌ Pilih toko & isi SKU/Harga")

                
                st.dataframe(dfs, use_container_width=True, height=500, hide_index=True)
                
                # Total kecil di pojok kanan
                total_omzet_all = int(dfh['Omzet'].sum())
                total_biaya_all = int(dfh['Biaya'].sum())
                total_laba_all = int(dfh['Laba_Bersih'].sum())
                margin_all = round(total_laba_all / total_omzet_all * 100, 1) if total_omzet_all > 0 else 0
                
                st.caption(f"💰 Omzet: {format_rupiah(total_omzet_all)} | 📢 Iklan: {format_rupiah(total_biaya_all)} | 💎 Laba: {format_rupiah(total_laba_all)} | 📊 Margin: {margin_all}%")
                
                st.download_button("📥 CSV", dfh[kol].to_csv(index=False).encode('utf-8-sig'), f"rekap_iklan_{datetime.now():%Y%m%d}.csv", use_container_width=True)
    
# =====================================
# 📋 MASTER PRODUK
# =====================================
elif active_menu == "📋 Master Produk":
    st.header("📋 Master Produk")
    
    # === FORM TAMBAH PRODUK ===
    with st.form("form_produk"):
        col1, col2 = st.columns(2)
        with col1:
            sku = st.text_input("SKU *")
            nama = st.text_input("Nama Produk *")
            supplier = st.text_input("Supplier")
        with col2:
            harga_beli = st.number_input("Harga Beli", min_value=0, step=1000)
            packing = st.number_input("Biaya Packing", min_value=0, step=500)
            harga_jual = st.number_input("Harga Jual", min_value=0, step=1000)
        
        if st.form_submit_button("💾 Simpan Produk Baru", type="primary", use_container_width=True):
            if sku and nama:
                if not st.session_state.master_produk.empty:
                    if str(sku).strip().upper() in st.session_state.master_produk['SKU'].astype(str).str.strip().str.upper().values:
                        st.error(f"❌ SKU '{sku}' sudah ada!")
                    else:
                        baru = pd.DataFrame([{
                            'SKU': str(sku).strip(),
                            'Nama_Produk': str(nama).strip(),
                            'Harga_Beli': harga_beli,
                            'Harga_Jual_Default': harga_jual,
                            'Packing': packing,
                            'Stok_Saat_Ini': 0,
                            'Stok_Opname': 0,
                            'Supplier_Default': str(supplier).strip()
                        }])
                        st.session_state.master_produk = pd.concat([st.session_state.master_produk, baru], ignore_index=True)
                        st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
                        st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
                        st.success(f"✅ {nama} ditambahkan!")
                        st.rerun()
                else:
                    baru = pd.DataFrame([{
                        'SKU': str(sku).strip(),
                        'Nama_Produk': str(nama).strip(),
                        'Harga_Beli': harga_beli,
                        'Harga_Jual_Default': harga_jual,
                        'Packing': packing,
                        'Stok_Saat_Ini': 0,
                        'Stok_Opname': 0,
                        'Supplier_Default': str(supplier).strip()
                    }])
                    st.session_state.master_produk = pd.concat([st.session_state.master_produk, baru], ignore_index=True)
                    st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
                    st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
                    st.success(f"✅ {nama} ditambahkan!")
                    st.rerun()
            else:
                st.error("SKU dan Nama wajib diisi!")
    
    # === DAFTAR PRODUK ===
    if not st.session_state.master_produk.empty:
        st.markdown("---")
        st.subheader("📊 Daftar Produk")
        
        # Tampilkan tabel
        df_produk = st.session_state.master_produk.copy()
        # Pilih kolom yang mau ditampilkan (sembunyikan Stok_Opname atau tidak)
        kolom_tampil = ['SKU', 'Nama_Produk', 'Harga_Beli', 'Harga_Jual_Default', 'Packing', 'Stok_Saat_Ini', 'Stok_Opname', 'Supplier_Default']
        kolom_ada = [c for c in kolom_tampil if c in df_produk.columns]
        # Tabel bisa diedit langsung
        edited_df = st.data_editor(
            df_produk[kolom_ada],
            use_container_width=True,
            height=300,
            num_rows="fixed",
            disabled=['SKU', 'Stok_Saat_Ini', 'Stok_Opname'],
            key="editor_master_produk"
        )
        
        if st.button("💾 Simpan Perubahan", type="primary", use_container_width=True):
            for idx, row in edited_df.iterrows():
                sku_row = str(row['SKU']).strip()
                mask = st.session_state.master_produk['SKU'].astype(str).str.strip() == sku_row
                if mask.any():
                    orig_idx = st.session_state.master_produk[mask].index[0]
                    for col in ['Nama_Produk', 'Harga_Beli', 'Harga_Jual_Default', 'Packing', 'Supplier_Default']:
                        if col in edited_df.columns:
                            st.session_state.master_produk.at[orig_idx, col] = row[col]
            
            st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
            st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
            st.success("✅ Perubahan disimpan!")
            st.rerun()
        
        # === STOK OPNAME ===
        st.markdown("---")
        st.subheader("🔍 Stok Opname (Hitung Fisik)")
        st.caption("Hitung stok fisik di gudang, sistem akan update Stok_Saat_Ini dan menyimpan ke Stok_Opname")
        
        list_sku = st.session_state.master_produk['SKU'].astype(str).str.strip().tolist()
        sku_opname = st.selectbox("Pilih SKU", list_sku, key="opname_sku_select")
        
        mask_op = st.session_state.master_produk['SKU'].astype(str).str.strip() == str(sku_opname).strip()
        
        if mask_op.any():
            idx_op = st.session_state.master_produk[mask_op].index[0]
            try:
                stok_sistem = int(float(st.session_state.master_produk.at[idx_op, 'Stok_Saat_Ini']))
            except:
                stok_sistem = 0
            
            try:
                stok_opname_lama = int(float(st.session_state.master_produk.at[idx_op, 'Stok_Opname']))
            except:
                stok_opname_lama = 0
            nama_produk_op = str(st.session_state.master_produk.at[idx_op, 'Nama_Produk'])
            
            col_a, col_b = st.columns(2)
            with col_a:
                st.text_input("Nama Produk", value=nama_produk_op, disabled=True)
                st.markdown(f"**Stok Sistem:** `{stok_sistem} pcs`")
                st.markdown(f"**Stok Opname Terakhir:** `{stok_opname_lama} pcs`")
            
            with col_b:
                stok_fisik = st.number_input("🔢 Stok Fisik (Hitung Ulang)", min_value=0, step=1, value=stok_sistem, key="stok_fisik_input")
                selisih = stok_fisik - stok_sistem
                if selisih > 0:
                    st.success(f"➕ Selisih: +{selisih} pcs")
                elif selisih < 0:
                    st.error(f"➖ Selisih: {selisih} pcs")
                else:
                    st.info("✅ Stok sesuai, tidak ada selisih")
            
            col_btn1, col_btn2, col_btn3 = st.columns(3)
            with col_btn1:
                if st.button("💾 Simpan Opname", key="simpan_opname_btn", type="primary", use_container_width=True):
                    # Update Stok_Opname
                    st.session_state.master_produk.at[idx_op, 'Stok_Opname'] = stok_fisik
                    
                    # Update stok batch FIFO
                    update_stok_saat_ini_dari_batch()
                    stok_batch_total = int(st.session_state.stok_batch[st.session_state.stok_batch['SKU'].astype(str).str.strip() == str(sku_opname).strip()]['Qty_Sisa'].sum()) if not st.session_state.stok_batch.empty else 0
                    selisih_opname = stok_fisik - stok_batch_total
                    if selisih_opname > 0:
                        tambah_stok_batch(sku_opname, selisih_opname, get_hpp_fifo(sku_opname), 'Opname', f'Penyesuaian +{selisih_opname}')
                    elif selisih_opname < 0 and stok_batch_total > 0:
                        kurangi_stok_fifo(sku_opname, abs(selisih_opname))
                    
                    update_stok_saat_ini_dari_batch()
                    
                    # Simpan ke CSV
                    st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
                    st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
                    
                    # Riwayat
                    if 'riwayat_opname' not in st.session_state:
                        st.session_state.riwayat_opname = []
                    st.session_state.riwayat_opname.append({
                        'Tanggal': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'SKU': str(sku_opname).strip(),
                        'Nama_Produk': nama_produk_op,
                        'Stok_Sistem_Lama': stok_sistem,
                        'Stok_Opname_Baru': stok_fisik,
                        'Selisih': selisih
                    })
                    
                    st.success(f"✅ Opname berhasil! {sku_opname}: {stok_sistem} → {stok_fisik} (Selisih: {'+' if selisih > 0 else ''}{selisih})")
                    st.rerun()
            
            with col_btn2:
                if st.button("🔄 Reset ke Opname", key="reset_stok_btn", type="secondary", use_container_width=True):
                    if stok_opname_lama > 0:
                        st.session_state.master_produk.at[idx_op, 'Stok_Saat_Ini'] = stok_opname_lama
                        st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
                        st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
                        st.success(f"✅ Stok {sku_opname} di-reset ke opname: {stok_opname_lama} pcs")
                        st.rerun()
                    else:
                        st.warning("⚠️ Belum ada data opname untuk produk ini")
            
            with col_btn3:
                if st.button("➕ Tambah Stok", key="tambah_stok_btn", type="secondary", use_container_width=True):
                    st.session_state.tambah_stok_sku = sku_opname
                    st.rerun()
        
        # === TAMBAH STOK CEPAT ===
        if 'tambah_stok_sku' in st.session_state and st.session_state.tambah_stok_sku:
            st.markdown("---")
            st.subheader(f"➕ Tambah Stok: {st.session_state.tambah_stok_sku}")
            qty_tambah = st.number_input("Jumlah", min_value=1, value=1, key="qty_tambah_cepat")
            if st.button("✅ Tambahkan", key="konfirmasi_tambah_cepat", type="primary"):
                tambah_stok_batch(st.session_state.tambah_stok_sku, qty_tambah, get_hpp_fifo(st.session_state.tambah_stok_sku), 'Manual', 'Tambah stok cepat')
                update_stok_saat_ini_dari_batch()
                st.success(f"✅ {qty_tambah} pcs ditambahkan ke {st.session_state.tambah_stok_sku}")
                del st.session_state.tambah_stok_sku
                st.rerun()
            if st.button("❌ Batal", key="batal_tambah_cepat"):
                del st.session_state.tambah_stok_sku
                st.rerun()
        
        # === RIWAYAT OPNAME ===
        if 'riwayat_opname' in st.session_state and st.session_state.riwayat_opname:
            with st.expander("📜 Riwayat Opname", expanded=False):
                df_riwayat = pd.DataFrame(st.session_state.riwayat_opname)
                if st.button("🗑️ Hapus Riwayat", key="hapus_riwayat_btn", type="secondary"):
                    st.session_state.riwayat_opname = []
                    st.rerun()
                st.dataframe(df_riwayat, use_container_width=True, height=200, hide_index=True)
        
        # === HAPUS PRODUK ===
        with st.expander("🗑️ Hapus Produk", expanded=False):
            sku_hapus = st.selectbox("Pilih SKU", list_sku, key="hapus_sku_select")
            if st.button("🗑️ Hapus Produk", key="hapus_produk_btn", type="secondary", use_container_width=True):
                mask_hapus = st.session_state.master_produk['SKU'].astype(str).str.strip() == str(sku_hapus).strip()
                if mask_hapus.any():
                    nama_hapus = st.session_state.master_produk[mask_hapus]['Nama_Produk'].iloc[0]
                    st.session_state.master_produk = st.session_state.master_produk[~mask_hapus]
                    st.session_state.master_produk.to_csv(MASTER_PRODUK_DB_FILE, index=False)
                    st.session_state.master_produk = pd.read_csv(MASTER_PRODUK_DB_FILE)
                    st.success(f"✅ {nama_hapus} ({sku_hapus}) dihapus!")
                    st.rerun()
    else:
        st.info("📭 Belum ada produk. Tambahkan produk pertama Anda!")
# =====================================
# 🛒 PEMBELIAN STOK
# =====================================
elif active_menu == "🛒 Pembelian Stok":
    st.header("🛒 Pembelian Stok (Multi-SKU)")
    
    if st.session_state.master_produk.empty:
        st.warning("⚠️ Tambah produk dulu di Master Produk!")
    else:
        # Init session untuk keranjang belanja
        if 'keranjang_beli' not in st.session_state:
            st.session_state.keranjang_beli = []
        if 'supplier_aktif' not in st.session_state:
            st.session_state.supplier_aktif = ''
        if 'no_invoice_aktif' not in st.session_state:
            st.session_state.no_invoice_aktif = f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        if 'metode_aktif' not in st.session_state:
            st.session_state.metode_aktif = 'Cash'
        
        # Daftar supplier
        daftar_supplier = []
        if not st.session_state.pembelian_data.empty and 'Supplier' in st.session_state.pembelian_data.columns:
            daftar_supplier = sorted(st.session_state.pembelian_data['Supplier'].dropna().unique().tolist())
        daftar_supplier = ['(Supplier Baru)'] + daftar_supplier
        
        # Supplier, No Invoice, Metode (dalam form)
        with st.form("form_info_invoice"):
            col1, col2, col3 = st.columns(3)
            with col1:
                if 'supplier_baru_input' in st.session_state and st.session_state.supplier_baru_input:
                    if st.session_state.supplier_baru_input not in daftar_supplier:
                        daftar_supplier.append(st.session_state.supplier_baru_input)
                
                supplier = st.selectbox("Supplier *", daftar_supplier, key="supplier_pilih")
                if supplier == '(Supplier Baru)':
                    supplier = st.text_input("Ketik Supplier Baru", key="supplier_baru_input", placeholder="Nama supplier...")
            with col2:
                no_invoice = st.text_input("No Invoice", value=st.session_state.no_invoice_aktif, key="no_inv_beli")
            with col3:
                metode = st.selectbox("Pembayaran", ['Cash','Transfer','Tempo 30','Tempo 60'], 
                                     index=['Cash','Transfer','Tempo 30','Tempo 60'].index(st.session_state.metode_aktif) if st.session_state.metode_aktif in ['Cash','Transfer','Tempo 30','Tempo 60'] else 0,
                                     key="metode_beli")
            
            submitted = st.form_submit_button("✅ Simpan Info Invoice", type="primary", use_container_width=True)
            if submitted:
                if supplier:
                    st.session_state.supplier_aktif = supplier
                    st.session_state.no_invoice_aktif = no_invoice
                    st.session_state.metode_aktif = metode
                    st.success("✅ Info invoice disimpan!")
                    st.rerun()
                else:
                    st.error("Supplier wajib diisi!")
        
        # SKU, Qty, Harga DI LUAR FORM (reaktif)
        if st.session_state.supplier_aktif:
            st.markdown("---")
            st.caption(f"📋 Invoice: {st.session_state.no_invoice_aktif} | 🏪 {st.session_state.supplier_aktif} | 💳 {st.session_state.metode_aktif}")
            st.caption("Tambah SKU:")
            
            col_sku, col_qty, col_harga, col_btn = st.columns([2, 1, 1, 1])
            with col_sku:
                list_sku_beli = sorted(st.session_state.master_produk['SKU'].tolist())
                sku_beli = st.selectbox("SKU", list_sku_beli, key="sku_beli_select")
                info_sku = st.session_state.master_produk[st.session_state.master_produk['SKU']==sku_beli].iloc[0]
                packing_sku = int(info_sku['Packing']) if 'Packing' in info_sku and pd.notna(info_sku['Packing']) else 0
                
                harga_default = 0
                if not st.session_state.pembelian_data.empty and 'SKU' in st.session_state.pembelian_data.columns:
                    df_riwayat = st.session_state.pembelian_data[
                        st.session_state.pembelian_data['SKU'].astype(str).str.strip() == sku_beli.strip()
                    ]
                    if not df_riwayat.empty:
                        harga_default = int(df_riwayat.sort_values('Tanggal', ascending=False)['Harga_Beli'].iloc[0])
                if harga_default == 0:
                    harga_default = int(info_sku['Harga_Beli']) if pd.notna(info_sku['Harga_Beli']) else 0
            with col_qty:
                qty_beli = st.number_input("Qty", min_value=1, value=1, key="qty_beli_input")
            with col_harga:
                harga_beli = st.number_input("Harga/Unit", min_value=0, value=harga_default, key=f"harga_beli_{sku_beli}")
                st.caption(f"📦 Packing: {format_rupiah(packing_sku)} | Total/Unit: {format_rupiah(harga_beli + packing_sku)}")
            with col_btn:
                st.write("")
                if st.button("➕ Tambah", key="tambah_sku_btn", use_container_width=True):
                    if sku_beli:
                        st.session_state.keranjang_beli.append({
                            'SKU': sku_beli,
                            'Nama_Barang': info_sku['Nama_Produk'],
                            'Qty': qty_beli,
                            'Harga_Beli': harga_beli,
                            'Packing': packing_sku,
                            'Total': qty_beli * harga_beli
                        })
                        st.success(f"✅ {sku_beli} ditambahkan")
                        st.rerun()
        else:
            st.info("📭 Simpan info invoice dulu (Supplier, No Invoice, Pembayaran)")
        
        # Tampilkan keranjang
        if st.session_state.keranjang_beli:
            st.markdown("---")
            st.subheader(f"📋 Keranjang")
            
            df_keranjang = pd.DataFrame(st.session_state.keranjang_beli)
            df_keranjang['No'] = range(1, len(df_keranjang)+1)
            df_display = df_keranjang[['No','SKU','Nama_Barang','Qty','Harga_Beli','Total']].copy()
            for col in ['Harga_Beli','Total']:
                df_display[col] = df_display[col].apply(format_rupiah)
            st.dataframe(df_display, use_container_width=True, height=250, hide_index=True)
            
            total_invoice = df_keranjang['Total'].sum()
            total_qty = df_keranjang['Qty'].sum()
            
            col_tot1, col_tot2 = st.columns(2)
            with col_tot1:
                st.markdown(f"### 📦 Total Qty: {int(total_qty)}")
            with col_tot2:
                st.markdown(f"### 💰 Total Belanja: {format_rupiah(int(total_invoice))}")
            
            col_simpan, col_batal, col_hapus = st.columns(3)
            with col_simpan:
                if st.button("💾 SIMPAN INVOICE", type="primary", use_container_width=True):
                    tgl = datetime.now()
                    for item in st.session_state.keranjang_beli:
                        baru = pd.DataFrame([{
                            'Tanggal': tgl.strftime('%Y-%m-%d'),
                            'No_Invoice': st.session_state.no_invoice_aktif,
                            'Supplier': st.session_state.supplier_aktif,
                            'SKU': item['SKU'],
                            'Nama_Barang': item['Nama_Barang'],
                            'Qty': item['Qty'],
                            'Harga_Beli': item['Harga_Beli'],
                            'Total': item['Total'],
                            'Metode': st.session_state.metode_aktif,
                            'Jatuh_Tempo': tgl.strftime('%Y-%m-%d'),
                            'Status': 'Lunas' if st.session_state.metode_aktif in ['Cash','Transfer'] else 'Hutang',
                            'No_Ref': '',
                            'Catatan': f"Packing: {format_rupiah(item['Packing'])}"
                        }])
                        st.session_state.pembelian_data = pd.concat([st.session_state.pembelian_data, baru], ignore_index=True)
                        save_pembelian_data(st.session_state.pembelian_data)
                        tambah_stok_batch(item['SKU'], item['Qty'], item['Harga_Beli'] + item['Packing'], 
                                         'Pembelian', f"INV: {st.session_state.no_invoice_aktif}")
                    
                    update_stok_saat_ini_dari_batch()
                    st.session_state.keranjang_beli = []
                    st.session_state.no_invoice_aktif = f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    st.success(f"✅ Invoice disimpan! ({len(df_keranjang)} SKU)")
                    st.rerun()
            
            with col_batal:
                if st.button("🗑️ BATAL", type="secondary", use_container_width=True):
                    st.session_state.keranjang_beli = []
                    st.session_state.no_invoice_aktif = f"INV-{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    st.rerun()
            
            with col_hapus:
                if len(st.session_state.keranjang_beli) > 0:
                    hapus_idx = st.number_input("Hapus baris No", min_value=1, max_value=len(st.session_state.keranjang_beli), step=1, key="hapus_idx")
                    if st.button("❌ Hapus Baris", type="secondary", use_container_width=True):
                        del st.session_state.keranjang_beli[hapus_idx - 1]
                        st.rerun()
        
        if not st.session_state.pembelian_data.empty:
            st.markdown("---")
            st.subheader("📜 Riwayat Pembelian")
            st.dataframe(st.session_state.pembelian_data.tail(20), use_container_width=True, height=250, hide_index=True)
# =====================================
# 📦 KARTU STOK (FIFO)
# =====================================
elif active_menu == "📦 Kartu Stok":
    st.header("📦 Kartu Stok FIFO")
    
    if not st.session_state.master_produk.empty and not st.session_state.stok_batch.empty:
        list_sku = st.session_state.master_produk['SKU'].astype(str).str.strip().tolist()
        sku_pilih = st.selectbox("Pilih SKU", list_sku)
        
        if sku_pilih:
            mask_mp = st.session_state.master_produk['SKU'].astype(str).str.strip() == sku_pilih
            if mask_mp.any():
                info = st.session_state.master_produk[mask_mp].iloc[0]
                st.markdown(f"**{info['Nama_Produk']}** | Stok: {info['Stok_Saat_Ini']} pcs | HPP FIFO: {format_rupiah(get_hpp_fifo(sku_pilih))}")
            
            st.subheader("📥 Batch Stok Tersedia")
            df_batch = st.session_state.stok_batch[st.session_state.stok_batch['SKU'].astype(str).str.strip() == sku_pilih].copy()
            df_batch = df_batch.sort_values('Tanggal_Masuk')
            if not df_batch.empty:
                df_batch_display = df_batch[['Tanggal_Masuk','Qty_Awal','Qty_Sisa','Harga_Satuan','Sumber','Keterangan']].copy()
                df_batch_display['Nilai_Sisa'] = df_batch_display['Qty_Sisa'] * df_batch_display['Harga_Satuan']
                for col in ['Harga_Satuan','Nilai_Sisa']:
                    df_batch_display[col] = df_batch_display[col].apply(format_rupiah)
                st.dataframe(df_batch_display, use_container_width=True, height=300, hide_index=True)
                
                st.subheader("📤 Mutasi Keluar")
                df_mutasi = df_batch[df_batch['Qty_Sisa'] < df_batch['Qty_Awal']]
                if not df_mutasi.empty:
                    df_mutasi['Qty_Keluar'] = df_mutasi['Qty_Awal'] - df_mutasi['Qty_Sisa']
                    df_mutasi['Nilai_Keluar'] = df_mutasi['Qty_Keluar'] * df_mutasi['Harga_Satuan']
                    df_mutasi_display = df_mutasi[['Tanggal_Masuk','Qty_Keluar','Harga_Satuan','Nilai_Keluar']].copy()
                    for col in ['Harga_Satuan','Nilai_Keluar']:
                        df_mutasi_display[col] = df_mutasi_display[col].apply(format_rupiah)
                    st.dataframe(df_mutasi_display, use_container_width=True, height=200, hide_index=True)
                else:
                    st.caption("📭 Belum ada mutasi keluar")
            else:
                st.info("📭 Tidak ada batch untuk SKU ini")
    else:
        st.info("📭 Data tidak lengkap. Tambahkan produk & pembelian dulu.")

# =====================================
# 🧮 SIMULASI ROAS
# =====================================
elif active_menu == "🧮 Simulasi ROAS":
    st.header("🧮 Simulasi ROAS")
    st.caption("Hitung ROAS Minimal & Harga Jual Ideal - Cukup masukkan HPP")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        hpp_input = st.number_input("HPP + Packing (Rp)", min_value=0, value=205000, step=1000, key="sim_hpp")
    with col2:
        margin_target = st.slider("Target Margin (%)", min_value=5, max_value=40, value=15, step=1, key="sim_margin")
    with col3:
        rasio_estimasi = st.slider("Rasio Estimasi Diterima", min_value=0.70, max_value=0.90, value=0.78, step=0.01, key="sim_rasio")
    
    if hpp_input > 0:
        cuan_target = int(hpp_input * margin_target / 100)
        estimasi_target = hpp_input + cuan_target
        harga_ideal = int(estimasi_target / rasio_estimasi)
        
        st.markdown("---")
        st.subheader("📊 Hasil Perhitungan")
        
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        with col_r1:
            st.metric("💎 HPP + Packing", format_rupiah(hpp_input))
        with col_r2:
            st.metric(f"🎯 Target Cuan {margin_target}%", format_rupiah(cuan_target))
        with col_r3:
            st.metric("📥 Estimasi Diterima Target", format_rupiah(estimasi_target))
        with col_r4:
            st.metric("💰 Harga Jual Ideal", format_rupiah(harga_ideal))
        
        st.markdown(f"""
        Estimasi Diterima = HPP + Cuan {margin_target}%
        Estimasi Diterima = {format_rupiah(hpp_input)} + {format_rupiah(cuan_target)}
        Estimasi Diterima = {format_rupiah(estimasi_target)}
        
        Harga Jual = {format_rupiah(estimasi_target)} / {rasio_estimasi}
        Harga Jual = {format_rupiah(harga_ideal)}
        """)
        
        # Simulasi range harga
        st.markdown("---")
        st.subheader("📈 SIMULASI RANGE HARGA")
        
        data_range = []
        for pct in [85, 88, 90, 92, 95, 98, 100, 102, 105, 108, 110]:
            harga_var = int(harga_ideal * pct / 100)
            estimasi_var = int(harga_var * rasio_estimasi)
            laba_kotor = estimasi_var - hpp_input
            laba_persen = round(laba_kotor / hpp_input * 100, 1)
            roas_min_var = round(harga_var / laba_kotor, 1) if laba_kotor > 0 else 0
            
            if laba_persen >= margin_target:
                status = "✅✅ SUPER"
            elif laba_persen >= margin_target * 0.7:
                status = "✅ AMAN"
            elif laba_persen > 0:
                status = "⚠️ TIPIS"
            else:
                status = "🔴 RUGI"
            
            data_range.append({
                'Harga Jual': format_rupiah(harga_var),
                'Estimasi (78%)': format_rupiah(estimasi_var),
                'Laba Bersih': format_rupiah(laba_kotor),
                '% dari HPP': f"{laba_persen}%",
                'ROAS Min': f"{roas_min_var}x",
                'Status': status
            })
        
        df_range = pd.DataFrame(data_range)
        st.dataframe(df_range, use_container_width=True, height=350, hide_index=True)
        
        # Simulasi ROAS
        st.markdown("---")
        st.subheader(f"📊 SIMULASI ROAS - HPP {format_rupiah(hpp_input)} - HARGA {format_rupiah(harga_ideal)}")
        
        laba_kotor_final = int(harga_ideal * rasio_estimasi) - hpp_input
        roas_min_final = round(harga_ideal / laba_kotor_final, 1) if laba_kotor_final > 0 else 0
        
        roas_data = []
        for target in [roas_min_final, roas_min_final + 0.1, 7.0, 8.0, 10.0, 12.0, 15.0]:
            target = round(target, 1)
            if target > 0:
                acos = round(1 / target * 100, 1)
                biaya_pcs = int(harga_ideal / target)
                laba_bersih = laba_kotor_final - biaya_pcs
                
                if target <= roas_min_final + 0.1:
                    status = "⚠️ BEP" if target == roas_min_final else "⚠️ Mendekati BEP"
                elif target < 7:
                    status = "✅"
                elif target < 10:
                    status = "✅✅ AMAN"
                else:
                    status = "✅✅✅ SUPER"
                
                roas_data.append({
                    'Target ROAS': f"{target}x",
                    'ACOS': f"{acos}%",
                    'Biaya Iklan/Pcs': format_rupiah(biaya_pcs),
                    'Laba Bersih/Pcs': f"{'+' if laba_bersih >= 0 else ''}{format_rupiah(laba_bersih)}",
                    'Status': status
                })
        
        df_bep = [r for r in roas_data if 'BEP' in r['Status']]
        df_rekom = [r for r in roas_data if 'BEP' not in r['Status']]
        
        if df_bep:
            st.markdown("**ROAS Minimal (BEP)**")
            st.dataframe(pd.DataFrame(df_bep), use_container_width=True, hide_index=True)
        
        if df_rekom:
            st.markdown("**Rekomendasi**")
            st.dataframe(pd.DataFrame(df_rekom), use_container_width=True, hide_index=True)
        
        st.markdown("---")
        st.subheader("🎯 KESIMPULAN")
        st.markdown(f"""
        Untuk **Cuan {margin_target}% dari HPP**
        
        | Settingan | Nilai |
        |-----------|-------|
        | HPP + Packing | {format_rupiah(hpp_input)} |
        | Target Cuan ({margin_target}%) | {format_rupiah(cuan_target)} |
        | Harga Jual Minimal | {format_rupiah(harga_ideal)} |
        | ROAS Minimal (BEP) | {roas_min_final}x |
        
        ⚡ **REKOMENDASI:** Pasang harga **{format_rupiah(harga_ideal)}** dengan Target ROAS **7.0x**! 🚀
        """)
# =====================================
# 🔍 FILTER DATA
# =====================================
elif active_menu == "🔍 Filter Data":
    if not st.session_state.shopee_data.empty:
        df = st.session_state.shopee_data.copy()
        
        if 'No. Pesanan' not in df.columns and 'No Pesanan' in df.columns: 
            df['No. Pesanan'] = df['No Pesanan']
        if 'SKU Induk' not in df.columns and 'SKU' in df.columns: 
            df['SKU Induk'] = df['SKU']
        if 'Terjual' not in df.columns:
            if 'Jumlah' in df.columns:
                df['Terjual'] = df['Jumlah']
            elif 'Jumlah Produk di Pesan' in df.columns:
                df['Terjual'] = df['Jumlah Produk di Pesan']
        if 'Status Pesanan' not in df.columns and 'Status' in df.columns: 
            df['Status Pesanan'] = df['Status']
        
        iklan_kode_map = {}
        if not st.session_state.iklan_data.empty:
            for _, row in st.session_state.iklan_data.iterrows():
                keyword = " ".join(str(row['Nama Iklan']).lower().split()[:4])
                kode = str(row['Kode Produk']).strip() if pd.notna(row.get('Kode Produk')) else ''
                if keyword and kode:
                    iklan_kode_map[keyword] = kode
        
        def cari_kode_produk(nama_produk):
            if not nama_produk or pd.isna(nama_produk):
                return ''
            nama_lower = str(nama_produk).lower()
            for keyword, kode in iklan_kode_map.items():
                if keyword in nama_lower:
                    return kode
            return ''
        
        df['Kode Produk (Iklan)'] = df['Nama Produk'].apply(cari_kode_produk) if 'Nama Produk' in df.columns else ''
        
        kolom_tersedia = ['No. Pesanan', 'Tanggal', 'Nama Toko', 'SKU Induk', 'Jumlah Produk di Pesan', 'Status Pesanan']
        kolom = [k for k in kolom_tersedia if k in df.columns]
        
        st.markdown('<div class="box"><div class="box-title">🔍 Filter Data</div>', unsafe_allow_html=True)
        
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            if 'Nama Toko' in df.columns:
                t = st.selectbox("🏪 Toko", ['Semua'] + sorted(df['Nama Toko'].dropna().unique()))
                if t != 'Semua': 
                    df = df[df['Nama Toko'] == t]
        with c2:
            if 'Status Pesanan' in df.columns:
                s = st.selectbox("📋 Status", ['Semua'] + sorted(df['Status Pesanan'].dropna().unique()))
                if s != 'Semua': 
                    df = df[df['Status Pesanan'].str.lower() == s.lower()]
        with c3:
            if 'Waktu Pesanan Dibuat' in df.columns:
                df['Tanggal'] = pd.to_datetime(df['Waktu Pesanan Dibuat'], errors='coerce')
                tgl_min = df['Tanggal'].min().date()
                tgl_max = df['Tanggal'].max().date()
                range_tgl = st.date_input("📅 Tanggal", value=(tgl_min, tgl_max), key="filter_tgl")
                if len(range_tgl) == 2:
                    df = df[(df['Tanggal'].dt.date >= range_tgl[0]) & (df['Tanggal'].dt.date <= range_tgl[1])]
        with c4:
            q = st.text_input("🔎 Cari", placeholder="No. Pesanan / SKU...")
            if q:
                mask = (
                    (df['No. Pesanan'].astype(str).str.contains(q, case=False, na=False)) |
                    (df['SKU Induk'].astype(str).str.contains(q, case=False, na=False)) |
                    (df['Nama Produk'].astype(str).str.contains(q, case=False, na=False))
                ) if 'Nama Produk' in df.columns else (
                    df['No. Pesanan'].astype(str).str.contains(q, case=False, na=False)
                )
                df = df[mask]
        
        if 'Tanggal' in df.columns:
            df['Tanggal'] = df['Tanggal'].dt.strftime('%d/%m/%Y')
        
        st.caption(f"📊 Menampilkan {len(df)} data")
        st.dataframe(df[kolom], use_container_width=True, height=500, hide_index=True)
        
        st.download_button(
            "📥 Download CSV", 
            df[kolom].to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig'), 
            f"filtered_{datetime.now().strftime('%Y%m%d')}.csv", 
            "text/csv",
            use_container_width=True
        )
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="box"><div style="text-align:center;padding:40px;">📭 Belum ada data penjualan</div></div>', unsafe_allow_html=True)

# =====================================
# 💎 LAPORAN KEUANGAN
# =====================================
elif active_menu == "💎 Laporan Keuangan":
    st.header("💎 Laporan Keuangan Bulanan")
    
    # Init database
    PENGELUARAN_DB_FILE = str(DB_FOLDER / "pengeluaran.csv")
    if 'pengeluaran_data' not in st.session_state:
        if os.path.exists(PENGELUARAN_DB_FILE):
            st.session_state.pengeluaran_data = pd.read_csv(PENGELUARAN_DB_FILE)
        else:
            st.session_state.pengeluaran_data = pd.DataFrame(columns=['Tanggal','Kategori','Jumlah','Keterangan'])
    
    tab_laporan, tab_hutang, tab_rekon = st.tabs(["📊 Laporan", "💳 Hutang Supplier", "🏦 Rekonsiliasi"])
    
    # ============================================================
    # TAB LAPORAN
    # ============================================================
    with tab_laporan:
        col1, col2, col3 = st.columns(3)
        with col1:
            bulan_pilih = st.selectbox("📅 Bulan", range(1,13), index=datetime.now().month-1, 
                                       format_func=lambda x: ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Ags','Sep','Okt','Nov','Des'][x-1])
        with col2:
            tahun_pilih = st.selectbox("📅 Tahun", range(2024,2031), index=datetime.now().year-2024)
        
        # Hitung data bulanan
        total_penjualan = 0
        if not st.session_state.income_data.empty and 'Tanggal_Dana' in st.session_state.income_data.columns and 'Total_Dibayar' in st.session_state.income_data.columns:
            df_inc = st.session_state.income_data.copy()
            df_inc['Tanggal_Dana'] = pd.to_datetime(df_inc['Tanggal_Dana'], errors='coerce')
            # Filter sama persis dengan Cair: 1 April - 30 April
            tgl_a = datetime(tahun_pilih, bulan_pilih, 1)
            if bulan_pilih == 12:
                tgl_b = datetime(tahun_pilih+1, 1, 1) - timedelta(days=1)
            else:
                tgl_b = datetime(tahun_pilih, bulan_pilih+1, 1) - timedelta(days=1)
            mask = (df_inc['Tanggal_Dana'].dt.date >= tgl_a.date()) & (df_inc['Tanggal_Dana'].dt.date <= tgl_b.date())
            total_penjualan = int(df_inc[mask]['Total_Dibayar'].sum())
        
        total_iklan = 0
        if not st.session_state.iklan_data.empty and 'Tanggal' in st.session_state.iklan_data.columns and 'Biaya' in st.session_state.iklan_data.columns:
            df_ik = st.session_state.iklan_data.copy()
            df_ik['Tanggal'] = pd.to_datetime(df_ik['Tanggal'], errors='coerce')
            mask = (df_ik['Tanggal'].dt.year == tahun_pilih) & (df_ik['Tanggal'].dt.month == bulan_pilih)
            total_iklan = int(df_ik[mask]['Biaya'].sum())
        
        total_hpp = 0
        if not st.session_state.income_data.empty and not st.session_state.master_produk.empty:
            df_inc = st.session_state.income_data.copy()
            if 'Tanggal_Dana' in df_inc.columns and 'Total_Dibayar' in df_inc.columns:
                df_inc['Tanggal_Dana'] = pd.to_datetime(df_inc['Tanggal_Dana'], errors='coerce')
                mask = (df_inc['Tanggal_Dana'].dt.date >= tgl_a.date()) & (df_inc['Tanggal_Dana'].dt.date <= tgl_b.date())
                df_inc = df_inc[mask]
                
                # Gabung dengan shopee_data untuk dapat SKU & Qty
                if not st.session_state.shopee_data.empty and 'No Pesanan' in st.session_state.shopee_data.columns:
                    df_sp = st.session_state.shopee_data[['No Pesanan', 'Nomor Referensi SKU', 'Jumlah Produk di Pesan']].copy()
                    df_sp.columns = ['No_Pesanan', 'SKU', 'Qty']
                    df_inc = df_inc.merge(df_sp, left_on='No_Pesanan', right_on='No_Pesanan', how='left')
                    
                    hpp_map_laporan = {}
                    for _, row in st.session_state.master_produk.iterrows():
                        hpp_map_laporan[str(row['SKU']).strip().upper()] = int(row['Harga_Beli']) + int(row['Packing'])
                    
                    for _, row in df_inc.iterrows():
                        sku = str(row['SKU']).strip().upper() if pd.notna(row['SKU']) else ''
                        if sku:
                            try:
                                qty = int(float(row['Qty']))
                            except:
                                qty = 1
                            hpp_val = get_hpp_fifo(sku) if 'get_hpp_fifo' in dir() else 0
                            if hpp_val == 0:
                                hpp_val = hpp_map_laporan.get(sku, 0)
                            total_hpp += hpp_val * qty
        
        total_pengeluaran = 0
        if not st.session_state.pengeluaran_data.empty:
            df_peng = st.session_state.pengeluaran_data.copy()
            df_peng['Tanggal'] = pd.to_datetime(df_peng['Tanggal'], errors='coerce')
            mask = (df_peng['Tanggal'].dt.date >= tgl_a.date()) & (df_peng['Tanggal'].dt.date <= tgl_b.date())
            total_pengeluaran = int(df_peng[mask]['Jumlah'].sum())
        
        laba_bersih = total_penjualan - total_hpp - total_iklan - total_pengeluaran
        
        # Ringkasan
        st.markdown("---")
        st.subheader(f"📊 Ringkasan Bulan {['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Ags','Sep','Okt','Nov','Des'][bulan_pilih-1]} {tahun_pilih}")
        
        ca, cb, cc, cd, ce = st.columns(5)
        ca.metric("💰 Penjualan", format_rupiah(total_penjualan))
        cb.metric("📊 HPP", format_rupiah(total_hpp))
        cc.metric("📢 Iklan", format_rupiah(total_iklan))
        cd.metric("💸 Pengeluaran", format_rupiah(total_pengeluaran))
        with ce:
            warna = "#10b981" if laba_bersih >= 0 else "#ef4444"
            label = "💎 LABA BERSIH" if laba_bersih >= 0 else "💎 RUGI BERSIH"
            st.markdown(f"""
            <div style="background:{warna}20;border:2px solid {warna};border-radius:12px;padding:16px;text-align:center;">
                <div style="font-size:11px;color:{warna};">{label}</div>
                <div style="font-size:22px;font-weight:800;color:{warna};">{format_rupiah(abs(laba_bersih))}</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Input pengeluaran
        st.markdown("---")
        st.subheader("💸 Input Pengeluaran")
        with st.form("form_pengeluaran"):
            cp1, cp2, cp3 = st.columns(3)
            with cp1:
                tgl_peng = st.date_input("Tanggal", value=datetime.now())
            with cp2:
                kat_peng = st.selectbox("Kategori", ['Operasional','Gaji','Sewa','Listrik','Internet','Transport','Packaging','Lainnya'])
            with cp3:
                jml_peng = st.number_input("Jumlah (Rp)", min_value=0, step=10000)
            ket_peng = st.text_input("Keterangan", placeholder="Deskripsi...")
            if st.form_submit_button("💾 Simpan", type="primary", use_container_width=True):
                if jml_peng > 0:
                    baru = pd.DataFrame([{'Tanggal':tgl_peng.strftime('%Y-%m-%d'),'Kategori':kat_peng,'Jumlah':jml_peng,'Keterangan':ket_peng}])
                    st.session_state.pengeluaran_data = pd.concat([st.session_state.pengeluaran_data, baru], ignore_index=True)
                    st.session_state.pengeluaran_data.to_csv(PENGELUARAN_DB_FILE, index=False)
                    st.success(f"✅ {format_rupiah(jml_peng)} dicatat!")
                    st.rerun()
                else:
                    st.error("Jumlah harus > 0")
        # ============================================================
        # INPUT PEMASUKAN CHANKO
        # ============================================================
        with st.expander("➕ Chanko) - Buka/Tutup", expanded=False):
            with st.form("form_pemasukan"):
                cp1, cp2, cp3 = st.columns(3)
                with cp1:
                    tgl_masuk = st.date_input("Tanggal", value=datetime.now(), key="tgl_masuk")
                with cp2:
                    sumber_masuk = st.text_input("Sumber (Nama Mitra)", placeholder="Chanko, Mitra A...", key="sumber_masuk")
                with cp3:
                    jml_masuk = st.number_input("Jumlah", step=50000, key="jml_masuk")
                
                ket_masuk = st.text_input("Keterangan", placeholder="Setoran harian...", key="ket_masuk")
                
                if st.form_submit_button("💾 Simpan Pemasukan", type="primary", use_container_width=True):
                    if jml_masuk != 0 and sumber_masuk:
                        baru_masuk = pd.DataFrame([{
                            'Tanggal': tgl_masuk.strftime('%Y-%m-%d'),
                            'Kategori': f'Pemasukan - {sumber_masuk}',
                            'Jumlah': -jml_masuk,
                            'Keterangan': ket_masuk
                        }])
                        st.session_state.pengeluaran_data = pd.concat([st.session_state.pengeluaran_data, baru_masuk], ignore_index=True)
                        st.session_state.pengeluaran_data.to_csv(PENGELUARAN_DB_FILE, index=False)
                        st.success(f"✅ Pemasukan {format_rupiah(jml_masuk)} dari {sumber_masuk} dicatat!")
                        st.rerun()
                    else:
                        st.error("Jumlah tidak boleh 0 dan Sumber wajib diisi!")
        
        # Tabel pengeluaran
        if not st.session_state.pengeluaran_data.empty:
            st.markdown("---")
            st.subheader("📋 Pengeluaran Bulan Ini")
            df_pt = st.session_state.pengeluaran_data.copy()
            df_pt['Tanggal'] = pd.to_datetime(df_pt['Tanggal'], errors='coerce')
            mask = (df_pt['Tanggal'].dt.year == tahun_pilih) & (df_pt['Tanggal'].dt.month == bulan_pilih)
            df_pt = df_pt[mask].sort_values('Tanggal', ascending=False)
            if not df_pt.empty:
                df_pt['Tanggal'] = df_pt['Tanggal'].dt.strftime('%d/%m/%Y')
                df_pt_d = df_pt[['Tanggal','Kategori','Jumlah','Keterangan']].copy()
                df_pt_d['Jumlah'] = df_pt_d['Jumlah'].apply(format_rupiah)
                st.dataframe(df_pt_d, use_container_width=True, height=250, hide_index=True)
            else:
                st.caption("📭 Belum ada pengeluaran")
        
        # Grafik tren
        st.markdown("---")
        st.subheader("📈 Tren 6 Bulan")
        tren_data = []
        for i in range(5, -1, -1):
            bln = datetime.now().month - i
            thn = datetime.now().year
            if bln <= 0:
                bln += 12
                thn -= 1
            pen = 0
            if not st.session_state.income_data.empty and 'Tanggal_Dana' in st.session_state.income_data.columns:
                df_i = st.session_state.income_data.copy()
                df_i['Tgl'] = pd.to_datetime(df_i['Tanggal_Dana'], errors='coerce')
                pen = int(df_i[(df_i['Tgl'].dt.year==thn)&(df_i['Tgl'].dt.month==bln)]['Total_Dibayar'].sum())
            ikl = 0
            if not st.session_state.iklan_data.empty:
                df_ik2 = st.session_state.iklan_data.copy()
                df_ik2['Tgl'] = pd.to_datetime(df_ik2['Tanggal'], errors='coerce')
                ikl = int(df_ik2[(df_ik2['Tgl'].dt.year==thn)&(df_ik2['Tgl'].dt.month==bln)]['Biaya'].sum())
            tren_data.append({'Bulan':f"{['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Ags','Sep','Okt','Nov','Des'][bln-1]} {thn}",'Penjualan':pen,'Iklan':ikl,'Laba':pen-ikl})
        
        df_tren = pd.DataFrame(tren_data)
        fig = go.Figure()
        fig.add_trace(go.Bar(name='💰 Penjualan', x=df_tren['Bulan'], y=df_tren['Penjualan'], marker_color='#10b981'))
        fig.add_trace(go.Bar(name='📢 Iklan', x=df_tren['Bulan'], y=df_tren['Iklan'], marker_color='#ef4444'))
        fig.add_trace(go.Scatter(name='💎 Laba', x=df_tren['Bulan'], y=df_tren['Laba'], line=dict(color='#0D9488', width=3)))
        fig.update_layout(height=350, margin=dict(l=10,r=10,t=10,b=10), legend=dict(orientation='h', y=1.1))
        st.plotly_chart(fig, use_container_width=True)
        
        # Rekap per toko
        st.markdown("---")
        st.subheader("🏪 Rekap per Toko")
        if not st.session_state.income_data.empty and 'Nama_Toko' in st.session_state.income_data.columns:
            df_toko = st.session_state.income_data.copy()
            df_toko['Tgl'] = pd.to_datetime(df_toko['Tanggal_Dana'], errors='coerce')
            tgl_a = datetime(tahun_pilih, bulan_pilih, 1)
            if bulan_pilih == 12:
                tgl_b = datetime(tahun_pilih+1, 1, 1) - timedelta(days=1)
            else:
                tgl_b = datetime(tahun_pilih, bulan_pilih+1, 1) - timedelta(days=1)
            df_toko = df_toko[(df_toko['Tgl'].dt.date >= tgl_a.date()) & (df_toko['Tgl'].dt.date <= tgl_b.date())]
            if not df_toko.empty:
                rt = df_toko.groupby('Nama_Toko')['Total_Dibayar'].sum().reset_index()
                rt.columns = ['Nama Toko','Total Diterima']
                rt = rt.sort_values('Total Diterima', ascending=False)
                rt['Total Diterima'] = rt['Total Diterima'].apply(format_rupiah)
                st.dataframe(rt, use_container_width=True, height=200, hide_index=True)
            else:
                st.caption("📭 Tidak ada data")
        else:
            st.caption("📭 Upload Income dulu")
    
    # ============================================================
    # TAB HUTANG SUPPLIER
    # ============================================================
    with tab_hutang:
        st.subheader("💳 Hutang Supplier")
        
        if not st.session_state.pembelian_data.empty:
            df_hutang = st.session_state.pembelian_data.copy()
            if 'Status' in df_hutang.columns:
                df_hutang = df_hutang[df_hutang['Status'].astype(str).str.strip().str.lower() == 'hutang']
            
            if df_hutang.empty:
                st.success("🎉 Tidak ada hutang!")
            else:
                ch1, ch2 = st.columns(2)
                with ch1:
                    fj = st.selectbox("Jatuh Tempo", ['Semua','Hari Ini','Minggu Ini','Bulan Ini','Lewat'], key="fj")
                with ch2:
                    fs = st.selectbox("Supplier", ['Semua']+sorted(df_hutang['Supplier'].dropna().unique().tolist()), key="fs")
                
                if fs != 'Semua':
                    df_hutang = df_hutang[df_hutang['Supplier']==fs]
                if fj != 'Semua' and 'Jatuh_Tempo' in df_hutang.columns:
                    df_hutang['JT'] = pd.to_datetime(df_hutang['Jatuh_Tempo'], errors='coerce')
                    tdy = datetime.now().date()
                    if fj == 'Hari Ini':
                        df_hutang = df_hutang[df_hutang['JT'].dt.date==tdy]
                    elif fj == 'Minggu Ini':
                        df_hutang = df_hutang[(df_hutang['JT'].dt.date>=tdy)&(df_hutang['JT'].dt.date<=tdy+timedelta(days=7))]
                    elif fj == 'Bulan Ini':
                        df_hutang = df_hutang[df_hutang['JT'].dt.month==tdy.month]
                    elif fj == 'Lewat':
                        df_hutang = df_hutang[df_hutang['JT'].dt.date<tdy]
                
                if not df_hutang.empty:
                    th = int(df_hutang['Total'].sum())
                    c1,c2,c3 = st.columns(3)
                    c1.metric("💳 Total Hutang", format_rupiah(th))
                    c2.metric("📋 Invoice", len(df_hutang['No_Invoice'].unique()))
                    c3.metric("🏪 Supplier", df_hutang['Supplier'].nunique())
                    
                    st.markdown("---")
                    df_hd = df_hutang[['Tanggal','No_Invoice','Supplier','SKU','Nama_Barang','Qty','Harga_Beli','Total','Metode','Jatuh_Tempo']].copy()
                    if 'Jatuh_Tempo' in df_hd.columns:
                        df_hd['Jatuh_Tempo'] = pd.to_datetime(df_hd['Jatuh_Tempo']).dt.strftime('%d/%m/%Y')
                    for c in ['Harga_Beli','Total']:
                        df_hd[c] = df_hd[c].apply(format_rupiah)
                    st.dataframe(df_hd.sort_values('Jatuh_Tempo'), use_container_width=True, height=350, hide_index=True)
                    
                    st.markdown("---")
                    st.subheader("✅ Tandai Lunas")
                    cl1,cl2,cl3 = st.columns(3)
                    with cl1:
                        inv_l = st.selectbox("No Invoice", sorted(df_hutang['No_Invoice'].unique()), key="inv_l")
                    with cl2:
                        tgl_l = st.date_input("Tgl Bayar", value=datetime.now(), key="tgl_l")
                    with cl3:
                        if st.button("✅ LUNAS", type="primary", use_container_width=True):
                            mask = st.session_state.pembelian_data['No_Invoice']==inv_l
                            st.session_state.pembelian_data.loc[mask,'Status']='Lunas'
                            st.session_state.pembelian_data.loc[mask,'Jatuh_Tempo']=tgl_l.strftime('%Y-%m-%d')
                            save_pembelian_data(st.session_state.pembelian_data)
                            st.success(f"✅ {inv_l} LUNAS!")
                            st.rerun()
                    
                    st.markdown("---")
                    st.subheader("📊 Per Supplier")
                    hs = df_hutang.groupby('Supplier')['Total'].sum().reset_index()
                    hs = hs.sort_values('Total', ascending=False)
                    hs['Total'] = hs['Total'].apply(format_rupiah)
                    st.dataframe(hs, use_container_width=True, height=200, hide_index=True)
                else:
                    st.info("📭 Tidak ada data")
        else:
            st.info("📭 Belum ada pembelian")

    
    # ============================================================
    # TAB REKONSILIASI SHOPEE
    # ============================================================
    with tab_rekon:
        st.subheader("🏦 Rekonsiliasi Saldo Shopee")
        st.caption("Catat saldo awal & penarikan untuk cocokkan dengan saldo Shopee")
        
        # Init database penarikan
        PENARIKAN_DB_FILE = str(DB_FOLDER / "penarikan.csv")
        if 'penarikan_data' not in st.session_state:
            if os.path.exists(PENARIKAN_DB_FILE):
                st.session_state.penarikan_data = pd.read_csv(PENARIKAN_DB_FILE)
            else:
                st.session_state.penarikan_data = pd.DataFrame(columns=['Tanggal','Jumlah','Keterangan'])
        
        # Pilih bulan
        bln_rekon = st.selectbox("Bulan", range(1,13), index=datetime.now().month-1,
                                 format_func=lambda x: ['Jan','Feb','Mar','Apr','Mei','Jun','Jul','Ags','Sep','Okt','Nov','Des'][x-1], key="bln_rekon")
        thn_rekon = st.selectbox("Tahun", range(2024,2031), index=datetime.now().year-2024, key="thn_rekon")
        
        tgl_a = datetime(thn_rekon, bln_rekon, 1)
        if bln_rekon == 12:
            tgl_b = datetime(thn_rekon+1, 1, 1) - timedelta(days=1)
        else:
            tgl_b = datetime(thn_rekon, bln_rekon+1, 1) - timedelta(days=1)
        
        # Saldo awal
        if f'saldo_awal_{thn_rekon}_{bln_rekon}' not in st.session_state:
            st.session_state[f'saldo_awal_{thn_rekon}_{bln_rekon}'] = 0
        
        st.markdown("---")
        col_sa, col_pen = st.columns(2)
        with col_sa:
            saldo_awal = st.number_input("💰 Saldo Awal Bulan (Rp)", min_value=0, step=100000,
                                         value=st.session_state[f'saldo_awal_{thn_rekon}_{bln_rekon}'],
                                         key=f"saldo_awal_{thn_rekon}_{bln_rekon}_input")
            if st.button("💾 Simpan Saldo Awal", key="simpan_saldo_awal", type="secondary"):
                st.session_state[f'saldo_awal_{thn_rekon}_{bln_rekon}'] = saldo_awal
                st.success("✅ Saldo awal disimpan!")
                st.rerun()
        
        # Penjualan Cair bulan ini
        penjualan_bulan = 0
        if not st.session_state.income_data.empty and 'Tanggal_Dana' in st.session_state.income_data.columns:
            df_inc_rekon = st.session_state.income_data.copy()
            df_inc_rekon['Tgl'] = pd.to_datetime(df_inc_rekon['Tanggal_Dana'], errors='coerce')
            mask = (df_inc_rekon['Tgl'].dt.date >= tgl_a.date()) & (df_inc_rekon['Tgl'].dt.date <= tgl_b.date())
            penjualan_bulan = int(df_inc_rekon[mask]['Total_Dibayar'].sum()) if 'Total_Dibayar' in df_inc_rekon.columns else 0
        
        with col_pen:
            st.metric("📥 Penjualan Masuk (Cair)", format_rupiah(penjualan_bulan))
            st.caption("Otomatis dari Income")
        
        # Input penarikan
        st.markdown("---")
        st.subheader("💸 Input Penarikan")
        with st.form("form_penarikan"):
            cp1, cp2, cp3 = st.columns(3)
            with cp1:
                tgl_tarik = st.date_input("Tanggal Tarik", value=datetime.now(), key="tgl_tarik")
            with cp2:
                jml_tarik = st.number_input("Jumlah (Rp)", min_value=0, step=100000, key="jml_tarik")
            with cp3:
                ket_tarik = st.text_input("Keterangan", placeholder="Ke rekening...", key="ket_tarik")
            
            if st.form_submit_button("💾 Catat Penarikan", type="primary", use_container_width=True):
                if jml_tarik > 0:
                    baru_tarik = pd.DataFrame([{
                        'Tanggal': tgl_tarik.strftime('%Y-%m-%d'),
                        'Jumlah': jml_tarik,
                        'Keterangan': ket_tarik
                    }])
                    st.session_state.penarikan_data = pd.concat([st.session_state.penarikan_data, baru_tarik], ignore_index=True)
                    st.session_state.penarikan_data.to_csv(PENARIKAN_DB_FILE, index=False)
                    st.success(f"✅ Penarikan {format_rupiah(jml_tarik)} dicatat!")
                    st.rerun()
                else:
                    st.error("Jumlah harus > 0")
        
        # Tabel penarikan
        st.markdown("---")
        st.subheader("📋 Riwayat Penarikan Bulan Ini")
        
        df_tarik_bulan = st.session_state.penarikan_data.copy()
        df_tarik_bulan['Tanggal'] = pd.to_datetime(df_tarik_bulan['Tanggal'], errors='coerce')
        mask = (df_tarik_bulan['Tanggal'].dt.date >= tgl_a.date()) & (df_tarik_bulan['Tanggal'].dt.date <= tgl_b.date())
        df_tarik_bulan = df_tarik_bulan[mask].sort_values('Tanggal', ascending=False)
        
        total_tarik = 0
        if not df_tarik_bulan.empty:
            total_tarik = int(df_tarik_bulan['Jumlah'].sum())
            df_tarik_bulan['Tanggal'] = df_tarik_bulan['Tanggal'].dt.strftime('%d/%m/%Y')
            df_tarik_display = df_tarik_bulan[['Tanggal','Jumlah','Keterangan']].copy()
            df_tarik_display['Jumlah'] = df_tarik_display['Jumlah'].apply(format_rupiah)
            st.dataframe(df_tarik_display, use_container_width=True, height=200, hide_index=True)
            
            # Hapus penarikan
            with st.expander("🗑️ Hapus Penarikan"):
                if len(df_tarik_bulan) > 0:
                    idx_h = st.number_input("Hapus baris ke", min_value=1, max_value=len(df_tarik_bulan), step=1, key="hapus_tarik")
                    if st.button("🗑️ Hapus", type="secondary", key="btn_hapus_tarik"):
                        st.session_state.penarikan_data = st.session_state.penarikan_data.drop(df_tarik_bulan.index[idx_h-1])
                        st.session_state.penarikan_data.to_csv(PENARIKAN_DB_FILE, index=False)
                        st.success("✅ Dihapus!")
                        st.rerun()
        else:
            st.caption("📭 Belum ada penarikan bulan ini")
        
        # Ringkasan
        st.markdown("---")
        st.subheader("📊 Ringkasan Saldo")
        
        saldo_akhir = saldo_awal + penjualan_bulan - total_tarik
        
        cr1, cr2, cr3, cr4 = st.columns(4)
        with cr1:
            st.metric("💰 Saldo Awal", format_rupiah(saldo_awal))
        with cr2:
            st.metric("📥 + Penjualan", format_rupiah(penjualan_bulan))
        with cr3:
            st.metric("💸 - Penarikan", format_rupiah(total_tarik))
        with cr4:
            warna_saldo = "#10b981" if saldo_akhir >= 0 else "#ef4444"
            st.markdown(f"""
            <div style="background:{warna_saldo}20;border:2px solid {warna_saldo};border-radius:12px;padding:16px;text-align:center;">
                <div style="font-size:11px;color:{warna_saldo};">🏦 SALDO AKHIR</div>
                <div style="font-size:18px;font-weight:800;color:{warna_saldo};">{format_rupiah(saldo_akhir)}</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.caption("ℹ️ Bandingkan Saldo Akhir dengan Saldo Penjual di Aplikasi Shopee")

# =====================================
# FOOTER
# =====================================
st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    if st.button("💾 SAVE ALL", use_container_width=True):
        st.success("✅ Tersimpan!")
with col2:
    if st.button("🔄 REFRESH", use_container_width=True):
        st.rerun()

st.caption(f"BSELLER ANALYTICS PRO v5.0 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
